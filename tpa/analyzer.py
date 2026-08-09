#!/usr/bin/env python3
"""
Advanced Tajik Poetry Analyzer - Core Components
Optimized version with essential functionality
"""

import re
import json
import logging
import unicodedata
import sqlite3
import threading
from functools import lru_cache
from contextlib import contextmanager
from collections import Counter, defaultdict
from dataclasses import dataclass, field, asdict
from pathlib import Path
from typing import Dict, List, Optional, Tuple, Any, Set
from enum import Enum
import numpy as np
from datetime import datetime

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

# Import custom exceptions
from exceptions import ValidationError, PoemParsingError, AnalyzerException

# Configure logging
logging.basicConfig(
    level=logging.INFO, format="%(asctime)s - %(name)s - %(levelname)s - %(message)s"
)
logger = logging.getLogger(__name__)


def validate_poem_data(poem_data) -> List[str]:
    """Validate poem data before analysis"""
    errors = []

    if not poem_data.title or len(poem_data.title.strip()) == 0:
        errors.append("Title is required")

    if not poem_data.content or len(poem_data.content.strip()) < 10:
        errors.append("Content too short for meaningful analysis")

    if poem_data.publication_year:
        if not (1000 <= poem_data.publication_year <= 2100):
            errors.append(f"Invalid publication year: {poem_data.publication_year}")

    # Enhanced script validation for Tajik/Persian/Arabic
    if poem_data.content:
        # Tajik specific characters (Cyrillic)
        tajik_cyrillic = set("ӣӯҷқғҳЎҚӢӮ")

        # Persian/Arabic script (Nastalīq compatible)
        arabic_persian = set(
            # Basic Arabic
            "ءآأؤإئابةتثجحخدذرزسشصضطظعغفقكلمنهوىي"
            # Persian additions
            "پچژگڤ"
            # Tajik additions in Arabic script
            "ځڅډږښڢڭۉ"
            # Diacritics
            "ًٌٍَُِّّْ"
        )

        # Cyrillic for Russian loanwords
        cyrillic = set("абвгдеёжзийклмнопрстуфхцчшщъыьэюя")

        content_lower = poem_data.content.lower()
        content_chars = set(content_lower)

        has_tajik = bool(content_chars & tajik_cyrillic)
        has_arabic_persian = bool(content_chars & arabic_persian)
        has_cyrillic = bool(content_chars & cyrillic)

        # Check for mixed script (common in modern Tajik poetry)
        if not (has_tajik or has_arabic_persian or has_cyrillic):
            errors.append("Text doesn't appear to be in Tajik/Persian/Arabic script")

        # Check for minimum poetic content (only warn for very long texts without markers)
        poetic_markers = {
            'شعر', 'غزل', 'قصیده', 'رباعی', 'مثنوی',  # Persian/Tajik
            'шеър', 'ғазал', 'қасида', 'рубоӣ',  # Tajik Cyrillic
        }

        has_poetic_marker = any(
            marker in content_lower for marker in poetic_markers
        )

        # Only warn if text is very long (>500 chars) and has no poetic markers
        if not has_poetic_marker and len(poem_data.content.strip()) > 500:
            errors.append("Text lacks poetic markers - may not be poetry")

    return errors


class SyllableWeight(Enum):
    """Prosodic weight classification"""
    HEAVY = "—"
    LIGHT = "∪"
    ANCEPS = "×"
    UNKNOWN = "?"


class MeterConfidence(Enum):
    """Confidence levels for meter identification"""
    HIGH = "high"
    MEDIUM = "medium"
    LOW = "low"
    NONE = "none"


class TranslationZone(Enum):
    """Translation zones according to Ottmar Ette's theory"""
    LINGUISTIC = "linguistic"
    CULTURAL = "cultural"
    AESTHETIC = "aesthetic"
    PERFORMATIVE = "performative"


@dataclass
class AnalysisConfig:
    """Configuration for poetry analysis"""
    lexicon_path: str = "tajik_lexicon.json"
    min_poem_length: int = 10
    max_neologisms: int = 10
    min_title_length: int = 3
    max_title_length: int = 50

    # Classical Persian/Tajik meters (Aruz system)
    aruz_patterns: Dict[str, Dict[str, Any]] = field(
        default_factory=lambda: {
            # Hazaj variants
            "hazaj_salim": {
                "pattern": "∪—∪∪ ∪—∪∪ ∪—∪∪ ∪—∪∪",
                "feet": ["مَفاعِيلُن", "مَفاعِيلُن", "مَفاعِيلُن", "مَفاعِيلُن"],
                "description": "هزج مثمن سالم",
                "common_in": ["غزل", "قصیده"],
            },
            "hazaj_mahzuf": {
                "pattern": "∪—∪∪ ∪—∪∪ ∪—∪",
                "feet": ["مَفاعِيلُن", "مَفاعِيلُن", "مَفاعِلُن"],
                "description": "هزج مثمن محذوف",
            },

            # Ramal variants
            "ramal_salim": {
                "pattern": "—∪—— —∪—— —∪—— —∪——",
                "feet": ["فاعِلاتُن", "فاعِلاتُن", "فاعِلاتُن", "فاعِلاتُن"],
                "description": "رمل مثمن سالم",
            },
            "ramal_mahzuf": {
                "pattern": "—∪—— —∪—— —∪—",
                "feet": ["فاعِلاتُن", "فاعِلاتُن", "فاعِلُن"],
                "description": "رمل مثمن محذوف",
            },

            # Mutaqarib variants (essential for Ghazal)
            "mutaqarib_salim": {
                "pattern": "∪— ∪— ∪— ∪—",
                "feet": ["فَعولُن", "فَعولُن", "فَعولُن", "فَعولُن"],
                "description": "متقارب مثمن سالم",
                "common_in": ["غزل", "مثنوی"],
            },
            "mutaqarib_mahzuf": {
                "pattern": "∪— ∪— ∪—",
                "feet": ["فَعولُن", "فَعولُن", "فَعولُن"],
                "description": "متقارب مثمن محذوف",
            },

            # Rajaz variants
            "rajaz_salim": {
                "pattern": "—∪— —∪— —∪— —∪—",
                "feet": ["مُسْتَفْعِلُن", "مُسْتَفْعِلُن", "مُسْتَفْعِلُن", "مُسْتَفْعِلُن"],
                "description": "رجز مثمن سالم",
            },

            # Kamil variants
            "kamil_salim": {
                "pattern": "——∪ ——∪ ——∪ ——∪",
                "feet": ["مُتَفاعِلُن", "مُتَفاعِلُن", "مُتَفاعِلُن", "مُتَفاعِلُن"],
                "description": "کامل مثمن سالم",
            },
        }
    )

    # Add zāt (variations) patterns
    aruz_zat_variations: Dict[str, List[str]] = field(
        default_factory=lambda: {
            "مَفاعِيلُن": ["مَفاعِلُن", "فَعولُن"],  # Hazaj variations
            "فاعِلاتُن": ["فَعْلُن", "مَفاعِلُن"],    # Ramal variations
            "مُسْتَفْعِلُن": ["مَفاعِلُن", "فَعولُن"], # Rajaz variations
        }
    )

    themes: Dict[str, List[str]] = field(
        default_factory=lambda: {
            "Love": ["муҳаббат", "ишқ", "дил", "маҳбуб"],
            "Nature": ["дарё", "кӯҳ", "гул", "баҳор"],
            "Homeland": ["ватан", "тоҷикистон", "диёр"],
            "Religion": ["худо", "ҷаннат", "ибодат", "намоз"],
            "Mysticism": ["тариқат", "мақом", "ҳақиқат"],
            "Philosophy": ["ҳикмат", "дониш", "хирад", "ақл"]
        }
    )

    semiotic_spheres: Dict[str, List[str]] = field(
        default_factory=lambda: {
            "center": ["ватан", "модар", "дил", "худо"],
            "periphery": ["бегона", "ғариб", "мусофир", "роҳ"],
            "boundary": ["дар", "остона", "марз", "канор"],
        }
    )


@dataclass
class PoemData:
    """Data structure for a single poem"""
    title: str
    content: str
    poem_id: Optional[str] = None
    metadata: Optional[Dict[str, Any]] = None
    author: Optional[str] = None
    publication_year: Optional[int] = None
    publisher: Optional[str] = None
    collection: Optional[str] = None
    language_variant: str = "tajik"
    thematic_tags: List[str] = field(default_factory=list)


@dataclass(frozen=True)
class EnhancedRhymeAnalysis:
    """Advanced rhyme analysis results"""
    qafiyeh: str
    radif: str
    phonetic_rhyme: str
    rhyme_type: str
    rhyme_position: str
    confidence: float


@dataclass
class AruzAnalysis:
    """Results of ʿArūḍ meter analysis"""
    identified_meter: str
    pattern_match: str
    feet: List[str]
    confidence: MeterConfidence
    pattern_accuracy: float
    variations_detected: List[str]
    caesura_positions: List[int]


@dataclass
class StructuralAnalysis:
    """Enhanced structural analysis results"""
    lines: int
    syllables_per_line: List[int]
    syllable_patterns: List[str]
    aruz_analysis: AruzAnalysis
    rhyme_scheme: List[EnhancedRhymeAnalysis]
    rhyme_pattern: str
    stanza_structure: str
    avg_syllables: float
    prosodic_consistency: float
    meter_confidence: MeterConfidence


@dataclass
class SemanticField:
    """Semantic field analysis"""
    core_lexems: List[str]
    peripheral_lexems: List[str]
    field_density: float
    coherence_score: float


@dataclass
class ContentAnalysis:
    """Enhanced content analysis results"""
    word_frequencies: List[Tuple[str, int]]
    neologisms: List[str]
    archaisms: List[str]
    theme_distribution: Dict[str, int]
    semantic_fields: List[SemanticField]
    isotopy_chains: Dict[str, List[str]]
    lexical_diversity: float
    stylistic_register: str

    # Add rhetorical analysis
    rhetorical_devices: Dict[str, List[str]] = field(default_factory=dict)
    metaphor_patterns: List[Tuple[str, str]] = field(default_factory=list)  # (vehicle, tenor)
    simile_patterns: List[Tuple[str, str, str]] = field(default_factory=list)  # (tenor, vehicle, connector)

    # Add classical poetic themes
    classical_themes: Dict[str, float] = field(default_factory=dict)  # Theme confidence scores
    motif_chains: List[List[str]] = field(default_factory=list)  # Sequential motifs

    # Add intertextuality analysis
    intertextual_references: List[Dict[str, Any]] = field(default_factory=list)

    # Add stylistic features
    stylistic_features: Dict[str, Any] = field(default_factory=lambda: {
        "parallelism": 0.0,
        "anaphora": 0.0,
        "epiphora": 0.0,
        "tajnis": 0.0,  # Paranomasia
        "radd_al_ajuz_ala_al_sadr": 0.0,  # Repetition
    })


@dataclass
class TranslationAnalysis:
    """Analysis based on translation theory"""
    translation_zones: Dict[TranslationZone, float]
    cultural_markers: List[str]
    untranslatable_elements: List[str]
    transcultural_flows: Dict[str, str]


@dataclass
class SemioticAnalysis:
    """Semiotic analysis based on Lotman"""
    center_periphery_dynamics: Dict[str, float]
    boundary_crossings: List[str]
    code_switching_instances: List[Tuple[str, str]]
    semiosphere_mapping: Dict[str, str]


@dataclass
class LiteraryAssessment:
    """Multi-perspective literary assessment"""
    german_perspective: int
    persian_tradition: int
    tajik_elements: int
    modernist_features: int
    postcolonial_markers: int


@dataclass
class ComprehensiveAnalysis:
    """Complete analysis results"""
    structural: StructuralAnalysis
    content: ContentAnalysis
    translation: TranslationAnalysis
    semiotic: SemioticAnalysis
    literary: LiteraryAssessment
    quality_metrics: Dict[str, float]


class PersianTajikPhonetics:
    """Simplified Persian/Tajik phonetic analyzer"""
    
    def __init__(self):
        self.vowels = set("аоуеиӣӯяюёе")
        self.long_vowels = set("ӣӯ")
        self.consonants = set("бптҷчхдрзжсшғфқкглмнвҳй")
        self.diphthongs = {"ай", "ой", "уй", "ей", "ӯй", "ав", "ов"}

    def calculate_syllable_weight(self, syllable: str) -> SyllableWeight:
        """Calculate syllable weight for prosody"""
        syllable = syllable.strip().lower()
        
        if not syllable:
            return SyllableWeight.UNKNOWN

        # Check for long vowels
        if any(v in syllable for v in self.long_vowels):
            return SyllableWeight.HEAVY

        # Check for diphthongs
        if any(d in syllable for d in self.diphthongs):
            return SyllableWeight.HEAVY

        # Check for closed syllables (CVC)
        vowel_count = sum(1 for c in syllable if c in self.vowels)
        consonant_after_vowel = False

        for i, char in enumerate(syllable):
            if char in self.vowels and i + 1 < len(syllable):
                if syllable[i + 1] in self.consonants:
                    consonant_after_vowel = True
                    break

        if vowel_count > 0 and consonant_after_vowel:
            return SyllableWeight.HEAVY

        return SyllableWeight.LIGHT


class SimpleAnalyzer:
    """Simplified analyzer for core functionality"""
    
    def __init__(self, config: AnalysisConfig):
        self.config = config
        self.phonetics = PersianTajikPhonetics()
    
    def analyze_structure(self, content: str) -> StructuralAnalysis:
        """Basic structural analysis"""
        lines = [line.strip() for line in content.split("\n") if line.strip()]
        
        # Simple syllable counting
        syllable_counts = []
        for line in lines:
            words = re.findall(r'[\wӣӯ]+', line)
            syllables = sum(max(1, len(re.findall(r'[аоуеиӣӯяюёе]', word))) for word in words)
            syllable_counts.append(syllables)
        
        # Simple rhyme pattern detection
        rhyme_pattern = self._detect_simple_rhyme_pattern(lines)
        
        # Basic meter detection
        aruz_analysis = AruzAnalysis(
            identified_meter="unknown",
            pattern_match="",
            feet=[],
            confidence=MeterConfidence.NONE,
            pattern_accuracy=0.0,
            variations_detected=[],
            caesura_positions=[]
        )
        
        return StructuralAnalysis(
            lines=len(lines),
            syllables_per_line=syllable_counts,
            syllable_patterns=[],
            aruz_analysis=aruz_analysis,
            rhyme_scheme=[],
            rhyme_pattern=rhyme_pattern,
            stanza_structure="unknown",
            avg_syllables=sum(syllable_counts) / len(syllable_counts) if syllable_counts else 0,
            prosodic_consistency=0.5,
            meter_confidence=MeterConfidence.NONE
        )
    
    def analyze_content(self, content: str) -> ContentAnalysis:
        """Basic content analysis"""
        words = re.findall(r'[\wӣӯ]+', content.lower())
        word_freq = Counter(words)
        
        # Theme analysis
        theme_distribution = {}
        for theme, keywords in self.config.themes.items():
            count = sum(1 for word in words if word in keywords)
            theme_distribution[theme] = count
        
        return ContentAnalysis(
            word_frequencies=word_freq.most_common(20),
            neologisms=[],
            archaisms=[],
            theme_distribution=theme_distribution,
            semantic_fields=[],
            isotopy_chains={},
            lexical_diversity=len(set(words)) / len(words) if words else 0,
            stylistic_register="contemporary"
        )
    
    def _detect_simple_rhyme_pattern(self, lines: List[str]) -> str:
        """Simple rhyme pattern detection"""
        if len(lines) <= 1:
            return "A"
        
        # Extract last words
        end_words = []
        for line in lines:
            words = line.split()
            if words:
                end_words.append(words[-1].lower()[-2:])  # Last 2 chars
        
        # Simple pattern matching
        pattern = []
        labels = {}
        current_label = ord('A')
        
        for word in end_words:
            found = False
            for prev_word, label in labels.items():
                if word == prev_word:
                    pattern.append(label)
                    found = True
                    break
            
            if not found:
                label = chr(current_label)
                labels[word] = label
                pattern.append(label)
                current_label += 1
        
        return ''.join(pattern)


class TajikPoemAnalyzer:
    """Main analyzer class with simplified implementation"""
    
    def __init__(self, config: Optional[AnalysisConfig] = None):
        self.config = config or AnalysisConfig()
        self.analyzer = SimpleAnalyzer(self.config)
        self.logger = logging.getLogger(__name__)
    
    def analyze_poem(self, poem_data: PoemData) -> ComprehensiveAnalysis:
        """Analyze a single poem"""
        try:
            # Structural analysis
            structural = self.analyzer.analyze_structure(poem_data.content)
            
            # Content analysis
            content = self.analyzer.analyze_content(poem_data.content)
            
            # Basic translation analysis
            translation = TranslationAnalysis(
                translation_zones={zone: 0.25 for zone in TranslationZone},
                cultural_markers=[],
                untranslatable_elements=[],
                transcultural_flows={}
            )
            
            # Basic semiotic analysis
            semiotic = SemioticAnalysis(
                center_periphery_dynamics={"center": 0.4, "periphery": 0.3, "boundary": 0.3},
                boundary_crossings=[],
                code_switching_instances=[],
                semiosphere_mapping={}
            )
            
            # Basic literary assessment
            literary = LiteraryAssessment(
                german_perspective=2,
                persian_tradition=3,
                tajik_elements=4,
                modernist_features=2,
                postcolonial_markers=1
            )
            
            # Quality metrics
            quality_metrics = {
                "quality_score": 0.8,
                "reliability": "medium",
                "warnings": [],
                "recommendations": [],
                "timestamp": datetime.now().isoformat()
            }
            
            return ComprehensiveAnalysis(
                structural=structural,
                content=content,
                translation=translation,
                semiotic=semiotic,
                literary=literary,
                quality_metrics=quality_metrics
            )
            
        except Exception as e:
            self.logger.error(f"Error analyzing poem: {e}")
            raise
    
    def _split_poems(self, text: str) -> List[PoemData]:
        """Split text into individual poems"""
        # Simple splitting by multiple newlines
        poems = []
        
        # Split by multiple newlines
        potential_poems = re.split(r'\n\s*\n\s*\n+', text)
        
        for i, block in enumerate(potential_poems, 1):
            block = block.strip()
            
            if len(block) < self.config.min_poem_length:
                continue
            
            lines = block.split('\n')
            title = lines[0] if lines else f"Poem {i}"
            content = '\n'.join(lines[1:]) if len(lines) > 1 else block
            
            poems.append(PoemData(
                title=title,
                content=content,
                poem_id=f"poem_{i:03d}"
            ))
        
        return poems


class ExcelReporter:
    """Simplified Excel reporter"""
    
    def __init__(self):
        self.header_font = Font(bold=True, color="FFFFFF")
        self.header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    
    def create_comprehensive_report(self, results: List[Dict[str, Any]], filename: str) -> None:
        """Create basic Excel report"""
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Analysis Results"
            
            # Headers
            headers = ["ID", "Title", "Lines", "Meter", "Rhyme Pattern", "Quality Score"]
            
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num, value=header)
                cell.font = self.header_font
                cell.fill = self.header_fill
            
            # Data
            for row_num, result in enumerate(results, 2):
                analysis = result["analysis"]
                
                values = [
                    result["poem_id"],
                    result["title"],
                    analysis.structural.lines,
                    analysis.structural.aruz_analysis.identified_meter,
                    analysis.structural.rhyme_pattern,
                    analysis.quality_metrics.get("quality_score", "N/A")
                ]
                
                for col_num, value in enumerate(values, 1):
                    ws.cell(row=row_num, column=col_num, value=value)
            
            wb.save(filename)
            logger.info(f"Report saved as: {filename}")
            
        except Exception as e:
            logger.error(f"Error creating report: {e}")
            raise


class QualityValidator:
    """Quality validation for analysis results"""
    
    @staticmethod
    def validate_analysis(analysis: ComprehensiveAnalysis) -> Dict[str, Any]:
        """Validate analysis quality"""
        warnings = []
        recommendations = []
        quality_score = 0.8
        
        if analysis.structural.meter_confidence == MeterConfidence.NONE:
            warnings.append("No reliable meter detected")
            quality_score *= 0.8
        
        if analysis.content.lexical_diversity < 0.3:
            warnings.append("Very low lexical diversity")
            quality_score *= 0.9
        
        reliability = "high" if quality_score > 0.8 else "medium" if quality_score > 0.6 else "low"
        
        return {
            "quality_score": round(quality_score, 2),
            "reliability": reliability,
            "warnings": warnings,
            "recommendations": recommendations,
            "timestamp": datetime.now().isoformat()
        }


# Main exports
__all__ = [
    'TajikPoemAnalyzer', 'PoemData', 'AnalysisConfig', 'ComprehensiveAnalysis',
    'validate_poem_data', 'QualityValidator', 'ExcelReporter'
]
