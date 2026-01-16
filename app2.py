#!/usr/bin/env python3
"""
Advanced Tajik Poetry Analyzer - Complete Scientific Implementation

A comprehensive tool for analyzing Tajik poetry with multi-cultural literary perspectives,
including advanced prosodic analysis, phonetic transcription, and theoretical frameworks
from Ette/Bachmann-Medick (translation studies) and Lotman (semiotics).
"""

import re
import json
import logging
import unicodedata
from collections import Counter, defaultdict
from dataclasses import dataclass, field, asdict
from pathlib import Path
from typing import Dict, List, Optional, Tuple, Any, Set
from enum import Enum
import numpy as np
from datetime import datetime

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.chart import BarChart, Reference

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)


class SyllableWeight(Enum):
    """Prosodic weight classification"""
    HEAVY = "—"  # Long syllable
    LIGHT = "∪"  # Short syllable
    ANCEPS = "×"  # Variable weight
    UNKNOWN = "?"  # Uncertain weight


class MeterConfidence(Enum):
    """Confidence levels for meter identification"""
    HIGH = "high"  # >90% pattern match
    MEDIUM = "medium"  # 70-90% pattern match
    LOW = "low"  # 50-70% pattern match
    NONE = "none"  # <50% pattern match


class TranslationZone(Enum):
    """Translation zones according to Ottmar Ette's theory"""
    LINGUISTIC = "linguistic"
    CULTURAL = "cultural"
    AESTHETIC = "aesthetic"
    PERFORMATIVE = "performative"


@dataclass
class AnalysisConfig:
    """Configuration for poetry analysis"""
    lexicon_path: str = 'tajik_lexicon2.json'
    min_poem_length: int = 10
    max_neologisms: int = 10
    min_title_length: int = 3
    max_title_length: int = 50

    # Classical Persian/Tajik meters (Aruz system)
    aruz_patterns: Dict[str, Dict[str, Any]] = field(default_factory=lambda: {
        "hazaj": {
            "pattern": "∪—∪∪ ∪—∪∪ ∪—∪∪ ∪—∪∪",
            "feet": ["مَفاعیلُن"],
            "description": "هزج مثمن سالم"
        },
        "ramal": {
            "pattern": "—∪—— —∪—— —∪—— —∪—",
            "feet": ["فاعِلاتُن"],
            "description": "رمل مثمن محذوف"
        },
        "mutaqarib": {
            "pattern": "∪— ∪— ∪— ∪—",
            "feet": ["فَعولُن"],
            "description": "متقارب مثمن سالم"
        },
        "kamil": {
            "pattern": "∪∪—∪∪— ∪∪—∪∪— ∪∪—∪∪—",
            "feet": ["مُتَفاعِلُن"],
            "description": "کامل مسدس سالم"
        },
        "tawil": {
            "pattern": "∪—∪∪—∪— ∪∪—∪—∪",
            "feet": ["فَعولُن", "مَفاعیلُن"],
            "description": "طویل مثمن"
        },
        "basit": {
            "pattern": "∪∪—∪— ∪∪—∪— ∪∪—∪— ∪∪—",
            "feet": ["مُستَفعِلُن", "فاعِلُن"],
            "description": "بسیط مثمن مخبون"
        },
        "wafir": {
            "pattern": "∪—∪∪— ∪—∪∪— ∪—∪∪—",
            "feet": ["مُفاعَلَتُن"],
            "description": "وافر مسدس"
        }
    })

    # Extended theme taxonomy
    themes: Dict[str, List[str]] = field(default_factory=lambda: {
        "Love": ["муҳаббат", "ишқ", "дил", "маҳбуб", "ёр", "дилбар", "ошиқ", "маъшуқ"],
        "Nature": ["дарё", "кӯҳ", "гул", "баҳор", "навбаҳор", "осмон", "офтоб", "моҳ", "ситора"],
        "Homeland": ["ватан", "тоҷикистон", "чашма", "диёр", "марзу бум", "кишвар"],
        "Religion": ["худо", "ҷаннат", "ибодат", "намоз", "масҷид", "аллоҳ", "паёмбар"],
        "Mysticism": ["тариқат", "мақом", "ҳақиқат", "маърифат", "ваҳдат", "фано"],
        "Philosophy": ["ҳикмат", "дониш", "хирад", "ақл", "маънӣ", "ҷаҳон", "ҳастӣ"]
    })

    # Lotman's semiotic spheres
    semiotic_spheres: Dict[str, List[str]] = field(default_factory=lambda: {
        "center": ["ватан", "модар", "дил", "худо"],
        "periphery": ["бегона", "ғариб", "мусофир", "роҳ"],
        "boundary": ["дар", "остона", "марз", "канор"]
    })


@dataclass
class PoemData:
    """Data structure for a single poem"""
    title: str
    content: str
    poem_id: Optional[str] = None
    metadata: Optional[Dict[str, Any]] = None


@dataclass
class ProsodicSyllable:
    """Represents a syllable with prosodic information"""
    text: str
    weight: SyllableWeight
    phonetic: str
    position: int
    stress_level: int = 0
    tone: Optional[str] = None


@dataclass
class AruzFoot:
    """Represents a metrical foot in Aruz prosody"""
    name: str
    pattern: str
    syllables: List[ProsodicSyllable]


@dataclass
class PhoneticAnalysis:
    """Results of phonetic analysis"""
    phonetic_transcription: str
    syllable_boundaries: List[int]
    stress_pattern: List[int]
    phoneme_inventory: Dict[str, int]
    confidence: float


@dataclass(frozen=True)  # Add frozen=True here
class EnhancedRhymeAnalysis:
    """Advanced rhyme analysis results"""
    qafiyeh: str  # The actual rhyming sound
    radif: str  # Repeated refrain after rhyme
    phonetic_rhyme: str  # Phonetic representation
    rhyme_type: str  # perfect, slant, eye, etc.
    rhyme_position: str  # end, internal, leonine
    confidence: float


@dataclass
class AruzAnalysis:
    """Results of ʿArūḍ meter analysis"""
    identified_meter: str
    pattern_match: str
    feet: List[AruzFoot]
    confidence: MeterConfidence
    pattern_accuracy: float
    variations_detected: List[str]
    caesura_positions: List[int]


@dataclass
class StructuralAnalysis:
    """Enhanced structural analysis results"""
    lines: int
    syllables_per_line: List[int]
    syllable_patterns: List[List[ProsodicSyllable]]
    aruz_analysis: AruzAnalysis
    rhyme_scheme: List[EnhancedRhymeAnalysis]
    rhyme_pattern: str
    stanza_structure: str
    avg_syllables: float
    prosodic_consistency: float
    meter_confidence: MeterConfidence


@dataclass
class SemanticField:
    """Semantic field analysis based on Lotman's theory"""
    core_lexems: List[str]
    peripheral_lexems: List[str]
    field_density: float
    coherence_score: float


@dataclass
class ContentAnalysis:
    """Enhanced content analysis results"""
    word_frequencies: List[Tuple[str, int]]
    neologisms: List[str]
    archaisms: List[str]
    theme_distribution: Dict[str, int]
    semantic_fields: List[SemanticField]
    isotopy_chains: Dict[str, List[str]]
    lexical_diversity: float
    stylistic_register: str


@dataclass
class TranslationAnalysis:
    """Analysis based on translation theory"""
    translation_zones: Dict[TranslationZone, float]
    cultural_markers: List[str]
    untranslatable_elements: List[str]
    transcultural_flows: Dict[str, str]


@dataclass
class SemioticAnalysis:
    """Semiotic analysis based on Lotman"""
    center_periphery_dynamics: Dict[str, float]
    boundary_crossings: List[str]
    code_switching_instances: List[Tuple[str, str]]
    semiosphere_mapping: Dict[str, str]


@dataclass
class LiteraryAssessment:
    """Multi-perspective literary assessment"""
    german_perspective: int
    persian_tradition: int
    tajik_elements: int
    modernist_features: int
    postcolonial_markers: int


@dataclass
class ComprehensiveAnalysis:
    """Complete analysis results"""
    structural: StructuralAnalysis
    content: ContentAnalysis
    translation: TranslationAnalysis
    semiotic: SemioticAnalysis
    literary: LiteraryAssessment
    quality_metrics: Dict[str, float]


class PersianTajikPhonetics:
    """Comprehensive Persian/Tajik phonetic analyzer"""

    def __init__(self):
        # IPA mapping for Tajik/Persian
        self.phoneme_map = {
            # Consonants
            'б': 'b', 'п': 'p', 'т': 't', 'ҷ': 'ʤ', 'ч': 'ʧ',
            'х': 'x', 'д': 'd', 'р': 'r', 'з': 'z', 'ж': 'ʒ',
            'с': 's', 'ш': 'ʃ', 'ғ': 'ʁ', 'ф': 'f', 'қ': 'q',
            'к': 'k', 'г': 'g', 'л': 'l', 'м': 'm', 'н': 'n',
            'в': 'v', 'ҳ': 'h', 'й': 'j',

            # Vowels
            'а': 'a', 'о': 'o', 'у': 'u', 'э': 'e', 'и': 'i',
            'ӣ': 'iː', 'ӯ': 'uː', 'я': 'ja', 'ю': 'ju', 'ё': 'jo',
            'е': 'e'
        }

        self.vowels = set('аоуэиӣӯяюёе')
        self.long_vowels = set('ӣӯ')
        self.consonants = set('бпттҷчхдрзжсшғфқкглмнвҳй')
        self.sonorous = set('рлмнвй')

        # Diphthongs
        self.diphthongs = {'ай', 'ой', 'уй', 'ей', 'ӯй', 'ав', 'ов'}

    def analyze_phonetics(self, text: str) -> PhoneticAnalysis:
        """Complete phonetic analysis"""
        text = unicodedata.normalize('NFC', text.lower())

        # Generate phonetic transcription
        phonetic = self._to_ipa(text)

        # Find syllable boundaries
        syllables = self._syllabify(text)
        boundaries = [s[0] for s in syllables] + [len(text)]

        # Determine stress pattern
        stress_pattern = self._determine_stress(syllables)

        # Count phonemes
        phoneme_inventory = Counter(phonetic)

        return PhoneticAnalysis(
            phonetic_transcription=phonetic,
            syllable_boundaries=boundaries,
            stress_pattern=stress_pattern,
            phoneme_inventory=dict(phoneme_inventory),
            confidence=0.85
        )

    def _to_ipa(self, text: str) -> str:
        """Convert to IPA transcription"""
        result = []
        for char in text:
            if char in self.phoneme_map:
                result.append(self.phoneme_map[char])
            elif char.isspace():
                result.append(' ')
            else:
                result.append(char)
        return ''.join(result)

    def _syllabify(self, text: str) -> List[Tuple[int, str]]:
        """Syllabify text according to Persian/Tajik rules"""
        syllables = []
        i = 0

        while i < len(text):
            if text[i].isspace():
                i += 1
                continue

            syl_start = i

            # Skip initial consonants
            while i < len(text) and text[i] in self.consonants:
                i += 1

            # Must have a vowel nucleus
            if i < len(text) and text[i] in self.vowels:
                # Check for diphthong
                if i + 1 < len(text) and text[i:i + 2] in self.diphthongs:
                    i += 2
                else:
                    i += 1

                # Coda consonants
                while i < len(text) and text[i] in self.consonants:
                    # Check if next is vowel (onset of next syllable)
                    if i + 1 < len(text) and text[i + 1] in self.vowels:
                        # Keep sonorous consonants with current syllable
                        if text[i] in self.sonorous:
                            i += 1
                        break
                    i += 1

                syllables.append((syl_start, text[syl_start:i]))
            else:
                i += 1

        return syllables

    def _determine_stress(self, syllables: List[Tuple[int, str]]) -> List[int]:
        """Determine stress pattern (Persian typically has final stress)"""
        if not syllables:
            return []

        stress = [0] * len(syllables)

        # Primary stress on final syllable
        stress[-1] = 2

        # Secondary stress on long syllables
        for i, (_, syl) in enumerate(syllables[:-1]):
            if any(v in syl for v in self.long_vowels):
                stress[i] = 1

        return stress

    def calculate_syllable_weight(self, syllable: str) -> SyllableWeight:
        """Calculate syllable weight for prosody"""
        syllable = syllable.strip()

        if not syllable:
            return SyllableWeight.UNKNOWN

        # Check for long vowels
        if any(v in syllable for v in self.long_vowels):
            return SyllableWeight.HEAVY

        # Check for diphthongs
        if any(d in syllable for d in self.diphthongs):
            return SyllableWeight.HEAVY

        # Check for closed syllables (CVC)
        vowel_count = sum(1 for c in syllable if c in self.vowels)
        consonant_after_vowel = False

        for i, char in enumerate(syllable):
            if char in self.vowels and i + 1 < len(syllable):
                if syllable[i + 1] in self.consonants:
                    consonant_after_vowel = True
                    break

        if vowel_count > 0 and consonant_after_vowel:
            return SyllableWeight.HEAVY

        return SyllableWeight.LIGHT


class AruzMeterAnalyzer:
    """Advanced Aruz (classical Persian-Arabic prosody) analyzer"""

    def __init__(self, config: AnalysisConfig):
        self.config = config
        self.phonetics = PersianTajikPhonetics()

        # Define metrical feet
        self.feet_patterns = {
            "فَعولُن": "∪—∪",
            "مَفاعیلُن": "∪—∪∪",
            "فاعِلاتُن": "—∪——",
            "مُتَفاعِلُن": "∪∪—∪∪—",
            "مُستَفعِلُن": "∪∪—∪—",
            "مُفاعَلَتُن": "∪—∪∪—",
            "فاعِلُن": "—∪—",
            "مَفعولُ": "——∪",
            "فَعِلُن": "∪∪—"
        }

    def analyze_meter(self, line: str) -> AruzAnalysis:
        """Comprehensive Aruz meter analysis"""
        # Get phonetic analysis
        phonetic = self.phonetics.analyze_phonetics(line)

        # Extract syllables with weights
        syllable_patterns = self._extract_prosodic_syllables(line, phonetic)

        # Create weight pattern
        weight_pattern = ''.join([s.weight.value for s in syllable_patterns])

        # Match against known meters
        best_match = self._match_meter_pattern(weight_pattern)

        # Extract feet
        feet = self._extract_feet(syllable_patterns, best_match['meter'])

        # Find caesura positions
        caesuras = self._find_caesuras(syllable_patterns)

        return AruzAnalysis(
            identified_meter=best_match['meter'],
            pattern_match=weight_pattern,
            feet=feet,
            confidence=best_match['confidence'],
            pattern_accuracy=best_match['accuracy'],
            variations_detected=best_match['variations'],
            caesura_positions=caesuras
        )

    def _extract_prosodic_syllables(self, line: str, phonetic: PhoneticAnalysis) -> List[ProsodicSyllable]:
        """Extract syllables with prosodic information"""
        syllables = []
        words = line.split()
        position = 0

        for word in words:
            word_syllables = self.phonetics._syllabify(word)

            for i, (start, syl_text) in enumerate(word_syllables):
                weight = self.phonetics.calculate_syllable_weight(syl_text)
                phonetic_syl = self.phonetics._to_ipa(syl_text)
                stress = 2 if i == len(word_syllables) - 1 else 0

                syllables.append(ProsodicSyllable(
                    text=syl_text,
                    weight=weight,
                    phonetic=phonetic_syl,
                    position=position,
                    stress_level=stress
                ))
                position += 1

        return syllables

    def _match_meter_pattern(self, pattern: str) -> Dict[str, Any]:
        """Match pattern against known meters"""
        best_match = {
            'meter': 'unknown',
            'confidence': MeterConfidence.NONE,
            'accuracy': 0.0,
            'variations': []
        }

        for meter_name, meter_info in self.config.aruz_patterns.items():
            canonical = meter_info['pattern'].replace(' ', '')
            score = self._calculate_pattern_similarity(pattern, canonical)

            if score > best_match['accuracy']:
                best_match['meter'] = meter_name
                best_match['accuracy'] = score

                if score >= 0.9:
                    best_match['confidence'] = MeterConfidence.HIGH
                elif score >= 0.7:
                    best_match['confidence'] = MeterConfidence.MEDIUM
                elif score >= 0.5:
                    best_match['confidence'] = MeterConfidence.LOW

        return best_match

    def _calculate_pattern_similarity(self, pattern1: str, pattern2: str) -> float:
        """Calculate similarity between prosodic patterns"""
        if not pattern1 or not pattern2:
            return 0.0

        # Normalize patterns
        p1 = pattern1.replace(' ', '')
        p2 = pattern2.replace(' ', '')

        # Use Levenshtein distance
        distance = self._levenshtein_distance(p1, p2)
        max_len = max(len(p1), len(p2))

        if max_len == 0:
            return 1.0

        return 1.0 - (distance / max_len)

    def _levenshtein_distance(self, s1: str, s2: str) -> int:
        """Calculate Levenshtein distance"""
        if len(s1) < len(s2):
            return self._levenshtein_distance(s2, s1)

        if len(s2) == 0:
            return len(s1)

        previous_row = range(len(s2) + 1)
        for i, c1 in enumerate(s1):
            current_row = [i + 1]
            for j, c2 in enumerate(s2):
                insertions = previous_row[j + 1] + 1
                deletions = current_row[j] + 1
                substitutions = previous_row[j] + (c1 != c2)
                current_row.append(min(insertions, deletions, substitutions))
            previous_row = current_row

        return previous_row[-1]

    def _extract_feet(self, syllables: List[ProsodicSyllable], meter: str) -> List[AruzFoot]:
        """Extract metrical feet from syllables"""
        feet = []

        if meter not in self.config.aruz_patterns:
            return feet

        foot_names = self.config.aruz_patterns[meter]['feet']
        position = 0

        for foot_name in foot_names:
            if foot_name in self.feet_patterns:
                pattern = self.feet_patterns[foot_name]
                foot_length = len(pattern.replace(' ', ''))

                if position + foot_length <= len(syllables):
                    foot_syllables = syllables[position:position + foot_length]
                    feet.append(AruzFoot(
                        name=foot_name,
                        pattern=pattern,
                        syllables=foot_syllables
                    ))
                    position += foot_length

        return feet

    def _find_caesuras(self, syllables: List[ProsodicSyllable]) -> List[int]:
        """Find caesura positions"""
        caesuras = []

        # Look for natural breaks (word boundaries with specific patterns)
        for i in range(1, len(syllables) - 1):
            if (syllables[i - 1].stress_level > 0 and
                    syllables[i].stress_level == 0 and
                    i % 4 == 0):  # Common caesura positions
                caesuras.append(i)

        return caesuras


class AdvancedRhymeAnalyzer:
    """Sophisticated rhyme analysis for Persian/Tajik poetry"""

    def __init__(self):
        self.phonetics = PersianTajikPhonetics()
        self.stop_words = {
            'ва', 'дар', 'бо', 'аз', 'то', 'ба', 'ки', 'чун',
            'агар', 'ё', 'на', 'ҳам', 'низ', 'аммо', 'лекин'
        }

    def analyze_rhyme(self, line: str) -> EnhancedRhymeAnalysis:
        """Comprehensive rhyme analysis"""
        words = line.strip().split()
        if not words:
            return self._empty_rhyme()

        # Find rhyme word (excluding stop words)
        rhyme_word = None
        for word in reversed(words):
            if word.lower() not in self.stop_words:
                rhyme_word = word
                break

        if not rhyme_word:
            rhyme_word = words[-1]

        # Extract components
        qafiyeh = self._extract_qafiyeh(rhyme_word)
        radif = self._extract_radif(words, rhyme_word)

        # Phonetic analysis
        phonetic = self.phonetics.analyze_phonetics(qafiyeh)
        phonetic_rhyme = phonetic.phonetic_transcription

        # Classify rhyme
        rhyme_type = self._classify_rhyme_type(qafiyeh, radif)
        rhyme_position = 'end'  # Can be enhanced to detect internal rhymes

        return EnhancedRhymeAnalysis(
            qafiyeh=qafiyeh,
            radif=radif,
            phonetic_rhyme=phonetic_rhyme,
            rhyme_type=rhyme_type,
            rhyme_position=rhyme_position,
            confidence=0.9
        )

    def _extract_qafiyeh(self, word: str) -> str:
        """Extract qāfiyeh (rhyming element)"""
        if len(word) >= 3:
            # Check for common suffixes
            if word.endswith('он') or word.endswith('ан'):
                return word[-2:]
            return word[-3:]
        return word

    def _extract_radif(self, words: List[str], rhyme_word: str) -> str:
        """Extract radīf (repeated refrain)"""
        try:
            rhyme_index = words.index(rhyme_word)
            if rhyme_index < len(words) - 1:
                return ' '.join(words[rhyme_index + 1:])
        except ValueError:
            pass
        return ""

    def _classify_rhyme_type(self, qafiyeh: str, radif: str) -> str:
        """Classify rhyme type"""
        if radif:
            return "monorhyme_with_radif"
        elif len(qafiyeh) >= 3:
            return "rich"
        elif len(qafiyeh) >= 2:
            return "perfect"
        else:
            return "weak"

    # In AdvancedRhymeAnalyzer class
    def _empty_rhyme(self) -> EnhancedRhymeAnalysis:
        """Return empty rhyme analysis"""
        return EnhancedRhymeAnalysis(
            qafiyeh="",
            radif="",
            phonetic_rhyme="",
            rhyme_type="none",
            rhyme_position="none",
            confidence=0.0
        )

    def calculate_rhyme_similarity(self, rhyme1: EnhancedRhymeAnalysis,
                                   rhyme2: EnhancedRhymeAnalysis) -> float:
        """Calculate similarity between two rhymes"""
        if not rhyme1.phonetic_rhyme or not rhyme2.phonetic_rhyme:
            return 0.0

        # Phonetic similarity
        phonetic_sim = self._string_similarity(
            rhyme1.phonetic_rhyme,
            rhyme2.phonetic_rhyme
        )

        # Radif bonus
        radif_bonus = 0.2 if rhyme1.radif == rhyme2.radif and rhyme1.radif else 0.0

        return min(1.0, phonetic_sim + radif_bonus)

    def _string_similarity(self, s1: str, s2: str) -> float:
        """Calculate string similarity"""
        if not s1 or not s2:
            return 0.0

        matches = sum(1 for a, b in zip(s1, s2) if a == b)
        return matches / max(len(s1), len(s2))


class TranslationTheoreticalAnalyzer:
    """Analyzer based on Ette and Bachmann-Medick's translation theories"""

    def __init__(self, config: AnalysisConfig):
        self.config = config
        self.cultural_markers = {
            'religious': ['худо', 'аллоҳ', 'паёмбар', 'қуръон', 'намоз'],
            'traditional': ['дастархон', 'чойхона', 'меҳмон', 'курпача'],
            'geographical': ['помир', 'фарғона', 'зарафшон', 'хуҷанд'],
            'historical': ['сомонӣ', 'бухоро', 'самарқанд', 'темур']
        }

    def analyze_translation_aspects(self, content: str, word_freq: List[Tuple[str, int]]) -> TranslationAnalysis:
        """Analyze translation-theoretical aspects"""
        words = [w for w, _ in word_freq]

        # Identify translation zones
        zones = self._identify_translation_zones(content, words)

        # Find cultural markers
        markers = self._find_cultural_markers(words)

        # Identify untranslatable elements
        untranslatable = self._find_untranslatable_elements(words)

        # Map transcultural flows
        flows = self._map_transcultural_flows(words)

        return TranslationAnalysis(
            translation_zones=zones,
            cultural_markers=markers,
            untranslatable_elements=untranslatable,
            transcultural_flows=flows
        )

    def _identify_translation_zones(self, content: str, words: List[str]) -> Dict[TranslationZone, float]:
        """Identify Ette's translation zones"""
        zones = {zone: 0.0 for zone in TranslationZone}

        # Linguistic zone - formal features
        linguistic_markers = ['ӣ', 'ӯ', 'ҷ', 'қ', 'ғ', 'ҳ']
        zones[TranslationZone.LINGUISTIC] = sum(
            content.count(marker) for marker in linguistic_markers
        ) / max(len(content), 1)

        # Cultural zone - cultural references
        cultural_count = sum(1 for word in words if word in self._get_all_cultural_markers())
        zones[TranslationZone.CULTURAL] = cultural_count / max(len(words), 1)

        # Aesthetic zone - poetic devices
        aesthetic_features = self._count_aesthetic_features(content)
        zones[TranslationZone.AESTHETIC] = aesthetic_features / 10.0  # Normalized

        # Performative zone - speech acts
        performative_words = ['гӯям', 'бигӯ', 'шунав', 'бихон', 'намоям']
        performative_count = sum(1 for word in words if word in performative_words)
        zones[TranslationZone.PERFORMATIVE] = performative_count / max(len(words), 1)

        # Normalize
        total = sum(zones.values())
        if total > 0:
            zones = {k: v / total for k, v in zones.items()}

        return zones

    def _get_all_cultural_markers(self) -> Set[str]:
        """Get all cultural markers"""
        all_markers = set()
        for markers in self.cultural_markers.values():
            all_markers.update(markers)
        return all_markers

    def _count_aesthetic_features(self, content: str) -> int:
        """Count aesthetic/poetic features"""
        features = 0

        # Alliteration
        lines = content.split('\n')
        for line in lines:
            words = line.split()
            if len(words) > 1:
                first_letters = [w[0] for w in words if w]
                if len(set(first_letters)) < len(first_letters) * 0.7:
                    features += 1

        # Repetition patterns
        if re.search(r'(\b\w+\b).*\1', content):
            features += 1

        # Parallel structures
        if re.search(r'(\w+\s+\w+).*\1', content):
            features += 1

        return features

    def _find_cultural_markers(self, words: List[str]) -> List[str]:
        """Find cultural markers in text"""
        found_markers = []
        for category, markers in self.cultural_markers.items():
            for word in words:
                if word in markers:
                    found_markers.append(f"{word} ({category})")
        return found_markers

    def _find_untranslatable_elements(self, words: List[str]) -> List[str]:
        """Identify potentially untranslatable elements"""
        untranslatable = []

        # Culture-specific concepts
        culture_specific = ['қурбон', 'ифтор', 'суҳбат', 'ҳамсоя']

        # Onomatopoeia
        onomatopoeia = ['шар-шар', 'тук-тук', 'жир-жир']

        # Emotional expressions
        emotional = ['воҳ', 'эҳ', 'оҳ', 'вой']

        for word in words:
            if word in culture_specific + onomatopoeia + emotional:
                untranslatable.append(word)

        return untranslatable

    def _map_transcultural_flows(self, words: List[str]) -> Dict[str, str]:
        """Map transcultural flows (loanwords and influences)"""
        flows = {}

        # Arabic loanwords
        arabic_patterns = ['ал-', 'ул-', '-ат', '-ият']
        persian_patterns = ['-гоҳ', '-истон', '-зор', '-нома']
        russian_patterns = ['-ция', '-ист', '-изм']

        for word in words:
            if any(word.startswith(p) or word.endswith(p) for p in arabic_patterns):
                flows[word] = 'Arabic'
            elif any(word.endswith(p) for p in persian_patterns):
                flows[word] = 'Persian'
            elif any(word.endswith(p) for p in russian_patterns):
                flows[word] = 'Russian'

        return flows


class LotmanSemioticAnalyzer:
    """Analyzer based on Lotman's semiotic theory"""

    def __init__(self, config: AnalysisConfig):
        self.config = config
        self.semiospheres = config.semiotic_spheres

    def analyze_semiotics(self, content: str, word_freq: List[Tuple[str, int]]) -> SemioticAnalysis:
        """Perform semiotic analysis based on Lotman's theory"""
        words = [w for w, _ in word_freq]

        # Analyze center-periphery dynamics
        dynamics = self._analyze_center_periphery(words)

        # Find boundary crossings
        crossings = self._find_boundary_crossings(content)

        # Detect code switching
        code_switches = self._detect_code_switching(content)

        # Map semiosphere
        mapping = self._map_semiosphere(words)

        return SemioticAnalysis(
            center_periphery_dynamics=dynamics,
            boundary_crossings=crossings,
            code_switching_instances=code_switches,
            semiosphere_mapping=mapping
        )

    def _analyze_center_periphery(self, words: List[str]) -> Dict[str, float]:
        """Analyze center-periphery dynamics"""
        dynamics = {'center': 0, 'periphery': 0, 'boundary': 0}

        for word in words:
            for sphere, sphere_words in self.semiospheres.items():
                if word in sphere_words:
                    dynamics[sphere] += 1

        total = sum(dynamics.values())
        if total > 0:
            dynamics = {k: v / total for k, v in dynamics.items()}

        # Calculate tension index
        dynamics['tension_index'] = abs(dynamics['center'] - dynamics['periphery'])

        return dynamics

    def _find_boundary_crossings(self, content: str) -> List[str]:
        """Find instances of boundary crossing"""
        crossings = []
        lines = content.split('\n')

        for line in lines:
            words = line.split()
            for i in range(len(words) - 1):
                w1, w2 = words[i], words[i + 1]

                sphere1 = self._get_sphere(w1)
                sphere2 = self._get_sphere(w2)

                if sphere1 and sphere2 and sphere1 != sphere2:
                    crossings.append(f"{w1} ({sphere1}) → {w2} ({sphere2})")

        return crossings

    def _detect_code_switching(self, content: str) -> List[Tuple[str, str]]:
        """Detect code switching instances"""
        switches = []

        # Language switches (Tajik/Persian/Arabic)
        arabic_religious = ['аллоҳ', 'бисмиллоҳ', 'иншоаллоҳ']
        persian_poetic = ['дилбар', 'ҷонон', 'нигор']

        lines = content.split('\n')
        for line in lines:
            if any(word in line for word in arabic_religious):
                switches.append(("Religious", line.strip()))
            elif any(word in line for word in persian_poetic):
                switches.append(("Classical", line.strip()))

        return switches

    def _get_sphere(self, word: str) -> Optional[str]:
        """Get semiotic sphere of a word"""
        for sphere, words in self.semiospheres.items():
            if word in words:
                return sphere
        return None

    def _map_semiosphere(self, words: List[str]) -> Dict[str, str]:
        """Map words to semiotic spheres"""
        mapping = {}
        for word in set(words):
            sphere = self._get_sphere(word)
            if sphere:
                mapping[word] = sphere
        return mapping


class SemanticFieldAnalyzer:
    """Analyzer for semantic fields and isotopies"""

    def __init__(self, config: AnalysisConfig):
        self.config = config
        self.semantic_relations = {
            'synonymy': [
                ('муҳаббат', 'ишқ'), ('дил', 'қалб'), ('ватан', 'диёр')
            ],
            'antonymy': [
                ('рӯз', 'шаб'), ('сафед', 'сиёҳ'), ('хушӣ', 'ғам')
            ],
            'hyponymy': [
                ('гул', 'лола'), ('гул', 'садбарг'), ('дарахт', 'чинор')
            ]
        }

    def analyze_semantic_fields(self, words: List[str]) -> List[SemanticField]:
        """Analyze semantic fields in the text"""
        fields = []

        # Group words by thematic fields
        for theme, theme_words in self.config.themes.items():
            field_words = [w for w in words if w in theme_words]

            if field_words:
                # Find related words through semantic relations
                expanded_field = self._expand_semantic_field(field_words, words)

                field = SemanticField(
                    core_lexems=field_words[:5],
                    peripheral_lexems=expanded_field[:5],
                    field_density=len(field_words) / max(len(words), 1),
                    coherence_score=self._calculate_coherence(field_words, words)
                )
                fields.append(field)

        return fields

    def find_isotopy_chains(self, content: str) -> Dict[str, List[str]]:
        """Find isotopy chains (recurring semantic elements)"""
        isotopies = defaultdict(list)
        lines = content.split('\n')

        # Define isotopy categories
        isotopy_patterns = {
            'visual': ['дидан', 'нигоҳ', 'чашм', 'рӯй', 'дида'],
            'movement': ['рафтан', 'омадан', 'гузаштан', 'давидан'],
            'emotion': ['хушӣ', 'ғам', 'шодӣ', 'андӯҳ', 'ҳасрат'],
            'nature': ['об', 'хок', 'оташ', 'ҳаво', 'замин']
        }

        for category, patterns in isotopy_patterns.items():
            for line in lines:
                line_lower = line.lower()
                for pattern in patterns:
                    if pattern in line_lower:
                        isotopies[category].append(line.strip())
                        break

        return dict(isotopies)

    def _expand_semantic_field(self, core_words: List[str], all_words: List[str]) -> List[str]:
        """Expand semantic field through relations"""
        expanded = set()

        for relation_type, relations in self.semantic_relations.items():
            for w1, w2 in relations:
                if w1 in core_words and w2 in all_words:
                    expanded.add(w2)
                elif w2 in core_words and w1 in all_words:
                    expanded.add(w1)

        return list(expanded)

    def _calculate_coherence(self, field_words: List[str], all_words: List[str]) -> float:
        """Calculate semantic field coherence"""
        if not field_words or not all_words:
            return 0.0

        # Check co-occurrence within sliding window
        window_size = 10
        co_occurrences = 0

        for i in range(len(all_words) - window_size):
            window = all_words[i:i + window_size]
            field_in_window = sum(1 for w in window if w in field_words)
            if field_in_window >= 2:
                co_occurrences += 1

        return co_occurrences / max(len(all_words) - window_size, 1)


class StructuralAnalyzer:
    """Enhanced structural analyzer"""

    def __init__(self, config: AnalysisConfig):
        self.config = config
        self.aruz_analyzer = AruzMeterAnalyzer(config)
        self.rhyme_analyzer = AdvancedRhymeAnalyzer()
        self.phonetics = PersianTajikPhonetics()

    def analyze(self, poem_content: str) -> StructuralAnalysis:
        """Comprehensive structural analysis"""
        lines = [line.strip() for line in poem_content.split('\n') if line.strip()]

        if not lines:
            raise ValueError("No valid lines found in poem")

        # Analyze each line
        line_analyses = []
        syllable_counts = []
        syllable_patterns = []
        rhyme_analyses = []

        for line in lines:
            # Syllable analysis
            phonetic = self.phonetics.analyze_phonetics(line)
            syllables = self.aruz_analyzer._extract_prosodic_syllables(line, phonetic)
            syllable_patterns.append(syllables)
            syllable_counts.append(len(syllables))

            # Rhyme analysis
            rhyme = self.rhyme_analyzer.analyze_rhyme(line)
            rhyme_analyses.append(rhyme)

            # Meter analysis
            aruz = self.aruz_analyzer.analyze_meter(line)

            line_analyses.append({
                'syllables': syllables,
                'rhyme': rhyme,
                'aruz': aruz
            })

        # Generate rhyme pattern
        rhyme_pattern = self._generate_rhyme_pattern(rhyme_analyses)

        # Detect stanza structure
        stanza_structure = self._detect_stanza_structure(lines, rhyme_pattern)

        # Calculate metrics
        avg_syllables = sum(syllable_counts) / len(syllable_counts) if syllable_counts else 0
        prosodic_consistency = self._calculate_prosodic_consistency(line_analyses)

        # Get overall meter
        meters = [la['aruz'] for la in line_analyses]
        overall_aruz = self._determine_overall_meter(meters)

        return StructuralAnalysis(
            lines=len(lines),
            syllables_per_line=syllable_counts,
            syllable_patterns=syllable_patterns,
            aruz_analysis=overall_aruz,
            rhyme_scheme=rhyme_analyses,
            rhyme_pattern=rhyme_pattern,
            stanza_structure=stanza_structure,
            avg_syllables=round(avg_syllables, 2),
            prosodic_consistency=prosodic_consistency,
            meter_confidence=overall_aruz.confidence
        )

    # In StructuralAnalyzer class
    def _generate_rhyme_pattern(self, rhyme_analyses: List[EnhancedRhymeAnalysis]) -> str:
        if not rhyme_analyses:
            return ""

        pattern = []
        rhyme_groups = {}
        next_label = 'A'

        for rhyme in rhyme_analyses:
            # Create a hashable key with ALL required fields
            rhyme_key = (rhyme.qafiyeh, rhyme.radif, rhyme.phonetic_rhyme,
                         rhyme.rhyme_type, rhyme.rhyme_position, rhyme.confidence)
            matched = False

            for prev_key, label in rhyme_groups.items():
                # Recreate object with all fields
                prev_rhyme = EnhancedRhymeAnalysis(*prev_key)
                similarity = self.rhyme_analyzer.calculate_rhyme_similarity(rhyme, prev_rhyme)
                if similarity > 0.7:
                    pattern.append(label)
                    matched = True
                    break

            if not matched:
                pattern.append(next_label)
                rhyme_groups[rhyme_key] = next_label
                next_label = chr(ord(next_label) + 1)

        return ''.join(pattern)

    def _detect_stanza_structure(self, lines: List[str], rhyme_pattern: str) -> str:
        """Detect stanza structure"""
        if len(lines) <= 2:
            return "monostich" if len(lines) == 1 else "couplet"

        # Check for ghazal pattern
        if rhyme_pattern.startswith('AA') and all(c in ['A', 'B'] for c in rhyme_pattern):
            return "ghazal"

        # Check for rubaiyat
        if len(lines) == 4 and rhyme_pattern in ['AABA', 'AAAA']:
            return "rubaiyat"

        # Check for qasida
        if len(lines) > 10 and rhyme_pattern[:2] == 'AA':
            return "qasida"

        return "free_verse"

    def _calculate_prosodic_consistency(self, line_analyses: List[Dict]) -> float:
        """Calculate overall prosodic consistency"""
        if not line_analyses:
            return 0.0

        # Meter consistency
        meters = [la['aruz'].identified_meter for la in line_analyses]
        unique_meters = set(meters)
        meter_consistency = 1.0 / len(unique_meters) if unique_meters else 0.0

        # Syllable consistency
        syllable_counts = [len(la['syllables']) for la in line_analyses]
        if syllable_counts:
            avg = sum(syllable_counts) / len(syllable_counts)
            variance = sum((c - avg) ** 2 for c in syllable_counts) / len(syllable_counts)
            syllable_consistency = 1.0 / (1.0 + variance / max(avg, 1))
        else:
            syllable_consistency = 0.0

        return (meter_consistency + syllable_consistency) / 2

    def _determine_overall_meter(self, meters: List[AruzAnalysis]) -> AruzAnalysis:
        """Determine overall meter from line analyses"""
        if not meters:
            return AruzAnalysis(
                identified_meter="unknown",
                pattern_match="",
                feet=[],
                confidence=MeterConfidence.NONE,
                pattern_accuracy=0.0,
                variations_detected=[],
                caesura_positions=[]
            )

        # Find most common meter
        meter_counts = Counter(m.identified_meter for m in meters)
        most_common = meter_counts.most_common(1)[0][0]

        # Return the best example of the most common meter
        best_example = max(
            (m for m in meters if m.identified_meter == most_common),
            key=lambda m: m.pattern_accuracy
        )

        return best_example

class EnhancedPoemSplitter:
    """Advanced poem splitter for Tajik Cyrillic poetry collections with manual correction support"""
    
    def __init__(self, config: AnalysisConfig):
        self.config = config
        self.logger = logging.getLogger(f"{__name__}.EnhancedPoemSplitter")
        
    def get_split_suggestions(self, text: str) -> List[int]:
        """
        Returns line indices where a new poem is likely to start.
        These suggestions are shown to the user for manual confirmation/correction.
        """
        lines = text.split('\n')
        suggestions = []
        
        for i, line in enumerate(lines):
            score = 0
            
            # 1. Title-like lines (strong signal)
            if self._looks_like_title(line):
                score += 2
            
            # 2. Empty line followed by a title-like line
            if i > 0 and not lines[i-1].strip() and len(line.strip()) > 0:
                score += 1.5
            
            # 3. Lines with poem markers like "***" or "---"
            if re.match(r'^[\*\-=]{3,}$', line.strip()):
                # Suggest split before this line
                suggestions.append(max(0, i-1))
                continue
                
            # 4. Line numbers (e.g., "1." or "(2)")
            if re.match(r'^\s*[\d]+[\.\)]\s*[A-ZА-Я]', line):
                score += 1
            
            # 5. Uppercase at the beginning of the line after an empty line
            if i > 0 and not lines[i-1].strip() and line.strip() and line.strip()[0].isupper():
                score += 0.5
            
            if score >= 1.5:  # Threshold
                suggestions.append(i)
        
        # Remove suggestions that are too close (within 3 lines)
        if suggestions:
            filtered = [suggestions[0]]
            for s in suggestions[1:]:
                if s - filtered[-1] > 3:
                    filtered.append(s)
            suggestions = filtered
        
        return suggestions

    def _looks_like_title(self, line: str) -> bool:
        """Simple heuristic to recognize title lines."""
        line = line.strip()
        if not line or len(line) > 150:
            return False
        
        # Does not end with punctuation
        if line.endswith(('.', '!', '?', ':', ',')):
            return False
        
        # Starts with an uppercase letter
        if not line[0].isupper():
            return False
        
        # Not written entirely in uppercase (not a "SCREAM")
        if line.isupper():
            return False
        
        return True
        
class ContentAnalyzer:
    """Enhanced content analyzer"""

    def __init__(self, config: AnalysisConfig):
        self.config = config
        self.semantic_analyzer = SemanticFieldAnalyzer(config)
        self.lexicon = self._load_lexicon()

        # Define archaic words
        self.archaisms = {
            'зи', 'ки', 'чу', 'зеро', 'балки', 'андар', 'бар', 'аз-ан-ки'
        }

    def analyze(self, poem_content: str) -> ContentAnalysis:
        """Comprehensive content analysis"""
        # Extract words
        words = re.findall(r'[\wӣӯ]+', poem_content.lower())
        word_freq = Counter(words)

        # Find neologisms and archaisms
        neologisms = self._find_neologisms(words)
        archaisms = self._find_archaisms(words)

        # Analyze themes
        theme_distribution = self._analyze_themes(words)

        # Analyze semantic fields
        semantic_fields = self.semantic_analyzer.analyze_semantic_fields(words)

        # Find isotopy chains
        isotopy_chains = self.semantic_analyzer.find_isotopy_chains(poem_content)

        # Calculate lexical diversity
        lexical_diversity = len(set(words)) / len(words) if words else 0

        # Determine stylistic register
        stylistic_register = self._determine_register(words, archaisms, neologisms)

        return ContentAnalysis(
            word_frequencies=word_freq.most_common(20),
            neologisms=neologisms[:self.config.max_neologisms],
            archaisms=list(archaisms),
            theme_distribution=theme_distribution,
            semantic_fields=semantic_fields,
            isotopy_chains=isotopy_chains,
            lexical_diversity=round(lexical_diversity, 3),
            stylistic_register=stylistic_register
        )

    def _load_lexicon(self) -> Set[str]:
        """Load lexicon from configured file path"""
        try:
            lexicon_path = Path(self.config.lexicon_path)
            if lexicon_path.exists():
                with open(lexicon_path, 'r', encoding='utf-8') as f:
                    return set(json.load(f))
            else:
                logger.warning(f"Lexicon file not found at: {lexicon_path}")
        except Exception as e:
            logger.error(f"Error loading lexicon: {e}")
        return set()  # Fallback to empty set

    def _find_neologisms(self, words: List[str]) -> List[str]:
        """Find neologisms (words not in standard lexicon)"""
        if not self.lexicon:
            return []

        neologisms = []
        for word in set(words):
            if word not in self.lexicon and word not in self.archaisms:
                if not word.isdigit() and len(word) > 2:
                    neologisms.append(word)

        return sorted(neologisms)

    def _find_archaisms(self, words: List[str]) -> Set[str]:
        """Find archaic words"""
        return set(word for word in words if word in self.archaisms)

    def _analyze_themes(self, words: List[str]) -> Dict[str, int]:
        """Analyze thematic distribution"""
        theme_counts = {}

        for theme, keywords in self.config.themes.items():
            count = sum(1 for word in words if word in keywords)
            theme_counts[theme] = count

        return theme_counts

    def _determine_register(self, words: List[str], archaisms: Set[str],
                            neologisms: List[str]) -> str:
        """Determine stylistic register"""
        total_words = len(words)

        if not total_words:
            return "unknown"

        archaic_ratio = len(archaisms) / total_words
        neologism_ratio = len(neologisms) / total_words

        if archaic_ratio > 0.05:
            return "classical"
        elif neologism_ratio > 0.05:
            return "modern"
        elif archaic_ratio > 0.02 and neologism_ratio < 0.02:
            return "neo-classical"
        else:
            return "contemporary"


class LiteraryAssessor:
    """Multi-perspective literary assessment"""

    @staticmethod
    def assess(structural: StructuralAnalysis, content: ContentAnalysis,
               translation: TranslationAnalysis, semiotic: SemioticAnalysis) -> LiteraryAssessment:
        """Comprehensive literary assessment"""

        # German perspective - formal perfection
        german_score = 0
        if structural.rhyme_pattern in ['ABAB', 'AABB', 'ABBA', 'ABCABC']:
            german_score += 2
        if 8 <= structural.avg_syllables <= 12:
            german_score += 2
        if structural.prosodic_consistency > 0.8:
            german_score += 1

        # Persian tradition - classical forms
        persian_score = 0
        if structural.aruz_analysis.identified_meter != "unknown":
            persian_score += 2
        if structural.stanza_structure in ['ghazal', 'qasida', 'rubaiyat']:
            persian_score += 2
        if content.stylistic_register in ['classical', 'neo-classical']:
            persian_score += 1

        # Tajik elements - cultural authenticity
        tajik_score = 0
        tajik_score += min(3, content.theme_distribution.get('Homeland', 0))
        if len(translation.cultural_markers) > 0:
            tajik_score += 1
        if semiotic.center_periphery_dynamics.get('center', 0) > 0.5:
            tajik_score += 1

        # Modernist features
        modern_score = 0
        if content.lexical_diversity > 0.7:
            modern_score += 2
        if len(content.neologisms) > 3:
            modern_score += 1
        if structural.stanza_structure == "free_verse":
            modern_score += 2

        # Postcolonial markers
        postcolonial_score = 0
        if translation.transcultural_flows:
            postcolonial_score += 2
        if semiotic.boundary_crossings:
            postcolonial_score += 2
        if translation.translation_zones[TranslationZone.CULTURAL] > 0.3:
            postcolonial_score += 1

        return LiteraryAssessment(
            german_perspective=min(5, german_score),
            persian_tradition=min(5, persian_score),
            tajik_elements=min(5, tajik_score),
            modernist_features=min(5, modern_score),
            postcolonial_markers=min(5, postcolonial_score)
        )


class QualityValidator:
    """Validate analysis quality for scientific rigor"""

    @staticmethod
    def validate_analysis(analysis: ComprehensiveAnalysis) -> Dict[str, Any]:
        """Validate analysis quality"""
        warnings = []
        recommendations = []
        quality_score = 1.0

        # Check structural quality
        if analysis.structural.meter_confidence == MeterConfidence.NONE:
            warnings.append("No reliable meter detected")
            recommendations.append("Manual prosodic verification recommended")
            quality_score *= 0.7

        if analysis.structural.prosodic_consistency < 0.5:
            warnings.append("Low prosodic consistency")
            recommendations.append("Check for textual corruption or free verse intention")
            quality_score *= 0.8

        # Check content quality
        if analysis.content.lexical_diversity < 0.3:
            warnings.append("Very low lexical diversity")
            recommendations.append("Verify text completeness")
            quality_score *= 0.9

        # Check translation analysis
        if sum(analysis.translation.translation_zones.values()) == 0:
            warnings.append("No translation zones identified")
            recommendations.append("Text may be too short for translation analysis")
            quality_score *= 0.8

        # Check semiotic analysis
        if not analysis.semiotic.semiosphere_mapping:
            warnings.append("No semiotic mapping possible")
            recommendations.append("Expand lexicon for better semiotic analysis")
            quality_score *= 0.9

        reliability = "high" if quality_score > 0.8 else "medium" if quality_score > 0.6 else "low"

        return {
            'quality_score': round(quality_score, 2),
            'reliability': reliability,
            'warnings': warnings,
            'recommendations': recommendations,
            'timestamp': datetime.now().isoformat()
        }


class ExcelReporter:
    """Enhanced Excel report generation"""

    def __init__(self):
        self.header_font = Font(bold=True, color="FFFFFF")
        self.header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        self.border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        self.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    def create_comprehensive_report(self, results: List[Dict[str, Any]],
                                    filename: str = "tajik_poetry_analysis.xlsx"):
        """Create comprehensive Excel report with all analyses"""
        try:
            wb = openpyxl.Workbook()

            # Overview sheet
            self._create_overview_sheet(wb, results)

            # Structural analysis sheet
            self._create_structural_sheet(wb, results)

            # Content analysis sheet
            self._create_content_sheet(wb, results)

            # Translation theory sheet
            self._create_translation_sheet(wb, results)

            # Semiotic analysis sheet
            self._create_semiotic_sheet(wb, results)

            # Literary assessment sheet
            self._create_literary_sheet(wb, results)

            # Quality metrics sheet
            self._create_quality_sheet(wb, results)

            # Save
            wb.save(filename)
            logger.info(f"Comprehensive report saved as: {filename}")

        except Exception as e:
            logger.error(f"Error creating report: {e}")
            raise

    def _create_overview_sheet(self, wb: openpyxl.Workbook, results: List[Dict[str, Any]]):
        """Create overview sheet"""
        ws = wb.active
        ws.title = "Overview"

        headers = [
            "ID", "Title", "Lines", "Meter", "Confidence", "Rhyme Pattern",
            "Stanza Form", "Register", "Main Theme", "Quality Score"
        ]

        # Set column widths
        for i, width in enumerate([12, 30, 8, 15, 12, 15, 15, 12, 15, 12], 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

        # Write headers
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.border = self.border
            cell.alignment = self.alignment

        # Write data
        for row_num, result in enumerate(results, 2):
            analysis = result["analysis"]
            validation = result.get("validation", {})

            # Find main theme
            main_theme = max(
                analysis.content.theme_distribution.items(),
                key=lambda x: x[1],
                default=("Unknown", 0)
            )[0]

            values = [
                result["poem_id"],
                result["title"],
                analysis.structural.lines,
                analysis.structural.aruz_analysis.identified_meter,
                analysis.structural.meter_confidence.value,
                analysis.structural.rhyme_pattern,
                analysis.structural.stanza_structure,
                analysis.content.stylistic_register,
                main_theme,
                validation.get("quality_score", "N/A")
            ]

            for col_num, value in enumerate(values, 1):
                cell = ws.cell(row=row_num, column=col_num, value=value)
                cell.border = self.border
                if col_num > 2:
                    cell.alignment = self.alignment

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

    def _create_structural_sheet(self, wb: openpyxl.Workbook, results: List[Dict[str, Any]]):
        """Create detailed structural analysis sheet"""
        ws = wb.create_sheet(title="Structural Analysis")

        headers = [
            "Poem ID", "Line #", "Line Text", "Syllables", "Meter Pattern",
            "Feet", "Rhyme (Qafiyeh)", "Radif", "Rhyme Type"
        ]

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill

        # Set column widths
        widths = [12, 8, 50, 10, 20, 30, 15, 20, 12]
        for i, width in enumerate(widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

        row_num = 2
        for result in results:
            poem_id = result["poem_id"]
            content = result["content"]
            structural = result["analysis"].structural

            lines = [line.strip() for line in content.split('\n') if line.strip()]

            for line_idx, line in enumerate(lines):
                # Get line-specific data
                syllable_count = structural.syllables_per_line[line_idx] if line_idx < len(
                    structural.syllables_per_line) else 0

                # Get syllable pattern
                if line_idx < len(structural.syllable_patterns):
                    pattern = ''.join([s.weight.value for s in structural.syllable_patterns[line_idx]])
                else:
                    pattern = ""

                # Get feet
                feet_str = ""
                if structural.aruz_analysis.feet:
                    # Simplified feet representation
                    feet_str = ", ".join([f.name for f in structural.aruz_analysis.feet[:3]])

                # Get rhyme info
                if line_idx < len(structural.rhyme_scheme):
                    rhyme = structural.rhyme_scheme[line_idx]
                    qafiyeh = rhyme.qafiyeh
                    radif = rhyme.radif
                    rhyme_type = rhyme.rhyme_type
                else:
                    qafiyeh = radif = rhyme_type = ""

                values = [
                    poem_id,
                    line_idx + 1,
                    line,
                    syllable_count,
                    pattern,
                    feet_str,
                    qafiyeh,
                    radif,
                    rhyme_type
                ]

                for col_num, value in enumerate(values, 1):
                    ws.cell(row=row_num, column=col_num, value=value)

                row_num += 1

    def _create_content_sheet(self, wb: openpyxl.Workbook, results: List[Dict[str, Any]]):
        """Create content analysis sheet"""
        ws = wb.create_sheet(title="Content Analysis")

        headers = [
            "Poem ID", "Top Words", "Neologisms", "Archaisms",
            "Themes", "Semantic Fields", "Isotopies", "Lexical Diversity"
        ]

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill

        # Set column widths
        widths = [12, 40, 30, 30, 30, 40, 40, 15]
        for i, width in enumerate(widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

        for row_num, result in enumerate(results, 2):
            content = result["analysis"].content

            # Format data
            top_words = ", ".join([f"{w}({c})" for w, c in content.word_frequencies[:10]])
            neologisms = ", ".join(content.neologisms[:10])
            archaisms = ", ".join(content.archaisms)
            themes = ", ".join([f"{k}({v})" for k, v in content.theme_distribution.items() if v > 0])

            # Semantic fields
            sem_fields = []
            for field in content.semantic_fields[:3]:
                sem_fields.append(f"{','.join(field.core_lexems[:3])} (d={field.field_density:.2f})")
            semantic_fields_str = "; ".join(sem_fields)

            # Isotopies
            isotopies = []
            for category, chains in content.isotopy_chains.items():
                isotopies.append(f"{category}({len(chains)})")
            isotopies_str = ", ".join(isotopies)

            values = [
                result["poem_id"],
                top_words,
                neologisms,
                archaisms,
                themes,
                semantic_fields_str,
                isotopies_str,
                content.lexical_diversity
            ]

            for col_num, value in enumerate(values, 1):
                cell = ws.cell(row=row_num, column=col_num, value=value)
                cell.border = self.border

    def _create_translation_sheet(self, wb: openpyxl.Workbook, results: List[Dict[str, Any]]):
        """Create translation theory analysis sheet"""
        ws = wb.create_sheet(title="Translation Analysis")

        headers = [
            "Poem ID", "Linguistic Zone", "Cultural Zone", "Aesthetic Zone",
            "Performative Zone", "Cultural Markers", "Untranslatable", "Transcultural Flows"
        ]

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill

        # Set column widths
        widths = [12, 15, 15, 15, 15, 40, 30, 40]
        for i, width in enumerate(widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

        for row_num, result in enumerate(results, 2):
            translation = result["analysis"].translation

            # Format zone values
            zones = translation.translation_zones

            # Format lists
            markers = ", ".join(translation.cultural_markers[:10])
            untranslatable = ", ".join(translation.untranslatable_elements[:10])
            flows = ", ".join([f"{k}({v})" for k, v in list(translation.transcultural_flows.items())[:10]])

            values = [
                result["poem_id"],
                f"{zones.get(TranslationZone.LINGUISTIC, 0):.2f}",
                f"{zones.get(TranslationZone.CULTURAL, 0):.2f}",
                f"{zones.get(TranslationZone.AESTHETIC, 0):.2f}",
                f"{zones.get(TranslationZone.PERFORMATIVE, 0):.2f}",
                markers,
                untranslatable,
                flows
            ]

            for col_num, value in enumerate(values, 1):
                cell = ws.cell(row=row_num, column=col_num, value=value)
                cell.border = self.border
                if 2 <= col_num <= 5:
                    cell.alignment = self.alignment

    def _create_semiotic_sheet(self, wb: openpyxl.Workbook, results: List[Dict[str, Any]]):
        """Create semiotic analysis sheet"""
        ws = wb.create_sheet(title="Semiotic Analysis")

        headers = [
            "Poem ID", "Center", "Periphery", "Boundary", "Tension Index",
            "Boundary Crossings", "Code Switches", "Semiosphere Mapping"
        ]

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill

        # Set column widths
        widths = [12, 10, 10, 10, 12, 40, 40, 40]
        for i, width in enumerate(widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

        for row_num, result in enumerate(results, 2):
            semiotic = result["analysis"].semiotic
            dynamics = semiotic.center_periphery_dynamics

            # Format complex data
            crossings = "; ".join(semiotic.boundary_crossings[:5])
            switches = "; ".join([f"{t}:{l[:30]}..." for t, l in semiotic.code_switching_instances[:3]])
            mapping = ", ".join([f"{k}({v})" for k, v in list(semiotic.semiosphere_mapping.items())[:10]])

            values = [
                result["poem_id"],
                f"{dynamics.get('center', 0):.2f}",
                f"{dynamics.get('periphery', 0):.2f}",
                f"{dynamics.get('boundary', 0):.2f}",
                f"{dynamics.get('tension_index', 0):.2f}",
                crossings,
                switches,
                mapping
            ]

            for col_num, value in enumerate(values, 1):
                cell = ws.cell(row=row_num, column=col_num, value=value)
                cell.border = self.border
                if 2 <= col_num <= 5:
                    cell.alignment = self.alignment

    def _create_literary_sheet(self, wb: openpyxl.Workbook, results: List[Dict[str, Any]]):
        """Create literary assessment sheet with visualization"""
        ws = wb.create_sheet(title="Literary Assessment")

        headers = [
            "Poem ID", "Title", "German", "Persian", "Tajik",
            "Modernist", "Postcolonial", "Total Score"
        ]

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill

        # Set column widths
        widths = [12, 30, 10, 10, 10, 10, 12, 12]
        for i, width in enumerate(widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

        for row_num, result in enumerate(results, 2):
            literary = result["analysis"].literary

            total_score = (
                    literary.german_perspective +
                    literary.persian_tradition +
                    literary.tajik_elements +
                    literary.modernist_features +
                    literary.postcolonial_markers
            )

            values = [
                result["poem_id"],
                result["title"],
                literary.german_perspective,
                literary.persian_tradition,
                literary.tajik_elements,
                literary.modernist_features,
                literary.postcolonial_markers,
                total_score
            ]

            for col_num, value in enumerate(values, 1):
                cell = ws.cell(row=row_num, column=col_num, value=value)
                cell.border = self.border
                if col_num > 2:
                    cell.alignment = self.alignment

        # Add chart if there are results
        if len(results) > 0:
            chart = BarChart()
            chart.title = "Literary Assessment Distribution"
            chart.x_axis.title = "Perspective"
            chart.y_axis.title = "Average Score"

            # Calculate averages
            perspectives = ["German", "Persian", "Tajik", "Modernist", "Postcolonial"]
            avg_row = len(results) + 3

            ws.cell(row=avg_row, column=1, value="Average")
            for i, perspective in enumerate(perspectives, 3):
                avg_formula = f"=AVERAGE({openpyxl.utils.get_column_letter(i)}2:{openpyxl.utils.get_column_letter(i)}{len(results) + 1})"
                ws.cell(row=avg_row, column=i, value=avg_formula)

            # Add data to chart
            data = Reference(ws, min_col=3, min_row=avg_row, max_col=7, max_row=avg_row)
            categories = Reference(ws, min_col=3, min_row=1, max_col=7, max_row=1)

            chart.add_data(data, from_rows=True)
            chart.set_categories(categories)

            ws.add_chart(chart, f"A{avg_row + 2}")

    def _create_quality_sheet(self, wb: openpyxl.Workbook, results: List[Dict[str, Any]]):
        """Create quality metrics sheet"""
        ws = wb.create_sheet(title="Quality Metrics")

        headers = [
            "Poem ID", "Quality Score", "Reliability", "Warnings", "Recommendations"
        ]

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill

        # Set column widths
        widths = [12, 15, 15, 50, 50]
        for i, width in enumerate(widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

        for row_num, result in enumerate(results, 2):
            validation = result.get("validation", {})

            warnings = "; ".join(validation.get("warnings", []))
            recommendations = "; ".join(validation.get("recommendations", []))

            values = [
                result["poem_id"],
                validation.get("quality_score", "N/A"),
                validation.get("reliability", "N/A"),
                warnings,
                recommendations
            ]

            for col_num, value in enumerate(values, 1):
                cell = ws.cell(row=row_num, column=col_num, value=value)
                cell.border = self.border
                cell.alignment = Alignment(wrap_text=True, vertical="top")

                # Color code based on reliability
                if col_num == 3 and value == "high":
                    cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                elif col_num == 3 and value == "medium":
                    cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                elif col_num == 3 and value == "low":
                    cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")


class TajikPoemAnalyzer:
    """Main analyzer class coordinating all components"""

    def __init__(self, config: Optional[AnalysisConfig] = None):
        self.config = config or AnalysisConfig()

        # Initialize all analyzers
        self.structural_analyzer = StructuralAnalyzer(self.config)
        self.content_analyzer = ContentAnalyzer(self.config)
        self.translation_analyzer = TranslationTheoreticalAnalyzer(self.config)
        self.semiotic_analyzer = LotmanSemioticAnalyzer(self.config)
        self.excel_reporter = ExcelReporter()

        logger.info("TajikPoemAnalyzer initialized with all scientific components")

    def analyze_poem(self, poem_content: str) -> ComprehensiveAnalysis:
        """Perform comprehensive analysis of a single poem"""
        if not poem_content or len(poem_content.strip()) < self.config.min_poem_length:
            raise ValueError("Poem content is too short or empty")

        # Structural analysis
        structural = self.structural_analyzer.analyze(poem_content)

        # Content analysis
        content = self.content_analyzer.analyze(poem_content)

        # Translation theoretical analysis
        translation = self.translation_analyzer.analyze_translation_aspects(
            poem_content, content.word_frequencies
        )

        # Semiotic analysis
        semiotic = self.semiotic_analyzer.analyze_semiotics(
            poem_content, content.word_frequencies
        )

        # Literary assessment
        literary = LiteraryAssessor.assess(structural, content, translation, semiotic)

        # Quality metrics
        analysis = ComprehensiveAnalysis(
            structural=structural,
            content=content,
            translation=translation,
            semiotic=semiotic,
            literary=literary,
            quality_metrics={}
        )

        # Validate and add quality metrics
        validation = QualityValidator.validate_analysis(analysis)
        analysis.quality_metrics = validation

        return analysis

    def analyze_text(self, text: str) -> List[Dict[str, Any]]:
        """Analyze text containing multiple poems"""
        # Preprocess text
        poems = self._split_poems(text)

        logger.info(f"Found {len(poems)} poems to analyze")

        results = []
        for poem in poems:
            try:
                analysis = self.analyze_poem(poem.content)
                validation = QualityValidator.validate_analysis(analysis)

                results.append({
                    "poem_id": poem.poem_id,
                    "title": poem.title,
                    "content": poem.content,
                    "analysis": analysis,
                    "validation": validation
                })
            except Exception as e:
                logger.error(f"Error analyzing poem {poem.poem_id}: {e}")
                continue

        return results

    def analyze_file(self, filepath: str, output_file: Optional[str] = None) -> List[Dict[str, Any]]:
        """Analyze poems from a file"""
        try:
            file_path = Path(filepath)
            if not file_path.exists():
                raise FileNotFoundError(f"File not found: {filepath}")

            with open(file_path, 'r', encoding='utf-8') as f:
                text = f.read()

            results = self.analyze_text(text)

            # Create output filename
            if output_file is None:
                output_file = f"{file_path.stem}_comprehensive_analysis.xlsx"

            # Create report
            self.excel_reporter.create_comprehensive_report(results, output_file)

            # Also save JSON
            json_file = f"{file_path.stem}_analysis.json"
            self._save_json_results(results, json_file)

            logger.info(f"Analysis complete. Results saved to {output_file} and {json_file}")
            return results

        except Exception as e:
            logger.error(f"Error analyzing file {filepath}: {e}")
            raise

    def _split_poems(self, text: str) -> List[PoemData]:
        """Split text into individual poems"""
        # Normalize text
        text = unicodedata.normalize('NFC', text)

        # Define separators
        separators = [
            r'\*{5,}',
            r'-{5,}',
            r'={5,}',
            r'_{5,}',
            r'#+\s*\d+\s*#+',
            r'\n\s*\n\s*\n+'
        ]

        # Split by separators
        pattern = '|'.join(separators)
        blocks = re.split(pattern, text)

        poems = []
        for i, block in enumerate(blocks, 1):
            block = block.strip()
            if len(block) < self.config.min_poem_length:
                logger.warning(f"Skipping block {i} - too short")
                continue

            # Extract title (first line if appropriate)
            lines = block.split('\n')
            if lines:
                first_line = lines[0].strip()
                if (self.config.min_title_length <= len(first_line) <= self.config.max_title_length
                        and not first_line.endswith(('.', '!', '?'))):
                    title = first_line
                    content = '\n'.join(lines[1:]).strip()
                else:
                    title = f"Poem {i}"
                    content = block

                poems.append(PoemData(
                    title=title,
                    content=content,
                    poem_id=f"poem_{i:03d}"
                ))

        return poems

    def _save_json_results(self, results: List[Dict[str, Any]], filename: str):
        """Save results as JSON"""
        # Convert dataclasses to dictionaries
        json_results = []

        for result in results:
            # Use asdict but handle non-serializable types
            json_result = {
                "poem_id": result["poem_id"],
                "title": result["title"],
                "content": result["content"],
                "validation": result["validation"]
            }

            # Manually convert analysis components
            analysis = result["analysis"]
            json_result["analysis"] = {
                "structural": self._serialize_structural(analysis.structural),
                "content": self._serialize_content(analysis.content),
                "translation": self._serialize_translation(analysis.translation),
                "semiotic": self._serialize_semiotic(analysis.semiotic),
                "literary": asdict(analysis.literary),
                "quality_metrics": analysis.quality_metrics
            }

            json_results.append(json_result)

        with open(filename, 'w', encoding='utf-8') as f:
            json.dump(json_results, f, ensure_ascii=False, indent=2)

    def _serialize_structural(self, structural: StructuralAnalysis) -> Dict:
        """Serialize structural analysis"""
        return {
            "lines": structural.lines,
            "syllables_per_line": structural.syllables_per_line,
            "rhyme_pattern": structural.rhyme_pattern,
            "stanza_structure": structural.stanza_structure,
            "avg_syllables": structural.avg_syllables,
            "prosodic_consistency": structural.prosodic_consistency,
            "meter_confidence": structural.meter_confidence.value,
            "meter": structural.aruz_analysis.identified_meter
        }

    def _serialize_content(self, content: ContentAnalysis) -> Dict:
        """Serialize content analysis"""
        return {
            "word_frequencies": content.word_frequencies[:20],
            "neologisms": content.neologisms,
            "archaisms": list(content.archaisms),
            "theme_distribution": content.theme_distribution,
            "lexical_diversity": content.lexical_diversity,
            "stylistic_register": content.stylistic_register,
            "isotopy_chains": {k: v[:5] for k, v in content.isotopy_chains.items()}
        }

    def _serialize_translation(self, translation: TranslationAnalysis) -> Dict:
        """Serialize translation analysis"""
        return {
            "translation_zones": {z.value: v for z, v in translation.translation_zones.items()},
            "cultural_markers": translation.cultural_markers[:10],
            "untranslatable_elements": translation.untranslatable_elements[:10],
            "transcultural_flows": dict(list(translation.transcultural_flows.items())[:10])
        }

    def _serialize_semiotic(self, semiotic: SemioticAnalysis) -> Dict:
        """Serialize semiotic analysis"""
        return {
            "center_periphery_dynamics": semiotic.center_periphery_dynamics,
            "boundary_crossings": semiotic.boundary_crossings[:10],
            "code_switching_instances": [
                {"type": t, "line": l[:100]} for t, l in semiotic.code_switching_instances[:5]
            ],
            "semiosphere_mapping": dict(list(semiotic.semiosphere_mapping.items())[:20])
        }


def create_sample_lexicon():
    """Create a comprehensive Tajik lexicon for testing"""
    lexicon = [
        # Basic vocabulary
        "ва", "дар", "бо", "аз", "то", "барои", "чун", "ки", "агар", "аммо",
        "ё", "на", "ҳам", "низ", "лекин", "вале", "пас", "зеро", "ҳар", "ин",

        # Love and emotion
        "муҳаббат", "ишқ", "дил", "маҳбуб", "ёр", "дилбар", "ошиқ", "маъшуқ",
        "қалб", "ҷон", "ҷонон", "нигор", "дӯст", "муҳаббатнома",

        # Nature
        "дарё", "кӯҳ", "гул", "баҳор", "навбаҳор", "осмон", "офтоб", "моҳ",
        "ситора", "абр", "борон", "барф", "дарахт", "баргу", "чашма", "соҳил",

        # Homeland and geography
        "ватан", "тоҷикистон", "диёр", "марзу", "бум", "кишвар", "шаҳр",
        "деҳа", "кӯча", "хона", "дар", "роҳ", "кӯпрук",

        # Religion and spirituality
        "худо", "аллоҳ", "ҷаннат", "ибодат", "намоз", "масҷид", "паёмбар",
        "қуръон", "дуо", "тасбеҳ", "ҳаҷ", "рӯза",

        # Cultural terms
        "дастархон", "чойхона", "меҳмон", "ош", "нон", "чой", "самбӯса",
        "курпача", "сузанӣ", "атлас", "чакан",

        # Time and seasons
        "рӯз", "шаб", "субҳ", "бегоҳ", "шом", "саҳар", "тобистон", "зимистон",
        "тирамоҳ", "баҳор", "имрӯз", "дирӯз", "фардо", "ҳафта", "моҳ", "сол",

        # Human and family
        "одам", "инсон", "мард", "зан", "кӯдак", "падар", "модар", "бародар",
        "хоҳар", "фарзанд", "оила", "хонавода",

        # Actions and states
        "рафтан", "омадан", "дидан", "шунидан", "гуфтан", "хондан", "навиштан",
        "хобидан", "хӯрдан", "нӯшидан", "кор", "кардан", "зиндагӣ", "мурдан"
    ]

    try:
        with open('tajik_lexicon2.json', 'w', encoding='utf-8') as f:
            json.dump(lexicon, f, ensure_ascii=False, indent=2)
        logger.info("Comprehensive lexicon created: tajik_lexicon2.json")
    except Exception as e:
        logger.warning(f"Could not create lexicon file: {e}")


def main():
    """Main function demonstrating the analyzer"""
    # Create lexicon
    # create_sample_lexicon()

    # Sample Tajik poetry
    sample_text = """
Дар кӯҳсори ватан гулҳо мешукуфанд,
Дили ошиқ аз муҳаббат меларзад.
Баҳори нав ба замин таҷдид меорад,
Навиди хушҳолии мардум мерасад.

*****

Эй ватан, эй модари меҳрубон,
Дар оғӯши ту ёфтам ҷон.
Кӯҳҳои ту сари фалак расида,
Дарёҳои ту ҷовидон.

*****

Дар чашмаҳои кӯҳистон об ҷӯшад,
Мисли ашки шодии модарон.
Замини ҳосилхез интизор аст,
То биёяд фасли деҳқон.

*****

Шаби торик гузашт, субҳи равшан омад,
Дар дили ман умеди тоза домад.
Ситораҳо чун чашмони маҳбубон,
Ба осмони шаб нигоҳ карданд, оҳ карданд.
"""

    try:
        custom_config = AnalysisConfig(
            lexicon_path='tajik_lexicon.json'
        )

        # Initialize analyzer
        analyzer = TajikPoemAnalyzer(config=custom_config)

        # Analyze sample text
        # results = analyzer.analyze_text(sample_text)
        results = analyzer.analyze_file("poems.txt", "output.xlsx")

        # Create comprehensive report
        analyzer.excel_reporter.create_comprehensive_report(results, "tajik_poetry_analysis_demo.xlsx")

        print("=== TAJIK POETRY ANALYZER - SCIENTIFIC ANALYSIS ===\n")
        print(f"Analyzed {len(results)} poems successfully")
        print("Report saved as: tajik_poetry_analysis_demo.xlsx\n")

        # Display summary for first poem
        if results:
            first = results[0]
            analysis = first["analysis"]
            validation = first["validation"]

            print(f"=== POEM: {first['title']} ===\n")

            # Structural
            print("STRUCTURAL ANALYSIS:")
            print(f"  - Lines: {analysis.structural.lines}")
            print(f"  - Meter: {analysis.structural.aruz_analysis.identified_meter}")
            print(f"  - Confidence: {analysis.structural.meter_confidence.value}")
            print(f"  - Rhyme Pattern: {analysis.structural.rhyme_pattern}")
            print(f"  - Stanza Form: {analysis.structural.stanza_structure}")
            print(f"  - Prosodic Consistency: {analysis.structural.prosodic_consistency:.2f}\n")

            # Content
            print("CONTENT ANALYSIS:")
            print(f"  - Lexical Diversity: {analysis.content.lexical_diversity:.3f}")
            print(f"  - Stylistic Register: {analysis.content.stylistic_register}")
            print(f"  - Main Themes: {', '.join([k for k, v in analysis.content.theme_distribution.items() if v > 0])}")
            print(
                f"  - Neologisms: {', '.join(analysis.content.neologisms[:5])}" if analysis.content.neologisms else "  - Neologisms: None")
            print(
                f"  - Archaisms: {', '.join(analysis.content.archaisms)}" if analysis.content.archaisms else "  - Archaisms: None\n")

            # Translation Theory
            print("TRANSLATION ANALYSIS (Ette/Bachmann-Medick):")
            for zone, value in analysis.translation.translation_zones.items():
                print(f"  - {zone.value.capitalize()} Zone: {value:.2f}")
            print(f"  - Cultural Markers: {len(analysis.translation.cultural_markers)}")
            print(f"  - Untranslatable Elements: {len(analysis.translation.untranslatable_elements)}\n")

            # Semiotic Analysis
            print("SEMIOTIC ANALYSIS (Lotman):")
            dynamics = analysis.semiotic.center_periphery_dynamics
            print(f"  - Center: {dynamics.get('center', 0):.2f}")
            print(f"  - Periphery: {dynamics.get('periphery', 0):.2f}")
            print(f"  - Boundary: {dynamics.get('boundary', 0):.2f}")
            print(f"  - Tension Index: {dynamics.get('tension_index', 0):.2f}")
            print(f"  - Boundary Crossings: {len(analysis.semiotic.boundary_crossings)}\n")

            # Literary Assessment
            print("LITERARY ASSESSMENT:")
            print(f"  - German Perspective: {analysis.literary.german_perspective}/5")
            print(f"  - Persian Tradition: {analysis.literary.persian_tradition}/5")
            print(f"  - Tajik Elements: {analysis.literary.tajik_elements}/5")
            print(f"  - Modernist Features: {analysis.literary.modernist_features}/5")
            print(f"  - Postcolonial Markers: {analysis.literary.postcolonial_markers}/5\n")

            # Quality Validation
            print("QUALITY METRICS:")
            print(f"  - Quality Score: {validation['quality_score']}")
            print(f"  - Reliability: {validation['reliability']}")
            if validation['warnings']:
                print("  - Warnings:")
                for warning in validation['warnings']:
                    print(f"    * {warning}")
            if validation['recommendations']:
                print("  - Recommendations:")
                for rec in validation['recommendations']:
                    print(f"    * {rec}")

        print("\n=== ANALYSIS COMPLETE ===")
        print("\nThe analyzer successfully integrates:")
        print("1. Classical Aruz prosody analysis")
        print("2. Advanced phonetic transcription")
        print("3. Translation theory (Ette/Bachmann-Medick)")
        print("4. Semiotic analysis (Lotman)")
        print("5. Multi-perspective literary assessment")
        print("6. Scientific quality validation")
        print("\nAll results are saved in the Excel report with detailed breakdowns.")

    except Exception as e:
        logger.error(f"Error in main: {e}")
        import traceback
        traceback.print_exc()


if __name__ == "__main__":
    main()

