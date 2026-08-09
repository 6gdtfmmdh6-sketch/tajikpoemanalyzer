#!/usr/bin/env python3
"""
Tajik Poetry Analyzer - Merged Enhanced Version
Scientific Research Grade with Proper ʿArūḍ Analysis

This implementation combines:
1. Enhanced ʿArūḍ (Classical Arabic-Persian prosody) analysis with 16 meters and Zihāfāt (metrical variations)
2. Phonetic-based rhyme detection (Qāfiyeh/Radīf)
3. Accurate syllable weight calculation
4. Scientific error handling and validation
5. Excel report generation
6. Free verse detection and modern metrics
7. Corpus contribution capabilities

Merged from analyzer.py and analyzer2.py enhancements.
"""

import re
import json
import logging
import unicodedata
import statistics
from collections import Counter, defaultdict
from dataclasses import dataclass, field, asdict
from pathlib import Path
from typing import Dict, List, Optional, Tuple, Any, Set, Union
from enum import Enum
from datetime import datetime

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.chart import BarChart, Reference, Series

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

# =============================================================================
# POEM PREPROCESSOR - Title/Dedication Detection
# =============================================================================

class PoemPreprocessor:
    """
    Preprocessor to clean poems before analysis.
    Detects and removes titles, dedications, and epigraphs.
    """
    
    # Patterns for title detection
    TITLE_PATTERNS = [
        r'^[А-ЯҒӢҚӮҲҶЁЪ\s\-–—]+$',  # All caps Cyrillic (with special chars)
        r'^[\d]+[\.\)]\s*',           # Numbered: "1. " or "1)"
        r'^\*{3,}$',                   # Stars
        r'^-{3,}$',                    # Dashes
        r'^={3,}$',                    # Equals
        r'^#{1,3}\s+',                 # Markdown headers
    ]
    
    # Words that indicate dedication/epigraph
    DEDICATION_MARKERS = [
        'ба ', 'барои ', 'тақдим ', 'бахшида ', 'ёдгор',
        'ба ёди ', 'ба муносибати ', 'пешкаш ',
        'Ba ', 'Baroi ', 'Taqdim ', 'Peshkash '
    ]
    
    # Epigraph indicators
    EPIGRAPH_PATTERNS = [
        r'^\s*\(.*\)\s*$',           # In parentheses
        r'^\s*\[.*\]\s*$',           # In brackets  
        r'^\s*«.*»\s*$',             # In guillemets
        r'^\s*".*"\s*$',             # In quotes
        r'^\s*_.+_\s*$',             # Italic markers
        r'^—\s*.+$',                 # Starting with em-dash (attribution)
    ]
    
    def __init__(self):
        self.logger = logging.getLogger(f"{__name__}.PoemPreprocessor")
    
    def is_title_line(self, line: str, position: int = 0) -> bool:
        """
        Detect if a line is a title or heading.
        
        Criteria:
        1. Matches title patterns (all caps, numbered, etc.)
        2. First line AND short without punctuation
        3. Title case and very short
        """
        line = line.strip()
        
        if not line:
            return False
        
        # Check regex patterns
        for pattern in self.TITLE_PATTERNS:
            if re.match(pattern, line):
                return True
        
        # First line heuristics
        if position == 0:
            words = line.split()
            word_count = len(words)
            
            # All uppercase and short (1-5 words)
            if line.isupper() and word_count <= 5:
                return True
            
            # Very short (1-3 words) without sentence-ending punctuation
            if word_count <= 3 and line[-1] not in '.!?،؛:':
                return True
            
            # Short line with no lowercase letters (likely title)
            if word_count <= 4 and not any(c.islower() for c in line if c.isalpha()):
                return True
        
        # Second line can also be title if very short
        if position == 1:
            words = line.split()
            if len(words) <= 2 and line.isupper():
                return True
        
        return False
    
    def is_dedication_line(self, line: str) -> bool:
        """Detect dedication lines"""
        line_stripped = line.strip()
        
        if not line_stripped:
            return False
        
        # Starts with dedication marker
        for marker in self.DEDICATION_MARKERS:
            if line_stripped.lower().startswith(marker.lower()):
                # Check if it's reasonably short (likely dedication)
                if len(line_stripped.split()) <= 12:
                    return True
        
        return False
    
    def is_epigraph_line(self, line: str) -> bool:
        """Detect epigraph/motto lines"""
        line_stripped = line.strip()
        
        if not line_stripped:
            return False
        
        for pattern in self.EPIGRAPH_PATTERNS:
            if re.match(pattern, line_stripped):
                return True
        
        return False
    
    def is_metadata_line(self, line: str, position: int) -> bool:
        """Check if line is any type of metadata (title, dedication, epigraph)"""
        return (self.is_title_line(line, position) or 
                self.is_dedication_line(line) or 
                self.is_epigraph_line(line))
    
    def extract_poem_body(self, text: str) -> dict:
        """
        Extract the actual poem body, separating metadata.
        
        Returns:
            dict with title, dedication, epigraph, body_lines, etc.
        """
        all_lines = [l for l in text.split('\n')]  # Keep empty lines for structure
        non_empty_lines = [(i, l.strip()) for i, l in enumerate(all_lines) if l.strip()]
        
        if not non_empty_lines:
            return {
                'title': '',
                'dedication': '',
                'epigraph': '',
                'body_lines': [],
                'body': '',
                'original': text,
                'metadata_lines': [],
                'effective_line_count': 0
            }
        
        title = ''
        dedication = ''
        epigraph = ''
        metadata_indices = set()
        
        # Check first few lines for metadata
        check_limit = min(4, len(non_empty_lines))
        
        for pos in range(check_limit):
            orig_idx, line = non_empty_lines[pos]
            
            if pos == 0 and self.is_title_line(line, pos):
                title = line
                metadata_indices.add(orig_idx)
                self.logger.debug(f"Detected title: '{line}'")
                
            elif self.is_dedication_line(line):
                dedication = line
                metadata_indices.add(orig_idx)
                self.logger.debug(f"Detected dedication: '{line}'")
                
            elif self.is_epigraph_line(line):
                epigraph = line
                metadata_indices.add(orig_idx)
                self.logger.debug(f"Detected epigraph: '{line}'")
                
            elif pos > 0 and self.is_title_line(line, pos):
                # Secondary title/subtitle
                if not title:
                    title = line
                else:
                    title += ' — ' + line
                metadata_indices.add(orig_idx)
        
        # Extract body (non-metadata lines)
        body_lines = []
        for i, line in enumerate(all_lines):
            if i not in metadata_indices and line.strip():
                body_lines.append(line.strip())
        
        return {
            'title': title,
            'dedication': dedication,
            'epigraph': epigraph,
            'body_lines': body_lines,
            'body': '\n'.join(body_lines),
            'original': text,
            'original_line_count': len([l for l in all_lines if l.strip()]),
            'effective_line_count': len(body_lines),
            'metadata_lines': list(metadata_indices),
            'lines_removed': len(metadata_indices)
        }
    
    def preprocess(self, text: str) -> dict:
        """Main preprocessing entry point"""
        return self.extract_poem_body(text)


# =============================================================================
# ENUMS
# =============================================================================

class SyllableWeight(Enum):
    """Prosodic weight classification for output"""
    HEAVY = "—"  # Long syllable
    LIGHT = "∪"  # Short syllable
    ANCEPS = "×"  # Variable weight
    UNKNOWN = "?"  # Uncertain weight

class SyllableType(Enum):
    """Internal type of syllable based on weight and context for prosody"""
    LIGHT = "light"      # CV (short vowel, open syllable)
    HEAVY = "heavy"      # CVV (long vowel/diphthong) or CVC (closed syllable)
    ANCEPS = "anceps"    # Can be either Light or Heavy depending on context/foot position
    UNKNOWN = "unknown"  # Could not be determined

class MeterConfidence(Enum):
    """Confidence levels for meter identification"""
    HIGH = "high"  # >90% pattern match
    MEDIUM = "medium"  # 70-90% pattern match
    LOW = "low"  # 50-70% pattern match
    NONE = "none"  # <50% pattern match

class FootType(Enum):
    """
    The fundamental metrical feet in classical ʿArūḍ prosody.
    
    Pattern notation: — = Heavy (ثقيل), ∪ = Light (خفيف)
    
    The 8 primary feet (الأصل الثمانية):
    - 2 five-letter feet (الخماسية)
    - 6 seven-letter feet (السباعية)
    """
    # === FIVE-LETTER FEET (الخماسية) ===
    FAULUN = "fa'ūlun"              # فَعُولُنْ      ∪ — —
    FAILUN = "fā'ilun"              # فَاعِلُنْ      — ∪ —
    
    # === SEVEN-LETTER FEET (السباعية) ===
    MAFALIYUN = "mafā'īlun"         # مَفَاعِيلُنْ    ∪ — — —
    FAILATUN = "fā'ilātun"          # فَاعِلَاتُنْ    — ∪ — —
    MUSTAFILUN = "mustaf'ilun"      # مُسْتَفْعِلُنْ   — — ∪ —
    MUFAWALATUN = "mufā'alatun"     # مُفَاعَلَتُنْ    ∪ — ∪ —
    MUTAFALIYUN = "mutafā'ilun"     # مُتَفَاعِلُنْ    ∪ ∪ — —
    MAFULATU = "maf'ūlātu"          # مَفْعُولَاتُ    — — — ∪

class ZihafType(Enum):
    """Types of metrical variations (زحافات) in ʿArūḍ"""
    KHABN = "khabn"          # خَبْن - dropping 2nd letter
    TAYY = "tayy"            # طَيّ - dropping 4th letter
    QABD = "qabd"            # قَبْض - dropping 5th letter
    KAFF = "kaff"            # كَفّ - dropping 7th letter
    WAQF = "waqf"            # وَقْف - making final letter quiescent
    QASR = "qasr"            # قَصْر - shortening
    HADHF = "hadhf"          # حَذْف - elision
    QATA = "qata"            # قَطْع - cutting
    BATR = "batr"            # بَتْر - amputation
    KHABAN_TAYY = "khaban_tayy"  # خَبْن + طَيّ combined

class StanzaForm(Enum):
    """Classical Persian/Tajik poetic forms"""
    GHAZAL = "ghazal"
    QASIDA = "qasida"
    QASIDA_WITH_RADIF = "qasida_with_radif"
    GHAZAL_WITH_RADIF = "ghazal_with_radif"
    RUBAI = "rubai"
    MASNAVI = "masnavi"
    QITA = "qita"
    MUKHAMMAS = "mukhammas"
    MUSADDAS = "musaddas"
    TARJI_BAND = "tarji_band"
    TARKIB_BAND = "tarkib_band"
    MUSTAZAD = "mustazad"
    FREE_VERSE = "free_verse"
    HAIKU = "haiku"
    UNKNOWN = "unknown"

# =============================================================================
# DATACLASSES
# =============================================================================

@dataclass
class AnalysisConfig:
    """Configuration for poetry analysis"""
    lexicon_path: str = 'data/tajik_lexicon.json'
    min_poem_length: int = 10
    max_neologisms: int = 10
    min_title_length: int = 3
    max_title_length: int = 50
    high_confidence_threshold: float = 0.90
    medium_confidence_threshold: float = 0.70
    low_confidence_threshold: float = 0.50

    # Extended theme taxonomy
    themes: Dict[str, List[str]] = field(default_factory=lambda: {
        "Love": ["муҳаббат", "ишқ", "дил", "маҳбуб", "ёр", "дилбар", "ошиқ", "маъшуқ", "ҷонон", "нигор"],
        "Nature": ["дарё", "кӯҳ", "гул", "баҳор", "навбаҳор", "осмон", "офтоб", "моҳ", "ситора", "боғ", "чаман", "сарв"],
        "Homeland": ["ватан", "тоҷикистон", "чашма", "диёр", "марзу бум", "кишвар", "миллат", "халқ"],
        "Religion": ["худо", "ҷаннат", "ибодат", "намоз", "масҷид", "аллоҳ", "паёмбар", "қуръон"],
        "Mysticism": ["тариқат", "мақом", "ҳақиқат", "маърифат", "ваҳдат", "фано", "суфӣ", "дарвеш"],
        "Philosophy": ["ҳикмат", "дониш", "хирад", "ақл", "маънӣ", "ҷаҳон", "ҳастӣ", "фано", "бақо"],
        "Time": ["замон", "рӯзгор", "умр", "ҳаёт", "марг", "фардо", "дирӯз", "имрӯз"],
        "Unity": ["ваҳдат", "ҳамбастагӣ", "иттиҳод", "якдилӣ", "ҳамдилӣ"]
    })


@dataclass
class PoemData:
    """Data structure for a single poem"""
    title: str
    content: str
    poem_id: Optional[str] = None
    metadata: Optional[Dict[str, Any]] = None

@dataclass
class InternalSyllable:
    """Detailed internal structure of a syllable for prosodic analysis"""
    text: str
    onset: str          # Initial consonant cluster
    nucleus: str        # Vowel or diphthong
    coda: str           # Final consonant cluster
    is_long_vowel: bool
    is_diphthong: bool
    is_closed: bool     # Has coda consonant(s)
    
    @property
    def weight(self) -> SyllableType:
        """Determine syllable weight based on structure"""
        if self.is_closed or self.is_long_vowel or self.is_diphthong:
            return SyllableType.HEAVY
        return SyllableType.LIGHT

@dataclass
class ProsodicSyllable:
    """Represents a syllable with prosodic information"""
    text: str
    weight: SyllableWeight
    phonetic: Optional[str] = None
    position: int = 0
    confidence: float = 1.0
    stress_level: int = 0

@dataclass
class FootVariant:
    """Represents a specific realization of a foot type"""
    name: str                           # e.g., "fa'ūlun (standard)"
    pattern: List[SyllableType]         # e.g., [LIGHT, HEAVY, HEAVY]
    foot_type: FootType
    is_standard: bool = True            # False if it's a zihāf variant
    zihaf_applied: Optional[ZihafType] = None
    
    def pattern_string(self) -> str:
        """Return pattern as string of symbols"""
        symbols = {
            SyllableType.LIGHT: "∪",
            SyllableType.HEAVY: "—",
            SyllableType.ANCEPS: "×",
            SyllableType.UNKNOWN: "?"
        }
        return " ".join(symbols.get(s, "?") for s in self.pattern)
    
    def __len__(self) -> int:
        return len(self.pattern)

@dataclass
class Zihaf:
    """A permitted metrical variation in ʿArūḍ"""
    zihaf_type: ZihafType
    name: str
    arabic_name: str
    description: str
    applicable_feet: List[FootType]
    transformation: Dict[FootType, List[SyllableType]]  # Original foot -> modified pattern
    
    def apply(self, foot_type: FootType) -> Optional[List[SyllableType]]:
        """Apply this zihāf to get the modified pattern"""
        return self.transformation.get(foot_type)

@dataclass
class MeterDefinition:
    """Defines a classical ʿArūḍ meter (بحر)"""
    name: str                                       # Arabic/Persian name
    transliteration: str                            # Latin transliteration
    description: str                                # Arabic script formula
    base_feet: Tuple[FootType, ...]                # Sequence of feet in one hemistich
    full_pattern: List[SyllableType]               # Full syllable pattern
    circle: str                                     # Which دائرة it belongs to
    frequency_weight: float = 1.0                   # How common in Tajik poetry
    
    def pattern_string(self) -> str:
        """Return full pattern as string of symbols"""
        symbols = {
            SyllableType.LIGHT: "∪",
            SyllableType.HEAVY: "—",
            SyllableType.ANCEPS: "×"
        }
        return " ".join(symbols.get(s, "?") for s in self.full_pattern)

@dataclass
class MatchedMeterResult:
    """Result of matching a parsed pattern to a meter definition"""
    meter: MeterDefinition
    confidence: MeterConfidence
    accuracy_score: float                   # 0.0 to 1.0
    variations_used: List[str]              # Names of zihāfāt applied
    matched_feet: List[FootVariant]         # The actual feet that matched
    alignment: List[Tuple[int, int]]        # (syllable_index, foot_index) pairs

@dataclass
class PhoneticAnalysis:
    """Results of phonetic analysis"""
    phonetic_transcription: str
    syllable_boundaries: List[int]
    stress_pattern: List[int]
    confidence: float
    phoneme_inventory: Dict[str, int] = field(default_factory=dict)
    syllable_count: int = 0

@dataclass
class RhymeAnalysis:
    """Advanced rhyme analysis results"""
    qafiya: str                     # The rhyming consonant(s) + vowel (قافیه)
    radif: str                      # Repeated refrain after rhyme (ردیف)
    rawi: str                       # The main rhyme consonant (روی)
    phonetic_rhyme: str             # Phonetic representation
    rhyme_type: str                 # perfect, imperfect, eye-rhyme, etc.
    rhyme_position: str = "end"
    confidence: float = 0.0

@dataclass
class RadifAnalysis:
    """Results of Radīf (refrain) detection"""
    radif_present: bool
    radif_text: str
    radif_words: List[str]          # Individual words in radif
    radif_frequency: float          # Percentage of lines with radif
    qafiya_pattern: str             # The qāfiya pattern (e.g., "-ون")
    lines_with_radif: List[int]     # Line indices with radif
    lines_without_radif: List[int]  # Line indices without radif
    cleaned_lines: List[str] = field(default_factory=list)  # Lines with radif removed

@dataclass
class AruzAnalysis:
    """Results of ʿArūḍ meter analysis - REVISED"""
    identified_meter: str
    meter_arabic: str                           # Arabic script name
    pattern_match: str                          # Foot sequence description
    full_pattern: str                           # Complete syllable pattern
    confidence: MeterConfidence
    pattern_accuracy: float
    variations_detected: List[str]              # Zihāfāt used
    line_scansion: List[ProsodicSyllable]       # Detailed syllable breakdown
    caesura_positions: List[int] = field(default_factory=list)
    feet_breakdown: List[FootVariant] = field(default_factory=list)
    alternative_meters: List[Tuple[str, float]] = field(default_factory=list)

@dataclass
class StructuralAnalysis:
    """Enhanced structural analysis results - REVISED"""
    lines: int
    syllables_per_line: List[int]
    syllable_patterns: List[List[ProsodicSyllable]]
    aruz_analysis: AruzAnalysis
    rhyme_scheme: List[RhymeAnalysis]
    rhyme_pattern: str
    radif_analysis: RadifAnalysis
    stanza_structure: StanzaForm
    avg_syllables: float
    syllable_std_dev: float                     # Standard deviation
    prosodic_consistency: float
    meter_confidence: MeterConfidence

@dataclass
class ContentAnalysis:
    """Content analysis results including lexical features"""
    word_frequencies: List[Tuple[str, int]]
    neologisms: List[str]
    archaisms: List[str]
    theme_distribution: Dict[str, int]
    primary_theme: str
    lexical_diversity: float                    # Type-token ratio
    stylistic_register: str                     # formal, informal, archaic, etc.
    total_words: int
    unique_words: int
    persian_arabic_ratio: float                 # Ratio of Persian/Arabic loanwords

@dataclass
class LiteraryAssessment:
    """Multi-perspective literary assessment"""
    classical_conformity: float         # How well it follows classical rules
    german_perspective: int             # 1-10 scale
    persian_tradition: int              # 1-10 scale
    tajik_elements: int                 # 1-10 scale
    modernist_features: int             # 1-10 scale
    overall_quality: float              # Weighted average

@dataclass
class ModernVerseMetrics:
    """Metrics for modern/free verse poetry"""
    enjambement_count: int = 0
    enjambement_ratio: float = 0.0
    semantic_density: float = 0.0  # Words per line
    line_length_variation: float = 0.0  # CV of syllables per line
    prose_poetry_score: float = 0.0
    visual_structure_score: float = 0.0
    caesura_distribution: List[int] = field(default_factory=list)
    syntactic_parallelism: float = 0.0
    lexical_repetition_score: float = 0.0
    breath_group_length: float = 0.0  # Average sentence length in words
    pause_frequency: float = 0.0  # Punctuation per line

@dataclass
class EnhancedStructuralAnalysis(StructuralAnalysis):
    """Enhanced structural analysis with modern metrics"""
    modern_metrics: Optional[ModernVerseMetrics] = None
    is_free_verse: bool = False
    free_verse_confidence: float = 0.0
    modern_features: Dict[str, float] = field(default_factory=dict)

@dataclass
class ComprehensiveAnalysis:
    """Complete analysis results"""
    structural: EnhancedStructuralAnalysis
    content: ContentAnalysis
    literary: LiteraryAssessment
    quality_metrics: Dict[str, Any]
    corpus_ready: bool = False
    contribution_id: Optional[str] = None


# =============================================================================
# SYLLABLE CONSISTENCY VALIDATOR
# =============================================================================

class SyllableConsistencyValidator:
    """Validate syllable consistency for meter assignment"""
    
    # Expected syllable counts for each meter (min, max, typical)
    # Based on standard Persian/Tajik versification
    METER_SYLLABLE_SPECS = {
        # Circle 1: المختلف
        'ṭawīl': {'min': 12, 'max': 16, 'typical': 14, 'feet': 4},
        'madīd': {'min': 9, 'max': 13, 'typical': 11, 'feet': 3},
        'basīṭ': {'min': 12, 'max': 16, 'typical': 14, 'feet': 4},
        
        # Circle 2: المؤتلف  
        'wāfir': {'min': 11, 'max': 15, 'typical': 13, 'feet': 4},
        'kāmil': {'min': 11, 'max': 15, 'typical': 13, 'feet': 4},
        
        # Circle 3: المجتلب
        'hazaj': {'min': 14, 'max': 18, 'typical': 16, 'feet': 4},
        'rajaz': {'min': 14, 'max': 18, 'typical': 16, 'feet': 4},
        'ramal': {'min': 13, 'max': 17, 'typical': 15, 'feet': 4},
        
        # Circle 4: المشتبه
        'sarīʿ': {'min': 10, 'max': 14, 'typical': 12, 'feet': 3},
        'munsariḥ': {'min': 10, 'max': 14, 'typical': 12, 'feet': 3},
        'khafīf': {'min': 13, 'max': 17, 'typical': 15, 'feet': 3},
        'muḍāriʿ': {'min': 10, 'max': 14, 'typical': 12, 'feet': 3},
        'muqtaḍab': {'min': 8, 'max': 12, 'typical': 10, 'feet': 2},
        'mujtath': {'min': 10, 'max': 14, 'typical': 12, 'feet': 3},
        
        # Circle 5: المتفق
        'mutaqārib': {'min': 14, 'max': 18, 'typical': 16, 'feet': 4},
        'mutadārik': {'min': 14, 'max': 18, 'typical': 16, 'feet': 4},
        
        # Special
        'free_verse': {'min': 3, 'max': 30, 'typical': 10, 'feet': 0},
        'haiku': {'min': 10, 'max': 20, 'typical': 17, 'feet': 0},
    }
    
    # Thresholds
    MAX_CV_CLASSICAL = 0.15      # 15% CV for classical meter
    MAX_CV_SEMI_FREE = 0.25     # 25% CV for semi-regular
    MAX_CV_FREE_VERSE = 0.35    # 35%+ suggests free verse
    
    def __init__(self):
        self.logger = logging.getLogger(f"{__name__}.SyllableValidator")
    
    def calculate_stats(self, syllable_counts: List[int]) -> dict:
        """Calculate comprehensive syllable statistics"""
        if not syllable_counts:
            return self._empty_stats()
        
        if len(syllable_counts) == 1:
            return {
                'count': 1,
                'mean': syllable_counts[0],
                'median': syllable_counts[0],
                'std_dev': 0,
                'cv': 0,
                'min': syllable_counts[0],
                'max': syllable_counts[0],
                'range': 0,
                'mode': syllable_counts[0],
                'outliers': []
            }
        
        n = len(syllable_counts)
        mean = sum(syllable_counts) / n
        sorted_counts = sorted(syllable_counts)
        median = sorted_counts[n // 2] if n % 2 else (sorted_counts[n//2 - 1] + sorted_counts[n//2]) / 2
        
        variance = sum((x - mean) ** 2 for x in syllable_counts) / n
        std_dev = variance ** 0.5
        cv = std_dev / mean if mean > 0 else 0
        
        # Find mode
        from collections import Counter
        count_freq = Counter(syllable_counts)
        mode = count_freq.most_common(1)[0][0]
        
        # Find outliers (more than 2 std devs from mean)
        outliers = [i for i, x in enumerate(syllable_counts) if abs(x - mean) > 2 * std_dev]
        
        return {
            'count': n,
            'mean': round(mean, 2),
            'median': median,
            'std_dev': round(std_dev, 2),
            'cv': round(cv, 3),
            'min': min(syllable_counts),
            'max': max(syllable_counts),
            'range': max(syllable_counts) - min(syllable_counts),
            'mode': mode,
            'outliers': outliers
        }
    
    def _empty_stats(self) -> dict:
        return {
            'count': 0, 'mean': 0, 'median': 0, 'std_dev': 0, 'cv': 0,
            'min': 0, 'max': 0, 'range': 0, 'mode': 0, 'outliers': []
        }
    
    def validate_meter(self, meter: str, syllable_counts: List[int]) -> dict:
        """
        Validate if a meter assignment is plausible.
        """
        if not syllable_counts:
            return {
                'valid': False,
                'confidence_adjustment': 0.0,
                'issues': ['No syllable data'],
                'stats': self._empty_stats(),
                'suggested_form': 'unknown'
            }
        
        stats = self.calculate_stats(syllable_counts)
        issues = []
        confidence_adj = 1.0
        
        # === Check coefficient of variation ===
        if stats['cv'] > self.MAX_CV_FREE_VERSE:
            issues.append(f"Very high syllable variation (CV={stats['cv']:.0%}) - likely free verse")
            confidence_adj *= 0.3
        elif stats['cv'] > self.MAX_CV_SEMI_FREE:
            issues.append(f"High syllable variation (CV={stats['cv']:.0%})")
            confidence_adj *= 0.5
        elif stats['cv'] > self.MAX_CV_CLASSICAL:
            issues.append(f"Moderate syllable variation (CV={stats['cv']:.0%})")
            confidence_adj *= 0.7
        
        # === Check against meter specifications ===
        meter_lower = meter.lower().replace('ʿ', '').replace('ā', 'a').replace('ī', 'i').replace('ū', 'u')
        
        # Try to find matching meter spec
        matched_spec = None
        for spec_name, spec in self.METER_SYLLABLE_SPECS.items():
            spec_normalized = spec_name.lower().replace('ʿ', '').replace('ā', 'a').replace('ī', 'i').replace('ū', 'u')
            if spec_normalized in meter_lower or meter_lower in spec_normalized:
                matched_spec = (spec_name, spec)
                break
        
        if matched_spec and meter != 'unknown':
            spec_name, spec = matched_spec
            
            # Check if mean is in expected range
            if stats['mean'] < spec['min'] - 2:
                issues.append(f"Average syllables ({stats['mean']:.1f}) below expected for {spec_name} ({spec['min']}-{spec['max']})")
                confidence_adj *= 0.6
            elif stats['mean'] > spec['max'] + 2:
                issues.append(f"Average syllables ({stats['mean']:.1f}) above expected for {spec_name} ({spec['min']}-{spec['max']})")
                confidence_adj *= 0.6
        
        # === Check for outliers ===
        if stats['outliers']:
            if len(stats['outliers']) > len(syllable_counts) * 0.2:
                issues.append(f"Many outlier lines ({len(stats['outliers'])} of {len(syllable_counts)})")
                confidence_adj *= 0.7
        
        # === Check range ===
        if stats['range'] > 6:
            issues.append(f"Large syllable range ({stats['min']}-{stats['max']})")
            confidence_adj *= 0.8
        
        # === Determine suggested form ===
        if stats['cv'] > self.MAX_CV_FREE_VERSE:
            suggested_form = 'free_verse'
        elif stats['cv'] > self.MAX_CV_SEMI_FREE:
            suggested_form = 'semi_regular'
        else:
            suggested_form = 'classical'
        
        return {
            'valid': len(issues) == 0,
            'confidence_adjustment': round(confidence_adj, 2),
            'issues': issues,
            'stats': stats,
            'suggested_form': suggested_form
        }
    
    def find_best_meter_match(self, syllable_counts: List[int]) -> List[tuple]:
        """
        Find meters that best match the syllable pattern.
        Returns list of (meter_name, match_score) tuples.
        """
        if not syllable_counts:
            return []
        
        stats = self.calculate_stats(syllable_counts)
        mean = stats['mean']
        
        matches = []
        for meter_name, spec in self.METER_SYLLABLE_SPECS.items():
            if meter_name in ('free_verse', 'haiku'):
                continue
            
            # Calculate how well the mean matches
            if spec['min'] <= mean <= spec['max']:
                # Perfect range match
                distance = abs(mean - spec['typical'])
                score = 1.0 - (distance / 5)  # Normalize
            else:
                # Out of range
                if mean < spec['min']:
                    distance = spec['min'] - mean
                else:
                    distance = mean - spec['max']
                score = max(0, 0.5 - (distance / 10))
            
            if score > 0.3:
                matches.append((meter_name, round(score, 2)))
        
        return sorted(matches, key=lambda x: -x[1])[:5]


# =============================================================================
# PHONETICS
# =============================================================================

class PersianTajikPhonetics:
    """Comprehensive Persian/Tajik phonetic analyzer"""

    def __init__(self):
        # IPA mapping for Tajik/Persian
        self.phoneme_map = {
            # Consonants (Cyrillic)
            'б': 'b', 'п': 'p', 'т': 't', 'ҷ': 'ʤ', 'ч': 'ʧ',
            'х': 'x', 'д': 'd', 'р': 'r', 'з': 'z', 'ж': 'ʒ',
            'с': 's', 'ш': 'ʃ', 'ғ': 'ʁ', 'ф': 'f', 'қ': 'q',
            'к': 'k', 'г': 'g', 'л': 'l', 'м': 'm', 'н': 'n',
            'в': 'v', 'ҳ': 'h', 'й': 'j',
            # Vowels (Cyrillic)
            'а': 'a', 'о': 'ɔ', 'у': 'u', 'э': 'e', 'и': 'i',
            'ӣ': 'iː', 'ӯ': 'ɵː', 'я': 'ja', 'ю': 'ju', 'ё': 'jɔ',
            'е': 'e',
            # Arabic script consonants
            'ب': 'b', 'پ': 'p', 'ت': 't', 'ث': 's', 'ج': 'ʤ', 'چ': 'ʧ',
            'ح': 'ħ', 'خ': 'x', 'д': 'd', 'ذ': 'z', 'ر': 'r', 'ز': 'z',
            'ژ': 'ʒ', 'س': 's', 'ش': 'ʃ', 'ص': 's', 'ض': 'z', 'ط': 't',
            'ظ': 'z', 'ع': 'ʔ', 'غ': 'ɣ', 'ф': 'f', 'ق': 'q', 'ک': 'k',
            'گ': 'g', 'ل': 'l', 'м': 'm', 'н': 'n', 'و': 'w', 'ه': 'h',
            'ی': 'j',
        }

        self.vowels = set('аоуэиӣӯяюёеАОУЭИӢӮЯЮЁЕ')
        self.long_vowels = set('ӣӯӢӮ')
        self.short_vowels = {'а', 'е', 'и', 'о', 'у', 'э', 'А', 'Е', 'И', 'О', 'У', 'Э'}
        # Complete Tajik consonant set (lowercase + uppercase)
        self.consonants = set('бвгғджзйклмнпрстфхҳчҷшъьқБВГҒДЖЗЙКЛМНПРСТФХҲЧҶШЪЬҚ')
        self.sonorous = set('рлмнвйРЛМНВЙ')
        self.diphthongs = {'ай', 'ой', 'уй', 'ей', 'ӯй', 'ав', 'ов', 'Ай', 'Ой', 'Уй', 'Ей', 'Ӯй', 'Ав', 'Ов', 'Ев'}
        self.diphthongs_ipa = {'aj', 'aw', 'oj', 'ej'}

    def analyze_phonetics(self, text: str) -> PhoneticAnalysis:
        """Complete phonetic analysis"""
        text = unicodedata.normalize('NFC', text.lower())

        # Generate phonetic transcription
        phonetic = self._to_ipa(text)

        # Find syllable boundaries
        syllables = self._syllabify(text)
        boundaries = [s[0] for s in syllables] + [len(text)] if syllables else []

        # Determine stress pattern
        stress_pattern = self._determine_stress(syllables)

        # Count phonemes
        phoneme_inventory = Counter(phonetic)

        return PhoneticAnalysis(
            phonetic_transcription=phonetic,
            syllable_boundaries=boundaries,
            stress_pattern=stress_pattern,
            phoneme_inventory=dict(phoneme_inventory),
            confidence=0.85,
            syllable_count=len(syllables)
        )

    def to_phonetic(self, text: str) -> PhoneticAnalysis:
        """Alias for analyze_phonetics for compatibility"""
        return self.analyze_phonetics(text)

    def _to_ipa(self, text: str) -> str:
        """Convert to IPA transcription"""
        result = []
        i = 0
        while i < len(text):
            char = text[i]
            
            # Check for diphthongs first (Cyrillic)
            if i + 1 < len(text):
                digraph = text[i:i+2]
                if digraph in self.diphthongs:
                    if digraph.endswith('й'):
                        result.append(self.phoneme_map.get(digraph[0], digraph[0]) + 'j')
                    elif digraph.endswith('в'):
                        result.append(self.phoneme_map.get(digraph[0], digraph[0]) + 'w')
                    i += 2
                    continue
            
            if char in self.phoneme_map:
                result.append(self.phoneme_map[char])
            elif char.isspace():
                result.append(' ')
            else:
                result.append(char)
            i += 1
        return ''.join(result)

    def _syllabify(self, text: str) -> List[Tuple[int, str]]:
        """
        Syllabify text according to Persian/Tajik phonological rules.
        
        This implementation combines:
        1. Correct word-final consonant cluster handling
        2. Maximal Onset Principle (MOP)
        3. Sonority Sequencing Principle for cluster splitting
        """
        text = unicodedata.normalize('NFC', text)
        
        # Process word by word for correct boundary handling
        words = self._split_text_into_words(text)
        all_syllables = []
        
        for word_start, word in words:
            word_syls = self._syllabify_single_word(word.lower())
            for rel_pos, syl in word_syls:
                all_syllables.append((word_start + rel_pos, syl))
        
        return all_syllables
    
    def _split_text_into_words(self, text: str) -> List[Tuple[int, str]]:
        """Split text into words with their positions"""
        words = []
        i = 0
        while i < len(text):
            # Skip non-word characters
            while i < len(text) and (text[i].isspace() or text[i] in '.,;:!?،؛؟«»()-–—\'"'):
                i += 1
            if i >= len(text):
                break
            # Collect word
            word_start = i
            while i < len(text) and not text[i].isspace() and text[i] not in '.,;:!?،؛؟«»()-–—\'"':
                i += 1
            if i > word_start:
                words.append((word_start, text[word_start:i]))
        return words
    
    def _syllabify_single_word(self, word: str) -> List[Tuple[int, str]]:
        """Syllabify a single word"""
        if not word:
            return []
        
        all_vowels = self.vowels | {'я', 'ю', 'ё'}
        all_vowels_lower = {v.lower() for v in all_vowels}
        diphthongs_lower = {d.lower() for d in self.diphthongs}
        consonants_lower = {c.lower() for c in self.consonants}
        sonorous_lower = {s.lower() for s in self.sonorous}
        
        # Valid onset clusters in Persian/Tajik (limited set)
        valid_onset_clusters = {'пр', 'бр', 'тр', 'др', 'кр', 'гр', 'фр',
                                'пл', 'бл', 'кл', 'гл', 'фл', 'сл'}
        
        # Find all vowel/nucleus positions
        vowel_positions = []
        i = 0
        while i < len(word):
            # Check for diphthong first
            if i + 1 < len(word) and word[i:i+2] in diphthongs_lower:
                vowel_positions.append((i, i + 2))
                i += 2
            elif word[i] in all_vowels_lower:
                vowel_positions.append((i, i + 1))
                i += 1
            else:
                i += 1
        
        # No vowels found - return whole word as single unit
        if not vowel_positions:
            return [(0, word)] if word else []
        
        # Build syllables by determining split points between vowels
        syllables = []
        prev_end = 0
        
        for idx, (v_start, v_end) in enumerate(vowel_positions):
            is_last = (idx == len(vowel_positions) - 1)
            
            if is_last:
                # Last syllable: include everything to end of word
                syllables.append((prev_end, word[prev_end:]))
            else:
                # Find consonants between this vowel and next
                next_v_start = vowel_positions[idx + 1][0]
                consonants = word[v_end:next_v_start]
                
                # Determine split point
                split = self._find_consonant_split_point(consonants, consonants_lower, 
                                                         sonorous_lower, valid_onset_clusters)
                syl_end = v_end + split
                syllables.append((prev_end, word[prev_end:syl_end]))
                prev_end = syl_end
        
        return syllables
    
    def _find_consonant_split_point(self, consonants: str, consonants_set: set,
                                     sonorous_set: set, valid_onsets: set) -> int:
        """
        Find optimal split point in a consonant cluster between syllables.
        
        Rules:
        1. No consonants → split at 0 (next vowel starts new syllable)
        2. One consonant → goes to onset of next syllable (split at 0)
        3. Two+ consonants → apply sonority and onset constraints
        """
        n = len(consonants)
        
        if n == 0:
            return 0
        
        if n == 1:
            # Single consonant becomes onset of next syllable
            return 0
        
        if n == 2:
            c1, c2 = consonants[0], consonants[1]
            
            # Sonorous + Obstruent → sonorous stays in coda (e.g., дил-бар)
            if c1 in sonorous_set and c2 not in sonorous_set:
                return 1
            
            # Valid onset cluster → both go to next syllable
            if consonants in valid_onsets:
                return 0
            
            # Default for 2 consonants: first stays in coda
            return 1
        
        # 3+ consonants
        last_two = consonants[-2:]
        
        # Check if last two form valid onset
        if last_two in valid_onsets:
            return n - 2
        
        # Sonorous + Obstruent at end
        if consonants[-2] in sonorous_set and consonants[-1] not in sonorous_set:
            return n - 2
        
        # Default: only last consonant goes to next syllable
        return n - 1

    def _determine_stress(self, syllables: List[Tuple[int, str]]) -> List[int]:
        """Determine stress pattern (Persian typically has final stress)"""
        if not syllables:
            return []

        stress = [0] * len(syllables)
        stress[-1] = 2  # Primary stress on final syllable

        for i, (_, syl) in enumerate(syllables[:-1]):
            if any(v in syl for v in self.long_vowels):
                stress[i] = 1

        return stress

    def calculate_syllable_weight(self, syllable: str) -> SyllableWeight:
        """Calculate syllable weight for prosody"""
        internal_syl = self._decompose_syllable(syllable)
        if internal_syl is None:
            return SyllableWeight.UNKNOWN
        
        if internal_syl.weight == SyllableType.HEAVY:
            return SyllableWeight.HEAVY
        elif internal_syl.weight == SyllableType.LIGHT:
            return SyllableWeight.LIGHT
        else:
            return SyllableWeight.UNKNOWN

    def _decompose_syllable(self, syllable: str) -> Optional[InternalSyllable]:
        """
        Decompose a syllable into onset, nucleus, and coda.
        Determines if the syllable is heavy or light.
        """
        syllable = unicodedata.normalize('NFC', syllable.strip().lower())
        if not syllable:
            return None
        
        all_vowels = self.vowels | {'я', 'ю', 'ё'}
        
        # Find nucleus position
        nucleus_start = -1
        nucleus_end = -1
        nucleus = ""
        is_long = False
        is_diphthong = False
        
        i = 0
        while i < len(syllable):
            # Check for diphthong
            if i + 1 < len(syllable):
                potential = syllable[i:i+2]
                if potential in {d.lower() for d in self.diphthongs}:
                    nucleus_start = i
                    nucleus_end = i + 2
                    nucleus = potential
                    is_diphthong = True
                    break
            
            # Check for single vowel
            if syllable[i] in {v.lower() for v in all_vowels}:
                nucleus_start = i
                nucleus_end = i + 1
                nucleus = syllable[i]
                is_long = syllable[i] in {v.lower() for v in self.long_vowels}
                break
            
            i += 1
        
        if nucleus_start == -1:
            return None
        
        onset = syllable[:nucleus_start]
        coda = syllable[nucleus_end:]
        is_closed = len(coda) > 0 and all(c in {x.lower() for x in self.consonants} for c in coda)
        
        return InternalSyllable(
            text=syllable,
            onset=onset,
            nucleus=nucleus,
            coda=coda,
            is_long_vowel=is_long,
            is_diphthong=is_diphthong,
            is_closed=is_closed
        )


# =============================================================================
# CLASSICAL PHONOLOGY FOR ʿARŪḍ
# =============================================================================

class ClassicalPhonology:
    """
    Handles phonological analysis specifically for ʿArūḍ prosody.
    Determines the weight (Light, Heavy, Anceps) of syllables.
    """

    def __init__(self):
        self.phonetics = PersianTajikPhonetics()
        self.logger = logging.getLogger(f"{__name__}.ClassicalPhonology")
        
        # Vowel classifications (Cyrillic Tajik)
        self.short_vowels = set('аеиоуэ')
        self.long_vowels = set('ӣӯ')
        self.compound_vowels = {'я': 'a', 'ю': 'u', 'ё': 'o'}
        
        # Diphthongs that make syllables heavy
        self.diphthongs = {'ай', 'ой', 'уй', 'ей', 'ӯй', 'ав', 'ов'}
        
        # Consonants for coda detection (corrected: removed vowels 'е' and 'и')
        self.consonants = set('бвгғджзйклмнпрстфхҳчҷшъьқ')
        
        # Special cases where 'о' represents Persian long 'ā'
        # Expanded list covering common words with Arabic/Persian long ā
        self.long_o_words = {
            # Core words with long ā
            'ҷон', 'ҷонон', 'осмон', 'замон', 'ҷаҳон', 'ватан', 'забон',
            'дӯстон', 'шоирон', 'инсон', 'армон', 'даҳон', 'нишон',
            # Family and people
            'модар', 'подшоҳ', 'шоҳ', 'меҳмон',
            # Nature and objects
            'боғ', 'моҳ', 'гоҳ', 'коҳ', 'роҳ', 'чоҳ',
            # Abstract concepts
            'ором', 'хоб', 'ном', 'ҷом', 'бод', 'дод', 'ёд', 'шод',
            # Compound endings
            'хона', 'бона', 'донишгоҳ', 'меҳмонхона',
            # Common nouns
            'китоб', 'ҷавоб', 'савоб', 'асос', 'хос',
            # More plurals
            'занон', 'мардон', 'бачагон', 'пирон', 'ҷавонон',
            'дилон', 'гулон', 'ситорагон', 'фариштагон'
        }
        
        # Suffixes that indicate long ā in the final syllable
        # Used for dynamic detection of plurals and other derived forms
        self.long_o_suffixes = {'он', 'гон', 'ён', 'ҳо', 'гоҳ'}

    def analyze_syllable(
        self,
        syllable_text: str,
        context: Optional[Dict[str, Any]] = None
    ) -> SyllableType:
        """
        Determine the prosodic type of a syllable.
        
        Args:
            syllable_text: The syllable text
            context: Optional context with 'position_in_line', 'total_syllables', 
                    'is_final', 'word' (the full word this syllable belongs to)
        
        Returns:
            SyllableType: LIGHT, HEAVY, or ANCEPS
        """
        syllable_text = unicodedata.normalize('NFC', syllable_text.strip().lower())
        
        if not syllable_text:
            return SyllableType.UNKNOWN
        
        context = context or {}
        
        try:
            internal = self._decompose_syllable(syllable_text)
            if internal is None:
                self.logger.warning(f"Could not decompose syllable: {syllable_text}")
                return SyllableType.UNKNOWN
            
            # === Rule 1: Line-final syllable is ALWAYS Anceps in classical prosody ===
            # This is a fundamental rule of ʿArūḍ: the last syllable of a line
            # can be scanned as either heavy or light regardless of its actual weight.
            # This rule must be checked FIRST, before any other weight determination.
            if context.get('is_final', False):
                return SyllableType.ANCEPS
            
            # === Rule 2: Closed syllable (CVC+) is ALWAYS Heavy ===
            if internal.is_closed:
                return SyllableType.HEAVY
            
            # === Rule 3: Long vowel makes syllable Heavy ===
            if internal.is_long_vowel:
                return SyllableType.HEAVY
            
            # === Rule 4: Diphthong makes syllable Heavy ===
            if internal.is_diphthong:
                return SyllableType.HEAVY
            
            # === Rule 5: Check for Arabic/Persian long 'ā' represented as 'о' ===
            word = context.get('word', '')
            if 'о' in internal.nucleus:
                # Direct word match
                if word.lower() in self.long_o_words:
                    return SyllableType.HEAVY
                # Suffix-based detection for plurals and derived forms
                word_lower = word.lower()
                for suffix in self.long_o_suffixes:
                    if word_lower.endswith(suffix):
                        return SyllableType.HEAVY
            
            # === Rule 6: Open syllable with short vowel is Light ===
            return SyllableType.LIGHT
            
        except Exception as e:
            self.logger.error(f"Error analyzing syllable '{syllable_text}': {e}")
            return SyllableType.UNKNOWN

    def _decompose_syllable(self, syllable_text: str) -> Optional[InternalSyllable]:
        """Decompose syllable into components"""
        return self.phonetics._decompose_syllable(syllable_text)

    def analyze_line(self, line: str) -> List[Tuple[str, SyllableType]]:
        """
        Analyze an entire line and return syllables with their weights.
        
        Args:
            line: A single line of poetry
            
        Returns:
            List of (syllable_text, SyllableType) tuples
        """
        syllables = self.phonetics._syllabify(line)
        total = len(syllables)
        
        results = []
        for i, (_, syl_text) in enumerate(syllables):
            context = {
                'position_in_line': i,
                'total_syllables': total,
                'is_final': (i == total - 1)
            }
            weight = self.analyze_syllable(syl_text, context)
            results.append((syl_text, weight))
        
        return results


# =============================================================================
# ZIHĀFĀT (METRICAL VARIATIONS) DEFINITIONS
# =============================================================================

class ZihafatRegistry:
    """
    Registry of all permitted metrical variations (زحافات) in classical ʿArūḍ.
    
    Zihāfāt are systematic modifications to the standard feet patterns that
    are permitted within classical prosody rules.
    """
    
    def __init__(self):
        self.logger = logging.getLogger(f"{__name__}.ZihafatRegistry")
        self._zihafat: Dict[ZihafType, Zihaf] = {}
        self._build_registry()
    
    def _build_registry(self):
        """Build the complete registry of zihāfāt"""
        
        # === خَبْن (KHABN): Dropping the 2nd letter of سَبَب خَفِيف ===
        # مُسْتَفْعِلُنْ → مُتَفْعِلُنْ (— — ∪ — → — ∪ ∪ —)
        self._zihafat[ZihafType.KHABN] = Zihaf(
            zihaf_type=ZihafType.KHABN,
            name="Khabn",
            arabic_name="خَبْن",
            description="Dropping the 2nd quiescent letter of sabab khafīf",
            applicable_feet=[FootType.MUSTAFILUN, FootType.FAILATUN, FootType.MAFULATU],
            transformation={
                FootType.MUSTAFILUN: [SyllableType.HEAVY, SyllableType.LIGHT, SyllableType.LIGHT, SyllableType.HEAVY],
                FootType.FAILATUN: [SyllableType.LIGHT, SyllableType.LIGHT, SyllableType.HEAVY, SyllableType.HEAVY],
            }
        )
        
        # === طَيّ (TAYY): Dropping the 4th letter ===
        # مُسْتَفْعِلُنْ → مُسْتَعِلُنْ (— — ∪ — → — — ∪)
        self._zihafat[ZihafType.TAYY] = Zihaf(
            zihaf_type=ZihafType.TAYY,
            name="Tayy",
            arabic_name="طَيّ",
            description="Dropping the 4th quiescent letter",
            applicable_feet=[FootType.MUSTAFILUN, FootType.MAFULATU],
            transformation={
                FootType.MUSTAFILUN: [SyllableType.HEAVY, SyllableType.HEAVY, SyllableType.LIGHT],
            }
        )
        
        # === قَبْض (QABD): Dropping the 5th letter of وَتِد مَجْمُوع ===
        # فَعُولُنْ → فَعُولُ (∪ — — → ∪ — ∪)
        self._zihafat[ZihafType.QABD] = Zihaf(
            zihaf_type=ZihafType.QABD,
            name="Qabd",
            arabic_name="قَبْض",
            description="Dropping the 5th quiescent letter of watid majmū'",
            applicable_feet=[FootType.FAULUN, FootType.MAFALIYUN],
            transformation={
                FootType.FAULUN: [SyllableType.LIGHT, SyllableType.HEAVY, SyllableType.LIGHT],
                FootType.MAFALIYUN: [SyllableType.LIGHT, SyllableType.HEAVY, SyllableType.HEAVY, SyllableType.LIGHT],
            }
        )
        
        # === كَفّ (KAFF): Dropping the 7th letter ===
        # مَفَاعِيلُنْ → مَفَاعِيلُ (∪ — — — → ∪ — — ∪)
        self._zihafat[ZihafType.KAFF] = Zihaf(
            zihaf_type=ZihafType.KAFF,
            name="Kaff",
            arabic_name="كَفّ",
            description="Dropping the 7th quiescent letter",
            applicable_feet=[FootType.MAFALIYUN, FootType.FAILATUN],
            transformation={
                FootType.MAFALIYUN: [SyllableType.LIGHT, SyllableType.HEAVY, SyllableType.HEAVY, SyllableType.LIGHT],
                FootType.FAILATUN: [SyllableType.HEAVY, SyllableType.LIGHT, SyllableType.HEAVY, SyllableType.LIGHT],
            }
        )
        
        # === حَذْف (HADHF): Elision of final sabab khafīf ===
        # فَعُولُنْ → فَعُو (∪ — — → ∪ —)
        self._zihafat[ZihafType.HADHF] = Zihaf(
            zihaf_type=ZihafType.HADHF,
            name="Hadhf",
            arabic_name="حَذْف",
            description="Elision of the final sabab khafīf",
            applicable_feet=[FootType.FAULUN, FootType.MAFALIYUN, FootType.FAILATUN],
            transformation={
                FootType.FAULUN: [SyllableType.LIGHT, SyllableType.HEAVY],
                FootType.MAFALIYUN: [SyllableType.LIGHT, SyllableType.HEAVY, SyllableType.HEAVY],
                FootType.FAILATUN: [SyllableType.HEAVY, SyllableType.LIGHT, SyllableType.HEAVY],
            }
        )
        
        # === قَصْر (QASR): Shortening ===
        self._zihafat[ZihafType.QASR] = Zihaf(
            zihaf_type=ZihafType.QASR,
            name="Qasr",
            arabic_name="قَصْر",
            description="Shortening the long vowel and dropping the following consonant",
            applicable_feet=[FootType.FAULUN, FootType.FAILATUN],
            transformation={
                FootType.FAULUN: [SyllableType.LIGHT, SyllableType.LIGHT, SyllableType.HEAVY],
                FootType.FAILATUN: [SyllableType.HEAVY, SyllableType.LIGHT, SyllableType.LIGHT, SyllableType.HEAVY],
            }
        )
        
        # === Combined: خَبْن + طَيّ (KHABN + TAYY) ===
        # مُسْتَفْعِلُنْ → مُتَعِلُنْ (— — ∪ — → — ∪ ∪)
        self._zihafat[ZihafType.KHABAN_TAYY] = Zihaf(
            zihaf_type=ZihafType.KHABAN_TAYY,
            name="Khabn+Tayy",
            arabic_name="خَبْن و طَيّ",
            description="Combined khabn and tayy",
            applicable_feet=[FootType.MUSTAFILUN],
            transformation={
                FootType.MUSTAFILUN: [SyllableType.HEAVY, SyllableType.LIGHT, SyllableType.LIGHT],
            }
        )
    
    def get_zihaf(self, zihaf_type: ZihafType) -> Optional[Zihaf]:
        """Get a specific zihāf by type"""
        return self._zihafat.get(zihaf_type)
    
    def get_all_variants_for_foot(self, foot_type: FootType) -> List[Tuple[ZihafType, List[SyllableType]]]:
        """Get all possible zihāf variants for a given foot type"""
        variants = []
        for zihaf in self._zihafat.values():
            if foot_type in zihaf.applicable_feet:
                pattern = zihaf.apply(foot_type)
                if pattern:
                    variants.append((zihaf.zihaf_type, pattern))
        return variants
    
    def find_matching_zihaf(
        self,
        foot_type: FootType,
        actual_pattern: List[SyllableType]
    ) -> Optional[ZihafType]:
        """
        Find which zihāf (if any) would transform the standard foot pattern
        into the actual observed pattern.
        """
        for zihaf in self._zihafat.values():
            if foot_type in zihaf.applicable_feet:
                transformed = zihaf.apply(foot_type)
                if transformed == actual_pattern:
                    return zihaf.zihaf_type
        return None


# =============================================================================
# METER DEFINITIONS (16 CLASSICAL METERS)
# =============================================================================

class MeterRegistry:
    """
    Registry of the 16 classical ʿArūḍ meters (بحور العَروض).
    """
    
    def __init__(self):
        self.logger = logging.getLogger(f"{__name__}.MeterRegistry")
        self._meters: Dict[str, MeterDefinition] = {}
        self._build_registry()
    
    def _build_registry(self):
        """Build the complete registry of 16 classical meters"""
        
        # ===================================================================
        # CIRCLE 1: دائرة المُخْتَلِف (Circle of the Different)
        # ===================================================================
        
        # 1. الطَّوِيل (al-Ṭawīl) - "The Long"
        # فَعُولُنْ مَفَاعِيلُنْ فَعُولُنْ مَفَاعِيلُنْ
        self._meters["tawil"] = MeterDefinition(
            name="الطَّوِيل",
            transliteration="ṭawīl",
            description="فَعُولُنْ مَفَاعِيلُنْ فَعُولُنْ مَفَاعِيلُنْ",
            base_feet=(FootType.FAULUN, FootType.MAFALIYUN, FootType.FAULUN, FootType.MAFALIYUN),
            full_pattern=[
                SyllableType.LIGHT, SyllableType.HEAVY, SyllableType.HEAVY,  # فَعُولُنْ
                SyllableType.LIGHT, SyllableType.HEAVY, SyllableType.HEAVY, SyllableType.HEAVY,  # مَفَاعِيلُنْ
                SyllableType.LIGHT, SyllableType.HEAVY, SyllableType.HEAVY,  # فَعُولُنْ
                SyllableType.LIGHT, SyllableType.HEAVY, SyllableType.HEAVY, SyllableType.HEAVY,  # مَفَاعِيلُنْ
            ],
            circle="المُخْتَلِف",
            frequency_weight=1.2  # Very common in Persian/Tajik
        )
        
        # 2. المَدِيد (al-Madīd) - "The Extended"
        # فَاعِلَاتُنْ فَاعِلُنْ فَاعِلَاتُنْ
        self._meters["madid"] = MeterDefinition(
            name="المَدِيد",
            transliteration="madīd",
            description="فَاعِلَاتُنْ فَاعِلُنْ فَاعِلَاتُنْ",
            base_feet=(FootType.FAILATUN, FootType.FAILUN, FootType.FAILATUN),
            full_pattern=[
                SyllableType.HEAVY, SyllableType.LIGHT, SyllableType.HEAVY, SyllableType.HEAVY,  # فَاعِلَاتُنْ
                SyllableType.HEAVY, SyllableType.LIGHT, SyllableType.HEAVY,  # فَاعِلُنْ
                SyllableType.HEAVY, SyllableType.LIGHT, SyllableType.HEAVY, SyllableType.HEAVY,  # فَاعِلَاتُنْ
            ],
            circle="المُخْتَلِف",
            frequency_weight=0.6
        )
        
        # 3. البَسِيط (al-Basīṭ) - "The Simple"
        # مُسْتَفْعِلُنْ فَاعِلُنْ مُسْتَفْعِلُنْ فَاعِلُنْ
        self._meters["basit"] = MeterDefinition(
            name="البَسِيط",
            transliteration="basīṭ",
            description="مُسْتَفْعِلُنْ فَاعِلُنْ مُسْتَفْعِلُنْ فَاعِلُنْ",
            base_feet=(FootType.MUSTAFILUN, FootType.FAILUN, FootType.MUSTAFILUN, FootType.FAILUN),
            full_pattern=[
                SyllableType.HEAVY, SyllableType.HEAVY, SyllableType.LIGHT, SyllableType.HEAVY,  # مُسْتَفْعِلُنْ
                SyllableType.HEAVY, SyllableType.LIGHT, SyllableType.HEAVY,  # فَاعِلُنْ
                SyllableType.HEAVY, SyllableType.HEAVY, SyllableType.LIGHT, SyllableType.HEAVY,  # مُسْتَفْعِلُنْ
                SyllableType.HEAVY, SyllableType.LIGHT, SyllableType.HEAVY,  # فَاعِلُنْ
            ],
            circle="المُخْتَلِف",
            frequency_weight=1.0
        )
        
        # ===================================================================
        # CIRCLE 2: دائرة المُؤْتَلِف (Circle of the Harmonious)
        # ===================================================================
        
        # 4. الوَافِر (al-Wāfir) - "The Abundant"
        # مُفَاعَلَتُنْ مُفَاعَلَتُنْ مُفَاعَلَتُنْ
        self._meters["wafir"] = MeterDefinition(
            name="الوَافِر",
            transliteration="wāfir",
            description="مُفَاعَلَتُنْ مُفَاعَلَتُنْ مُفَاعَلَتُنْ",
            base_feet=(FootType.MUFAWALATUN, FootType.MUFAWALATUN, FootType.MUFAWALATUN),
            full_pattern=[
                SyllableType.LIGHT, SyllableType.HEAVY, SyllableType.LIGHT, SyllableType.HEAVY,  # مُفَاعَلَتُنْ
                SyllableType.LIGHT, SyllableType.HEAVY, SyllableType.LIGHT, SyllableType.HEAVY,  # مُفَاعَلَتُنْ
                SyllableType.LIGHT, SyllableType.HEAVY, SyllableType.LIGHT, SyllableType.HEAVY,  # مُفَاعَلَتُنْ
            ],
            circle="المُؤْتَلِف",
            frequency_weight=0.8
        )
        
        # 5. الكَامِل (al-Kāmil) - "The Complete"
        # مُتَفَاعِلُنْ مُتَفَاعِلُنْ مُتَفَاعِلُنْ
        self._meters["kamil"] = MeterDefinition(
            name="الكَامِل",
            transliteration="kāmil",
            description="مُتَفَاعِلُنْ مُتَفَاعِلُنْ مُتَفَاعِلُنْ",
            base_feet=(FootType.MUTAFALIYUN, FootType.MUTAFALIYUN, FootType.MUTAFALIYUN),
            full_pattern=[
                SyllableType.LIGHT, SyllableType.LIGHT, SyllableType.HEAVY, SyllableType.HEAVY,  # مُتَفَاعِلُنْ
                SyllableType.LIGHT, SyllableType.LIGHT, SyllableType.HEAVY, SyllableType.HEAVY,  # مُتَفَاعِلُنْ
                SyllableType.LIGHT, SyllableType.LIGHT, SyllableType.HEAVY, SyllableType.HEAVY,  # مُتَفَاعِلُنْ
            ],
            circle="المُؤْتَلِف",
            frequency_weight=0.9
        )
        
        # ===================================================================
        # CIRCLE 3: دائرة المُجْتَلَب (Circle of the Acquired)
        # ===================================================================
        
        # 6. الهَزَج (al-Hazaj) - "The Trembling"
        # مَفَاعِيلُنْ مَفَاعِيلُنْ مَفَاعِيلُنْ مَفَاعِيلُنْ
        self._meters["hazaj"] = MeterDefinition(
            name="الهَزَج",
            transliteration="hazaj",
            description="مَفَاعِيلُنْ مَفَاعِيلُنْ مَفَاعِيلُنْ مَفَاعِيلُنْ",
            base_feet=(FootType.MAFALIYUN, FootType.MAFALIYUN, FootType.MAFALIYUN, FootType.MAFALIYUN),
            full_pattern=[
                SyllableType.LIGHT, SyllableType.HEAVY, SyllableType.HEAVY, SyllableType.HEAVY,  # مَفَاعِيلُنْ
                SyllableType.LIGHT, SyllableType.HEAVY, SyllableType.HEAVY, SyllableType.HEAVY,  # مَفَاعِيلُنْ
                SyllableType.LIGHT, SyllableType.HEAVY, SyllableType.HEAVY, SyllableType.HEAVY,  # مَفَاعِيلُنْ
                SyllableType.LIGHT, SyllableType.HEAVY, SyllableType.HEAVY, SyllableType.HEAVY,  # مَفَاعِيلُنْ
            ],
            circle="المُجْتَلَب",
            frequency_weight=1.3  # Very common in Persian/Tajik
        )
        
        # 7. الرَّجَز (al-Rajaz) - "The Tremor"
        # مُسْتَفْعِلُنْ مُسْتَفْعِلُنْ مُسْتَفْعِلُنْ
        self._meters["rajaz"] = MeterDefinition(
            name="الرَّجَز",
            transliteration="rajaz",
            description="مُسْتَفْعِلُنْ مُسْتَفْعِلُنْ مُسْتَفْعِلُنْ",
            base_feet=(FootType.MUSTAFILUN, FootType.MUSTAFILUN, FootType.MUSTAFILUN),
            full_pattern=[
                SyllableType.HEAVY, SyllableType.HEAVY, SyllableType.LIGHT, SyllableType.HEAVY,  # مُسْتَفْعِلُنْ
                SyllableType.HEAVY, SyllableType.HEAVY, SyllableType.LIGHT, SyllableType.HEAVY,  # مُسْتَفْعِلُنْ
                SyllableType.HEAVY, SyllableType.HEAVY, SyllableType.LIGHT, SyllableType.HEAVY,  # مُسْتَفْعِلُنْ
            ],
            circle="المُجْتَلَب",
            frequency_weight=1.1
        )
        
        # 8. الرَّمَل (al-Ramal) - "The Running"
        # فَاعِلَاتُنْ فَاعِلَاتُنْ فَاعِلَاتُنْ
        self._meters["ramal"] = MeterDefinition(
            name="الرَّمَل",
            transliteration="ramal",
            description="فَاعِلَاتُنْ فَاعِلَاتُنْ فَاعِلَاتُنْ",
            base_feet=(FootType.FAILATUN, FootType.FAILATUN, FootType.FAILATUN),
            full_pattern=[
                SyllableType.HEAVY, SyllableType.LIGHT, SyllableType.HEAVY, SyllableType.HEAVY,  # فَاعِلَاتُنْ
                SyllableType.HEAVY, SyllableType.LIGHT, SyllableType.HEAVY, SyllableType.HEAVY,  # فَاعِلَاتُنْ
                SyllableType.HEAVY, SyllableType.LIGHT, SyllableType.HEAVY, SyllableType.HEAVY,  # فَاعِلَاتُنْ
            ],
            circle="المُجْتَلَб",
            frequency_weight=1.2  # Very common in Persian/Tajik
        )
        
        # ===================================================================
        # CIRCLE 4: دائرة المُشْتَبِه (Circle of the Similar)
        # ===================================================================
        
        # 9. السَّرِيع (al-Sarīʿ) - "The Swift"
        # مُسْتَفْعِلُنْ مُسْتَفْعِلُنْ مَفْعُولَاتُ
        self._meters["sari"] = MeterDefinition(
            name="السَّرِيع",
            transliteration="sarīʿ",
            description="مُسْتَفْعِلُنْ مُسْتَفْعِلُنْ مَفْعُولَاتُ",
            base_feet=(FootType.MUSTAFILUN, FootType.MUSTAFILUN, FootType.MAFULATU),
            full_pattern=[
                SyllableType.HEAVY, SyllableType.HEAVY, SyllableType.LIGHT, SyllableType.HEAVY,  # مُسْتَفْعِلُنْ
                SyllableType.HEAVY, SyllableType.HEAVY, SyllableType.LIGHT, SyllableType.HEAVY,  # مُسْتَفْعِلُنْ
                SyllableType.HEAVY, SyllableType.HEAVY, SyllableType.HEAVY, SyllableType.LIGHT,  # مَفْعُولَاتُ
            ],
            circle="المُشْتَبِه",
            frequency_weight=0.7
        )
        
        # 10. المُنْسَرِح (al-Munsarih) - "The Flowing"
        # مُسْتَفْعِلُنْ مَفْعُولَاتُ مُسْتَفْعِلُنْ
        self._meters["munsarih"] = MeterDefinition(
            name="المُنْسَرِح",
            transliteration="munsariḥ",
            description="مُسْتَفْعِلُنْ مَفْعُولَاتُ مُسْتَفْعِلُنْ",
            base_feet=(FootType.MUSTAFILUN, FootType.MAFULATU, FootType.MUSTAFILUN),
            full_pattern=[
                SyllableType.HEAVY, SyllableType.HEAVY, SyllableType.LIGHT, SyllableType.HEAVY,  # مُسْتَفْعِلُنْ
                SyllableType.HEAVY, SyllableType.HEAVY, SyllableType.HEAVY, SyllableType.LIGHT,  # مَفْعُولَاتُ
                SyllableType.HEAVY, SyllableType.HEAVY, SyllableType.LIGHT, SyllableType.HEAVY,  # مُسْتَفْعِلُنْ
            ],
            circle="المُشْتَبِه",
            frequency_weight=0.5
        )
        
        # 11. الخَفِيف (al-Khafīf) - "The Light"
        # فَاعِلَاتُنْ مُسْتَفْعِلُنْ فَاعِلَاتُنْ
        self._meters["khafif"] = MeterDefinition(
            name="الخَفِيف",
            transliteration="khafīf",
            description="فَاعِلَاتُنْ مُسْتَفْعِلُنْ فَاعِلَاتُنْ",
            base_feet=(FootType.FAILATUN, FootType.MUSTAFILUN, FootType.FAILATUN),
            full_pattern=[
                SyllableType.HEAVY, SyllableType.LIGHT, SyllableType.HEAVY, SyllableType.HEAVY,  # فَاعِلَاتُنْ
                SyllableType.HEAVY, SyllableType.HEAVY, SyllableType.LIGHT, SyllableType.HEAVY,  # مُسْتَفْعِلُنْ
                SyllableType.HEAVY, SyllableType.LIGHT, SyllableType.HEAVY, SyllableType.HEAVY,  # فَاعِلَاتُنْ
            ],
            circle="المُشْتَبِه",
            frequency_weight=0.9
        )
        
        # 12. المُضَارِع (al-Muḍāriʿ) - "The Similar"
        # مَفَاعِيلُنْ فَاعِلَاتُنْ مَفَاعِيلُنْ
        self._meters["mudari"] = MeterDefinition(
            name="المُضَارِع",
            transliteration="muḍāriʿ",
            description="مَفَاعِيلُنْ فَاعِلَاتُنْ مَفَاعِيلُنْ",
            base_feet=(FootType.MAFALIYUN, FootType.FAILATUN, FootType.MAFALIYUN),
            full_pattern=[
                SyllableType.LIGHT, SyllableType.HEAVY, SyllableType.HEAVY, SyllableType.HEAVY,  # مَفَاعِילُنْ
                SyllableType.HEAVY, SyllableType.LIGHT, SyllableType.HEAVY, SyllableType.HEAVY,  # فَاعِلَاتُنْ
                SyllableType.LIGHT, SyllableType.HEAVY, SyllableType.HEAVY, SyllableType.HEAVY,  # مَفَاعِيلُنْ
            ],
            circle="المُشْتَبِه",
            frequency_weight=0.4
        )
        
        # 13. المُقْتَضَب (al-Muqtaḍab) - "The Cut Short"
        # مَفْعُولَاتُ مُسْتَفْعِلُنْ مُسْتَفْعِلُنْ
        self._meters["muqtadab"] = MeterDefinition(
            name="المُقْتَضَب",
            transliteration="muqtaḍab",
            description="مَفْعُولَاتُ مُسْتَفْعِلُنْ مُسْتَفْعِلُنْ",
            base_feet=(FootType.MAFULATU, FootType.MUSTAFILUN, FootType.MUSTAFILUN),
            full_pattern=[
                SyllableType.HEAVY, SyllableType.HEAVY, SyllableType.HEAVY, SyllableType.LIGHT,  # مَفْعُولَاتُ
                SyllableType.HEAVY, SyllableType.HEAVY, SyllableType.LIGHT, SyllableType.HEAVY,  # مُسْتَفْعِلُنْ
                SyllableType.HEAVY, SyllableType.HEAVY, SyllableType.LIGHT, SyllableType.HEAVY,  # مُسْتَفْعِلُنْ
            ],
            circle="المُشْتَبِه",
            frequency_weight=0.3
        )
        
        # 14. المُجْتَثّ (al-Mujtathth) - "The Cut Off"
        # مُسْتَفْعِلُنْ فَاعِلَاتُنْ فَاعِلَاتُنْ
        self._meters["mujtathth"] = MeterDefinition(
            name="المُجْتَثّ",
            transliteration="mujtathth",
            description="مُسْتَفْعِلُنْ فَاعِلَاتُنْ فَاعِلَاتُنْ",
            base_feet=(FootType.MUSTAFILUN, FootType.FAILATUN, FootType.FAILATUN),
            full_pattern=[
                SyllableType.HEAVY, SyllableType.HEAVY, SyllableType.LIGHT, SyllableType.HEAVY,  # مُسْتَفْعِلُنْ
                SyllableType.HEAVY, SyllableType.LIGHT, SyllableType.HEAVY, SyllableType.HEAVY,  # فَاعِلَاتُنْ
                SyllableType.HEAVY, SyllableType.LIGHT, SyllableType.HEAVY, SyllableType.HEAVY,  # فَاعِلَاتُنْ
            ],
            circle="المُشْتَبِه",
            frequency_weight=0.4
        )
        
        # ===================================================================
        # CIRCLE 5: دائرة المُتَّفِق (Circle of the Agreeing)
        # ===================================================================
        
        # 15. المُتَقَارِب (al-Mutaqārib) - "The Approaching"
        # فَعُولُنْ فَعُولُنْ فَعُولُنْ فَعُولُنْ
        self._meters["mutaqarib"] = MeterDefinition(
            name="المُتَقَارِب",
            transliteration="mutaqārib",
            description="فَعُولُنْ فَعُولُنْ فَعُولُنْ فَعُولُنْ",
            base_feet=(FootType.FAULUN, FootType.FAULUN, FootType.FAULUN, FootType.FAULUN),
            full_pattern=[
                SyllableType.LIGHT, SyllableType.HEAVY, SyllableType.HEAVY,  # فَعُولُنْ
                SyllableType.LIGHT, SyllableType.HEAVY, SyllableType.HEAVY,  # فَعُولُنْ
                SyllableType.LIGHT, SyllableType.HEAVY, SyllableType.HEAVY,  # فَعُولُنْ
                SyllableType.LIGHT, SyllableType.HEAVY, SyllableType.HEAVY,  # فَعُولُنْ
            ],
            circle="المُتَّفِق",
            frequency_weight=1.1
        )
        
        # 16. المُتَدَارِك (al-Mutadārik) - "The Overtaking"
        # فَاعِلُنْ فَاعِلُنْ فَاعِلُنْ فَاعِلُنْ
        self._meters["mutadarik"] = MeterDefinition(
            name="المُتَدَارِك",
            transliteration="mutadārik",
            description="فَاعِلُنْ فَاعِلُنْ فَاعِلُنْ فَاعِلُنْ",
            base_feet=(FootType.FAILUN, FootType.FAILUN, FootType.FAILUN, FootType.FAILUN),
            full_pattern=[
                SyllableType.HEAVY, SyllableType.LIGHT, SyllableType.HEAVY,  # فَاعِلُنْ
                SyllableType.HEAVY, SyllableType.LIGHT, SyllableType.HEAVY,  # فَاعِلُنْ
                SyllableType.HEAVY, SyllableType.LIGHT, SyllableType.HEAVY,  # فَاعِلُنْ
                SyllableType.HEAVY, SyllableType.LIGHT, SyllableType.HEAVY,  # فَاعِلُنْ
            ],
            circle="المُتَّفِق",
            frequency_weight=0.8
        )
        
        # ===================================================================
        # COMMON PERSIAN/TAJIK VARIANTS (Truncated meters - مجزوء)
        # ===================================================================
        
        # مجزوء الهزج - Truncated Hazaj (very common in Persian/Tajik)
        # مَفَاعِيلُنْ مَفَاعِيلُنْ
        self._meters["hazaj_majzu"] = MeterDefinition(
            name="مجزوء الهَزَج",
            transliteration="hazaj majzū'",
            description="مَفَاعِيلُنْ مَفَاعِيلُنْ",
            base_feet=(FootType.MAFALIYUN, FootType.MAFALIYUN),
            full_pattern=[
                SyllableType.LIGHT, SyllableType.HEAVY, SyllableType.HEAVY, SyllableType.HEAVY,  # مَفَاعِيلُنْ
                SyllableType.LIGHT, SyllableType.HEAVY, SyllableType.HEAVY, SyllableType.HEAVY,  # مَفَاعِيلُنْ
            ],
            circle="المُجْتَلَب",
            frequency_weight=1.4  # Very common
        )
        
        # مجزوء الرمل - Truncated Ramal
        # فَاعِلَاتُنْ فَاعِلَاتُنْ
        self._meters["ramal_majzu"] = MeterDefinition(
            name="مجزوء الرَّمَل",
            transliteration="ramal majzū'",
            description="فَاعِلَاتُنْ فَاعِلَاتُنْ",
            base_feet=(FootType.FAILATUN, FootType.FAILATUN),
            full_pattern=[
                SyllableType.HEAVY, SyllableType.LIGHT, SyllableType.HEAVY, SyllableType.HEAVY,  # فَاعِلَاتُنْ
                SyllableType.HEAVY, SyllableType.LIGHT, SyllableType.HEAVY, SyllableType.HEAVY,  # فَاعِلَاتُنْ
            ],
            circle="المُجْتَلَب",
            frequency_weight=1.3
        )
        
        # مجزوء الخفيف - Truncated Khafif
        # فَاعِلَاتُنْ مُسْتَفْعِلُنْ
        self._meters["khafif_majzu"] = MeterDefinition(
            name="مجزوء الخَفِيف",
            transliteration="khafīf majzū'",
            description="فَاعِلَاتُنْ مُسْتَفْعِلُنْ",
            base_feet=(FootType.FAILATUN, FootType.MUSTAFILUN),
            full_pattern=[
                SyllableType.HEAVY, SyllableType.LIGHT, SyllableType.HEAVY, SyllableType.HEAVY,  # فَاعِلَاتُنْ
                SyllableType.HEAVY, SyllableType.HEAVY, SyllableType.LIGHT, SyllableType.HEAVY,  # مُسْتَفْعِلُنْ
            ],
            circle="المُشْتَبِه",
            frequency_weight=1.0
        )

    def get_meter(self, name: str) -> Optional[MeterDefinition]:
        """Get a meter by its transliteration name"""
        return self._meters.get(name.lower())
    
    def get_all_meters(self) -> List[MeterDefinition]:
        """Get all registered meters"""
        return list(self._meters.values())
    
    def get_meters_by_circle(self, circle: str) -> List[MeterDefinition]:
        """Get all meters belonging to a specific circle"""
        return [m for m in self._meters.values() if m.circle == circle]
    
    def get_meters_sorted_by_frequency(self) -> List[MeterDefinition]:
        """Get meters sorted by frequency weight (most common first)"""
        return sorted(self._meters.values(), key=lambda m: -m.frequency_weight)
    
# =============================================================================
# EXTENDED ZIHAFAT REGISTRY - All Classical Variations
# =============================================================================

class ExtendedZihafatRegistry:
    """
    Complete registry of Zihāfāt (metrical variations/licenses).
    
    Zihāfāt are the permissible modifications to the base feet in Arabic/Persian prosody.
    They allow flexibility while maintaining the essential character of the meter.
    """
    
    def __init__(self):
        self.logger = logging.getLogger(f"{__name__}.ExtendedZihafatRegistry")
        self.zihafat = self._build_zihafat()
        self.foot_variants = self._build_foot_variants()
    
    def _build_zihafat(self) -> Dict[str, dict]:
        """
        Build complete zihāfāt dictionary.
        
        Categories:
        1. Single-letter modifications (حذف حرف واحد)
        2. Double modifications (زحافات مركبة)
        3. Final-foot modifications (علل)
        """
        return {
            # === SINGLE MODIFICATIONS (زحافات مفردة) ===
            
            # خبن (Khabn) - Remove second letter (س from مُسْتَفْعِلُنْ → مُتَفْعِلُنْ)
            'khabn': {
                'name_ar': 'خَبْن',
                'description': 'Removal of second quiescent letter',
                'transforms': {
                    # Original → Modified
                    (SyllableType.HEAVY, SyllableType.HEAVY, SyllableType.LIGHT, SyllableType.HEAVY): 
                        (SyllableType.LIGHT, SyllableType.HEAVY, SyllableType.LIGHT, SyllableType.HEAVY),
                    # مُسْتَفْعِلُنْ → مُتَفْعِلُنْ  (- - u - → u - u -)
                    (SyllableType.HEAVY, SyllableType.HEAVY, SyllableType.LIGHT, SyllableType.HEAVY):
                        (SyllableType.LIGHT, SyllableType.HEAVY, SyllableType.LIGHT, SyllableType.HEAVY),
                    # فَاعِلَاتُنْ → فَعِلَاتُنْ (- u - - → u u - -)
                    (SyllableType.HEAVY, SyllableType.LIGHT, SyllableType.HEAVY, SyllableType.HEAVY):
                        (SyllableType.LIGHT, SyllableType.LIGHT, SyllableType.HEAVY, SyllableType.HEAVY),
                },
                'affected_feet': ['mustafʿilun', 'fāʿilātun', 'mafāʿīlun']
            },
            
            # طي (Ṭayy) - Remove fourth letter
            'tayy': {
                'name_ar': 'طَيّ',
                'description': 'Removal of fourth quiescent letter',
                'transforms': {
                    # مُسْتَفْعِلُنْ → مُسْتَعِلُنْ (- - u - → - - u u -)
                    (SyllableType.HEAVY, SyllableType.HEAVY, SyllableType.LIGHT, SyllableType.HEAVY):
                        (SyllableType.HEAVY, SyllableType.HEAVY, SyllableType.LIGHT, SyllableType.LIGHT, SyllableType.HEAVY),
                    # مَفْعُولَاتُ → مَفْعُلَاتُ
                },
                'affected_feet': ['mustafʿilun', 'mafʿūlātu']
            },
            
            # قبض (Qabḍ) - Remove fifth letter
            'qabd': {
                'name_ar': 'قَبْض',
                'description': 'Removal of fifth quiescent letter',
                'transforms': {
                    # فَعُولُنْ → فَعُولُ (u - - → u - u)
                    (SyllableType.LIGHT, SyllableType.HEAVY, SyllableType.HEAVY):
                        (SyllableType.LIGHT, SyllableType.HEAVY, SyllableType.LIGHT),
                    # مَفَاعِيلُنْ → مَفَاعِلُنْ (u - - - → u - u -)
                    (SyllableType.LIGHT, SyllableType.HEAVY, SyllableType.HEAVY, SyllableType.HEAVY):
                        (SyllableType.LIGHT, SyllableType.HEAVY, SyllableType.LIGHT, SyllableType.HEAVY),
                },
                'affected_feet': ['faʿūlun', 'mafāʿīlun']
            },
            
            # عصب (ʿAṣb) - Make fifth letter quiescent
            'asb': {
                'name_ar': 'عَصْب',
                'description': 'Making fifth letter quiescent',
                'transforms': {
                    # مُفَاعَلَتُنْ → مُفَاعَلْتُنْ (u - u u - → u - - -)
                    (SyllableType.LIGHT, SyllableType.HEAVY, SyllableType.LIGHT, SyllableType.LIGHT, SyllableType.HEAVY):
                        (SyllableType.LIGHT, SyllableType.HEAVY, SyllableType.HEAVY, SyllableType.HEAVY),
                },
                'affected_feet': ['mufāʿalatun']
            },
            
            # وقص (Waqṣ) - Remove second letter of mutafāʿilun
            'waqs': {
                'name_ar': 'وَقْص',
                'description': 'Removal of second letter from mutafāʿilun',
                'transforms': {
                    # مُتَفَاعِلُنْ → مُفَاعِلُنْ (u u - u - → u - u -)
                    (SyllableType.LIGHT, SyllableType.LIGHT, SyllableType.HEAVY, SyllableType.LIGHT, SyllableType.HEAVY):
                        (SyllableType.LIGHT, SyllableType.HEAVY, SyllableType.LIGHT, SyllableType.HEAVY),
                },
                'affected_feet': ['mutafāʿilun']
            },
            
            # إضمار (Iḍmār) - Make second letter quiescent  
            'idmar': {
                'name_ar': 'إِضْمَار',
                'description': 'Making second letter quiescent',
                'transforms': {
                    # مُتَفَاعِلُنْ → مُتْفَاعِلُنْ (u u - u - → - - u -)
                    (SyllableType.LIGHT, SyllableType.LIGHT, SyllableType.HEAVY, SyllableType.LIGHT, SyllableType.HEAVY):
                        (SyllableType.HEAVY, SyllableType.HEAVY, SyllableType.LIGHT, SyllableType.HEAVY),
                },
                'affected_feet': ['mutafāʿilun']
            },
            
            # === COMPOUND MODIFICATIONS (زحافات مركبة) ===
            
            # خبل (Khabl) = خبن + طي
            'khabl': {
                'name_ar': 'خَبْل',
                'description': 'Combination of khabn and ṭayy',
                'transforms': {
                    # مُسْتَفْعِلُنْ → مُتَعِلُنْ (- - u - → u u u -)
                    (SyllableType.HEAVY, SyllableType.HEAVY, SyllableType.LIGHT, SyllableType.HEAVY):
                        (SyllableType.LIGHT, SyllableType.LIGHT, SyllableType.LIGHT, SyllableType.HEAVY),
                },
                'affected_feet': ['mustafʿilun']
            },
            
            # شكل (Shakl) = خبن + كف
            'shakl': {
                'name_ar': 'شَكْل',
                'description': 'Combination of khabn and kaff',
                'transforms': {
                    # فَاعِلَاتُنْ → فَعِلَاتُ (- u - - → u u - u)
                    (SyllableType.HEAVY, SyllableType.LIGHT, SyllableType.HEAVY, SyllableType.HEAVY):
                        (SyllableType.LIGHT, SyllableType.LIGHT, SyllableType.HEAVY, SyllableType.LIGHT),
                },
                'affected_feet': ['fāʿilātun']
            },
            
            # خرب (Kharb) = خرم + كف
            'kharb': {
                'name_ar': 'خَرْب',
                'description': 'Combination of kharm and kaff',
                'transforms': {
                    # مَفَاعِيلُنْ → فَاعِيلُ (u - - - → - - u)
                    (SyllableType.LIGHT, SyllableType.HEAVY, SyllableType.HEAVY, SyllableType.HEAVY):
                        (SyllableType.HEAVY, SyllableType.HEAVY, SyllableType.LIGHT),
                },
                'affected_feet': ['mafāʿīlun']
            },
            
            # === FINAL-FOOT MODIFICATIONS (علل) ===
            
            # قطع (Qaṭʿ) - Cut final letter
            'qat': {
                'name_ar': 'قَطْع',
                'description': 'Cutting the final letter and making previous quiescent',
                'transforms': {
                    # فَاعِلُنْ → فَاعِلْ (- u - → - u -)
                    (SyllableType.HEAVY, SyllableType.LIGHT, SyllableType.HEAVY):
                        (SyllableType.HEAVY, SyllableType.LIGHT, SyllableType.HEAVY),
                    # مُتَفَاعِلُنْ → مُتَفَاعِلْ
                },
                'affected_feet': ['fāʿilun', 'mutafāʿilun'],
                'position': 'final'
            },
            
            # حذف (Ḥadhf) - Remove final watad
            'hadhf': {
                'name_ar': 'حَذْف',
                'description': 'Removal of final watad (two letters)',
                'transforms': {
                    # فَعُولُنْ → فَعُو (u - - → u -)
                    (SyllableType.LIGHT, SyllableType.HEAVY, SyllableType.HEAVY):
                        (SyllableType.LIGHT, SyllableType.HEAVY),
                    # مَفَاعِيلُنْ → مَفَاعِي (u - - - → u - -)
                    (SyllableType.LIGHT, SyllableType.HEAVY, SyllableType.HEAVY, SyllableType.HEAVY):
                        (SyllableType.LIGHT, SyllableType.HEAVY, SyllableType.HEAVY),
                },
                'affected_feet': ['faʿūlun', 'mafāʿīlun'],
                'position': 'final'
            },
            
            # قصر (Qaṣr) - Shortening
            'qasr': {
                'name_ar': 'قَصْر',
                'description': 'Shortening the final syllable',
                'transforms': {
                    # فَاعِلَاتُنْ → فَاعِلَاتْ (- u - - → - u - -)
                    # Already heavy, just confirms
                },
                'affected_feet': ['fāʿilātun'],
                'position': 'final'
            },
            
            # بتر (Batr) - حذف + قطع
            'batr': {
                'name_ar': 'بَتْر',
                'description': 'Combination of ḥadhf and qaṭʿ',
                'transforms': {
                    # فَعُولُنْ → فَعْ (u - - → -)
                    (SyllableType.LIGHT, SyllableType.HEAVY, SyllableType.HEAVY):
                        (SyllableType.HEAVY,),
                },
                'affected_feet': ['faʿūlun'],
                'position': 'final'
            },
            
            # تسبيغ (Tasbīgh) - Adding letter to final sabab
            'tasbigh': {
                'name_ar': 'تَسْبِيغ',
                'description': 'Adding a letter to final light sabab',
                'transforms': {
                    # فَاعِلَاتُنْ → فَاعِلَاتَانْ
                },
                'affected_feet': ['fāʿilātun'],
                'position': 'final'
            },
        }
    
    def _build_foot_variants(self) -> Dict[str, List[tuple]]:
        """
        Build all acceptable variants for each foot type.
        Returns dict mapping foot name to list of acceptable syllable patterns.
        """
        variants = {
            # فَعُولُنْ (faʿūlun) - u - -
            'faʿūlun': [
                (SyllableType.LIGHT, SyllableType.HEAVY, SyllableType.HEAVY),      # Standard
                (SyllableType.LIGHT, SyllableType.HEAVY, SyllableType.LIGHT),      # قبض (qabḍ)
                (SyllableType.LIGHT, SyllableType.HEAVY),                          # حذف (ḥadhf)
                (SyllableType.HEAVY,),                                             # بتر (batr)
            ],
            
            # مَفَاعِيلُنْ (mafāʿīlun) - u - - -
            'mafāʿīlun': [
                (SyllableType.LIGHT, SyllableType.HEAVY, SyllableType.HEAVY, SyllableType.HEAVY),  # Standard
                (SyllableType.LIGHT, SyllableType.HEAVY, SyllableType.LIGHT, SyllableType.HEAVY),  # قبض
                (SyllableType.LIGHT, SyllableType.HEAVY, SyllableType.HEAVY),                      # حذف
                (SyllableType.HEAVY, SyllableType.HEAVY, SyllableType.HEAVY),                      # خرم (kharm)
                (SyllableType.HEAVY, SyllableType.HEAVY, SyllableType.LIGHT),                      # خرب
            ],
            
            # فَاعِلَاتُنْ (fāʿilātun) - - u - -
            'fāʿilātun': [
                (SyllableType.HEAVY, SyllableType.LIGHT, SyllableType.HEAVY, SyllableType.HEAVY),  # Standard
                (SyllableType.LIGHT, SyllableType.LIGHT, SyllableType.HEAVY, SyllableType.HEAVY),  # خبن
                (SyllableType.HEAVY, SyllableType.LIGHT, SyllableType.HEAVY, SyllableType.LIGHT),  # كف (kaff)
                (SyllableType.LIGHT, SyllableType.LIGHT, SyllableType.HEAVY, SyllableType.LIGHT),  # شكل
                (SyllableType.HEAVY, SyllableType.LIGHT, SyllableType.HEAVY),                      # حذف
            ],
            
            # فَاعِلُنْ (fāʿilun) - - u -
            'fāʿilun': [
                (SyllableType.HEAVY, SyllableType.LIGHT, SyllableType.HEAVY),      # Standard
                (SyllableType.LIGHT, SyllableType.LIGHT, SyllableType.HEAVY),      # خبن
                (SyllableType.HEAVY, SyllableType.LIGHT, SyllableType.LIGHT),      # قطع → فاعلْ
                (SyllableType.HEAVY, SyllableType.HEAVY),                          # تشعيث
            ],
            
            # مُسْتَفْعِلُنْ (mustafʿilun) - - - u -
            'mustafʿilun': [
                (SyllableType.HEAVY, SyllableType.HEAVY, SyllableType.LIGHT, SyllableType.HEAVY),  # Standard
                (SyllableType.LIGHT, SyllableType.HEAVY, SyllableType.LIGHT, SyllableType.HEAVY),  # خبن
                (SyllableType.HEAVY, SyllableType.LIGHT, SyllableType.LIGHT, SyllableType.HEAVY),  # طي
                (SyllableType.LIGHT, SyllableType.LIGHT, SyllableType.LIGHT, SyllableType.HEAVY),  # خبل
                (SyllableType.HEAVY, SyllableType.HEAVY, SyllableType.HEAVY),                      # طي + قطع
            ],
            
            # مُتَفَاعِلُنْ (mutafāʿilun) - u u - u -
            'mutafāʿilun': [
                (SyllableType.LIGHT, SyllableType.LIGHT, SyllableType.HEAVY, SyllableType.LIGHT, SyllableType.HEAVY),  # Standard
                (SyllableType.LIGHT, SyllableType.HEAVY, SyllableType.LIGHT, SyllableType.HEAVY),  # وقص
                (SyllableType.HEAVY, SyllableType.HEAVY, SyllableType.LIGHT, SyllableType.HEAVY),  # إضمار
                (SyllableType.LIGHT, SyllableType.LIGHT, SyllableType.HEAVY, SyllableType.HEAVY),  # عصب
            ],
            
            # مُفَاعَلَتُنْ (mufāʿalatun) - u - u u -
            'mufāʿalatun': [
                (SyllableType.LIGHT, SyllableType.HEAVY, SyllableType.LIGHT, SyllableType.LIGHT, SyllableType.HEAVY),  # Standard
                (SyllableType.LIGHT, SyllableType.HEAVY, SyllableType.HEAVY, SyllableType.HEAVY),  # عصب
                (SyllableType.LIGHT, SyllableType.HEAVY, SyllableType.LIGHT, SyllableType.HEAVY),  # عقل
            ],
            
            # مَفْعُولَاتُ (mafʿūlātu) - - - u -
            'mafʿūlātu': [
                (SyllableType.HEAVY, SyllableType.HEAVY, SyllableType.HEAVY, SyllableType.LIGHT),  # Standard
                (SyllableType.HEAVY, SyllableType.HEAVY, SyllableType.LIGHT, SyllableType.LIGHT),  # طي
                (SyllableType.HEAVY, SyllableType.LIGHT, SyllableType.HEAVY, SyllableType.LIGHT),  # خبن
            ],
        }
        
        return variants
    
    def get_all_variants(self, foot_name: str) -> List[tuple]:
        """Get all acceptable variants for a foot"""
        foot_key = foot_name.lower().replace('ʿ', '').replace('ā', 'a').replace('ī', 'i').replace('ū', 'u')
    
        # Direct lookup in foot_variants dictionary
        if foot_key in self.foot_variants:
            return self.foot_variants[foot_key]
        
        # If not found, try partial matching
        for key in self.foot_variants:
            # Try to match without diacritics or with partial name
            clean_key = key.replace('ʿ', '').replace('ā', 'a').replace('ī', 'i').replace('ū', 'u')
            if clean_key == foot_key or foot_key in clean_key or clean_key in foot_key:
                return self.foot_variants[key]
        
        # Return empty list if no match found
        self.logger.warning(f"No variants found for foot: {foot_name} (key: {foot_key})")
        return []


# =============================================================================
# FOOT VARIANT BUILDER
# =============================================================================

class FootVariantBuilder:
    """
    Builds all possible foot variants including standard forms and zihāf variations.
    """
    
    def __init__(self):
        self.logger = logging.getLogger(f"{__name__}.FootVariantBuilder")
        self.zihafat = ZihafatRegistry()
        self._variants: Dict[FootType, List[FootVariant]] = {}
        self._build_all_variants()
    
    def _build_all_variants(self):
        """Build standard and variant forms for all foot types"""
        
        # === Standard Foot Patterns ===
        standard_patterns = {
            # Five-letter feet (الخماسية)
            FootType.FAULUN: [SyllableType.LIGHT, SyllableType.HEAVY, SyllableType.HEAVY],
            FootType.FAILUN: [SyllableType.HEAVY, SyllableType.LIGHT, SyllableType.HEAVY],
            
            # Seven-letter feet (السباعية)
            FootType.MAFALIYUN: [SyllableType.LIGHT, SyllableType.HEAVY, SyllableType.HEAVY, SyllableType.HEAVY],
            FootType.FAILATUN: [SyllableType.HEAVY, SyllableType.LIGHT, SyllableType.HEAVY, SyllableType.HEAVY],
            FootType.MUSTAFILUN: [SyllableType.HEAVY, SyllableType.HEAVY, SyllableType.LIGHT, SyllableType.HEAVY],
            FootType.MUFAWALATUN: [SyllableType.LIGHT, SyllableType.HEAVY, SyllableType.LIGHT, SyllableType.HEAVY],
            FootType.MUTAFALIYUN: [SyllableType.LIGHT, SyllableType.LIGHT, SyllableType.HEAVY, SyllableType.HEAVY],
            FootType.MAFULATU: [SyllableType.HEAVY, SyllableType.HEAVY, SyllableType.HEAVY, SyllableType.LIGHT],
        }
        
        # Build standard variants
        for foot_type, pattern in standard_patterns.items():
            self._variants[foot_type] = [
                FootVariant(
                    name=f"{foot_type.value} (standard)",
                    pattern=pattern,
                    foot_type=foot_type,
                    is_standard=True
                )
            ]
        
        # Add zihāf variants
        for foot_type in FootType:
            if foot_type not in self._variants:
                continue
                
            zihaf_variants = self.zihafat.get_all_variants_for_foot(foot_type)
            for zihaf_type, pattern in zihaf_variants:
                zihaf = self.zihafat.get_zihaf(zihaf_type)
                if zihaf:
                    self._variants[foot_type].append(
                        FootVariant(
                            name=f"{foot_type.value} ({zihaf.name})",
                            pattern=pattern,
                            foot_type=foot_type,
                            is_standard=False,
                            zihaf_applied=zihaf_type
                        )
                    )
    
    def get_variants(self, foot_type: FootType) -> List[FootVariant]:
        """Get all variants for a foot type"""
        return self._variants.get(foot_type, [])
    
    def get_standard_variant(self, foot_type: FootType) -> Optional[FootVariant]:
        """Get the standard (non-zihāf) variant"""
        variants = self._variants.get(foot_type, [])
        for v in variants:
            if v.is_standard:
                return v
        return None
    
    def get_all_variants(self) -> Dict[FootType, List[FootVariant]]:
        """Get all variants for all foot types"""
        return self._variants
    
    def find_matching_variant(
        self,
        pattern: List[SyllableType],
        preferred_foot: Optional[FootType] = None
    ) -> Optional[FootVariant]:
        """
        Find a foot variant that matches the given syllable pattern.
        
        Args:
            pattern: The syllable pattern to match
            preferred_foot: If specified, search this foot type first
        
        Returns:
            Matching FootVariant or None
        """
        # Search preferred foot first
        if preferred_foot:
            for variant in self._variants.get(preferred_foot, []):
                if self._patterns_match(variant.pattern, pattern):
                    return variant
        
        # Search all foot types
        for foot_type, variants in self._variants.items():
            if foot_type == preferred_foot:
                continue
            for variant in variants:
                if self._patterns_match(variant.pattern, pattern):
                    return variant
        
        return None
    
    def _patterns_match(
        self,
        variant_pattern: List[SyllableType],
        actual_pattern: List[SyllableType]
    ) -> bool:
        """Check if patterns match, accounting for ANCEPS flexibility"""
        if len(variant_pattern) != len(actual_pattern):
            return False
        
        for v, a in zip(variant_pattern, actual_pattern):
            if v == SyllableType.ANCEPS or a == SyllableType.ANCEPS:
                continue  # Anceps matches anything
            if v != a:
                return False
        
        return True


# =============================================================================
# ARUZ METER ANALYZER (REVISED)
# =============================================================================

class AruzMeterAnalyzer:
    """
    Classical ʿArūḍ analyzer focusing on FOOT identification and matching.
    """

    def __init__(self, config: Optional[AnalysisConfig] = None):
        self.config = config or AnalysisConfig()
        self.phonetics = PersianTajikPhonetics()
        self.phonology = ClassicalPhonology()
        self.foot_builder = FootVariantBuilder()
        self.meter_registry = MeterRegistry()
        self.zihafat = ZihafatRegistry()
        self.logger = logging.getLogger(f"{__name__}.AruzMeterAnalyzer")

    def analyze_meter(self, line: str) -> AruzAnalysis:
        """Analyze a single line of poetry for ʿArūḍ meter using foot-based approach."""
        try:
            clean_line = self._clean_line(line)
            if not clean_line:
                return self._create_empty_analysis("Empty line after cleaning")

            syllable_tuples = self.phonetics._syllabify(clean_line)
            
            if not syllable_tuples:
                self.logger.warning(f"No syllables found in line: {line[:50]}...")
                return self._create_empty_analysis("No syllables found")

            # Determine syllable weights
            syllable_weights = self._analyze_syllable_weights(syllable_tuples, clean_line)
            
            # Create prosodic syllables for output
            line_scansion = self._create_line_scansion(syllable_tuples, syllable_weights)
            
            # Try to match against known meters
            match_result = self._match_meter(syllable_weights)
            
            if match_result:
                return AruzAnalysis(
                    identified_meter=match_result.meter.transliteration,
                    meter_arabic=match_result.meter.name,
                    pattern_match=match_result.meter.description,
                    full_pattern=self._weights_to_string(syllable_weights),
                    confidence=match_result.confidence,
                    pattern_accuracy=match_result.accuracy_score,
                    variations_detected=match_result.variations_used,
                    line_scansion=line_scansion,
                    caesura_positions=self._find_caesura_positions(syllable_weights, match_result),
                    feet_breakdown=match_result.matched_feet,
                    alternative_meters=self._find_alternative_meters(syllable_weights, match_result.meter)
                )
            else:
                return AruzAnalysis(
                    identified_meter="unknown",
                    meter_arabic="غير معروف",
                    pattern_match="",
                    full_pattern=self._weights_to_string(syllable_weights),
                    confidence=MeterConfidence.NONE,
                    pattern_accuracy=0.0,
                    variations_detected=[],
                    line_scansion=line_scansion,
                    caesura_positions=[],
                    feet_breakdown=[],
                    alternative_meters=self._find_possible_meters(syllable_weights)
                )

        except Exception as e:
            self.logger.error(f"Meter analysis failed for line '{line[:50]}...': {e}", exc_info=True)
            return self._create_empty_analysis(f"Analysis error: {str(e)}")

    def _clean_line(self, line: str) -> str:
        """Clean and normalize a line for analysis"""
        # Remove punctuation except for prosodically relevant marks
        line = re.sub(r'[.,;:!?؟،؛«»()\[\]{}"\'\-]', '', line)
        # Normalize whitespace
        line = ' '.join(line.split())
        # Normalize Unicode
        line = unicodedata.normalize('NFC', line)
        return line.strip()

    def _analyze_syllable_weights(
        self,
        syllable_tuples: List[Tuple[int, str]],
        original_line: str = ""
    ) -> List[SyllableType]:
        """Analyze the prosodic weight of each syllable"""
        weights = []
        total = len(syllable_tuples)
        
        # Build word list for context lookup
        words = original_line.split() if original_line else []
        
        for i, (pos, syl_text) in enumerate(syllable_tuples):
            # Find which word this syllable belongs to
            current_word = self._find_word_at_position(original_line, pos) if original_line else ""
            
            context = {
                'position_in_line': i,
                'total_syllables': total,
                'is_final': (i == total - 1),
                'word': current_word
            }
            weight = self.phonology.analyze_syllable(syl_text, context)
            weights.append(weight)
        
        return weights
    
    def _find_word_at_position(self, line: str, pos: int) -> str:
        """Find the word that contains the character at position pos"""
        if not line or pos < 0 or pos >= len(line):
            return ""
        
        # Find word boundaries
        start = pos
        while start > 0 and not line[start - 1].isspace():
            start -= 1
        
        end = pos
        while end < len(line) and not line[end].isspace():
            end += 1
        
        return line[start:end]

    def _create_line_scansion(
        self,
        syllable_tuples: List[Tuple[int, str]],
        weights: List[SyllableType]
    ) -> List[ProsodicSyllable]:
        """Create detailed prosodic syllable information"""
        scansion = []
        
        for i, ((_, syl_text), weight) in enumerate(zip(syllable_tuples, weights)):
            # Convert SyllableType to SyllableWeight for output
            if weight == SyllableType.HEAVY:
                output_weight = SyllableWeight.HEAVY
            elif weight == SyllableType.LIGHT:
                output_weight = SyllableWeight.LIGHT
            elif weight == SyllableType.ANCEPS:
                output_weight = SyllableWeight.ANCEPS
            else:
                output_weight = SyllableWeight.UNKNOWN
            
            # Get phonetic transcription
            phonetic_analysis = self.phonetics.analyze_phonetics(syl_text)
            
            scansion.append(ProsodicSyllable(
                text=syl_text,
                weight=output_weight,
                phonetic=phonetic_analysis.phonetic_transcription,
                position=i,
                confidence=phonetic_analysis.confidence,
                stress_level=phonetic_analysis.stress_pattern[0] if phonetic_analysis.stress_pattern else 0
            ))
        
        return scansion

    def _match_meter(self, weights: List[SyllableType]) -> Optional[MatchedMeterResult]:
        """
        Match the syllable weight pattern against all known meters.
        Uses dynamic programming for flexible matching with zihāfāt.
        """
        best_match = None
        best_score = 0.0
        
        # Get meters sorted by frequency (try most common first)
        meters = self.meter_registry.get_meters_sorted_by_frequency()
        
        for meter in meters:
            result = self._match_against_meter(weights, meter)
            if result and result.accuracy_score > best_score:
                best_score = result.accuracy_score
                best_match = result
                
                # Early exit if perfect match
                if best_score >= 0.98:
                    break
        
        return best_match

    def _match_against_meter(
        self,
        weights: List[SyllableType],
        meter: MeterDefinition
    ) -> Optional[MatchedMeterResult]:
        """Match syllable weights against a specific meter definition
        
        Includes tolerance for:
        - Catalexis: missing syllable at line end (common in Persian/Tajik poetry)
        - Hypercatalexis: extra syllable at line end
        - Anceps: flexible syllable weight at certain positions
        """
        
        expected_len = len(meter.full_pattern)
        actual_len = len(weights)
        
        # Allow flexibility for catalexis/hypercatalexis (±2 syllables)
        if abs(expected_len - actual_len) > 2:
            return None
        
        # Try direct matching first
        match_count = 0
        mismatches = []
        variations_used = []
        
        min_len = min(expected_len, actual_len)
        
        for i in range(min_len):
            expected = meter.full_pattern[i]
            actual = weights[i]
            
            if expected == actual:
                match_count += 1
            elif expected == SyllableType.ANCEPS or actual == SyllableType.ANCEPS:
                match_count += 1  # Anceps is flexible
            else:
                mismatches.append((i, expected, actual))
        
        # === Handle catalexis (missing final syllable) ===
        # If actual is 1 shorter and core pattern matches well, accept it
        if actual_len == expected_len - 1 and match_count >= min_len * 0.85:
            variations_used.append("catalexis")
            # Credit the missing syllable as a partial match
            match_count += 0.5
        
        # === Handle hypercatalexis (extra final syllable) ===
        # If actual is 1 longer and core pattern matches well, accept it
        elif actual_len == expected_len + 1 and match_count >= min_len * 0.85:
            variations_used.append("hypercatalexis")
            # The extra syllable shouldn't heavily penalize the match
            match_count += 0.5
        
        # === Handle brachycatalexis (missing 2 final syllables) ===
        elif actual_len == expected_len - 2 and match_count >= min_len * 0.9:
            variations_used.append("brachycatalexis")
            match_count += 0.3
        
        # Calculate base accuracy
        # Use expected_len as denominator to not over-penalize short lines
        if expected_len > 0:
            accuracy = match_count / expected_len
        else:
            accuracy = 0.0
        
        # Gentle penalty for length mismatch (reduced from 0.03)
        if actual_len != expected_len and not variations_used:
            length_penalty = abs(actual_len - expected_len) * 0.02
            accuracy = max(0, accuracy - length_penalty)
        
        # Try to explain mismatches via zihāfāt
        if mismatches and accuracy >= 0.5:
            explained = self._explain_mismatches_via_zihafat(mismatches, meter, weights)
            if explained:
                variations_used = explained
                # Reduce penalty for explained variations
                accuracy += len(explained) * 0.02
                accuracy = min(accuracy, 1.0)
        
        # Determine confidence level
        if accuracy >= self.config.high_confidence_threshold:
            confidence = MeterConfidence.HIGH
        elif accuracy >= self.config.medium_confidence_threshold:
            confidence = MeterConfidence.MEDIUM
        elif accuracy >= self.config.low_confidence_threshold:
            confidence = MeterConfidence.LOW
        else:
            return None  # Too low to report
        
        # Build matched feet
        matched_feet = self._build_matched_feet(weights, meter)
        
        return MatchedMeterResult(
            meter=meter,
            confidence=confidence,
            accuracy_score=accuracy,
            variations_used=variations_used,
            matched_feet=matched_feet,
            alignment=[(i, i) for i in range(min_len)]
        )

    def _explain_mismatches_via_zihafat(
        self,
        mismatches: List[Tuple[int, SyllableType, SyllableType]],
        meter: MeterDefinition,
        weights: List[SyllableType]
    ) -> List[str]:
        """Try to explain mismatches using known zihāfāt"""
        explained = []
        
        # Group mismatches by foot position
        foot_positions = self._get_foot_positions(meter)
        
        for pos, expected, actual in mismatches:
            # Find which foot this position belongs to
            for foot_idx, (foot_type, start, end) in enumerate(foot_positions):
                if start <= pos < end:
                    # Check if any zihāf can explain this
                    zihaf_type = self.zihafat.find_matching_zihaf(
                        foot_type,
                        weights[start:end]
                    )
                    if zihaf_type:
                        zihaf = self.zihafat.get_zihaf(zihaf_type)
                        if zihaf and zihaf.name not in explained:
                            explained.append(zihaf.name)
                    break
        
        return explained

    def _get_foot_positions(
        self,
        meter: MeterDefinition
    ) -> List[Tuple[FootType, int, int]]:
        """Get the start/end positions of each foot in a meter"""
        positions = []
        current_pos = 0
        
        for foot_type in meter.base_feet:
            variant = self.foot_builder.get_standard_variant(foot_type)
            if variant:
                foot_len = len(variant.pattern)
                positions.append((foot_type, current_pos, current_pos + foot_len))
                current_pos += foot_len
        
        return positions

    def _build_matched_feet(
        self,
        weights: List[SyllableType],
        meter: MeterDefinition
    ) -> List[FootVariant]:
        """Build the sequence of foot variants that match the weights"""
        matched = []
        current_pos = 0
        
        for foot_type in meter.base_feet:
            variants = self.foot_builder.get_variants(foot_type)
            variant = self.foot_builder.get_standard_variant(foot_type)
            
            if variant:
                foot_len = len(variant.pattern)
                actual_pattern = weights[current_pos:current_pos + foot_len]
                
                # Try to find best matching variant
                best_variant = None
                for v in variants:
                    if self.foot_builder._patterns_match(v.pattern, actual_pattern):
                        best_variant = v
                        break
                
                matched.append(best_variant or variant)
                current_pos += foot_len
        
        return matched

    def _find_caesura_positions(
        self,
        weights: List[SyllableType],
        match_result: MatchedMeterResult
    ) -> List[int]:
        """Find caesura (pause) positions based on foot boundaries"""
        positions = []
        current_pos = 0
        
        for i, foot in enumerate(match_result.matched_feet[:-1]):  # Exclude last foot
            current_pos += len(foot.pattern)
            positions.append(current_pos)
        
        return positions

    def _find_alternative_meters(
        self,
        weights: List[SyllableType],
        primary_meter: MeterDefinition
    ) -> List[Tuple[str, float]]:
        """Find alternative meter possibilities"""
        alternatives = []
        
        for meter in self.meter_registry.get_all_meters():
            if meter.transliteration == primary_meter.transliteration:
                continue
            
            result = self._match_against_meter(weights, meter)
            if result and result.accuracy_score >= 0.5:
                alternatives.append((meter.transliteration, result.accuracy_score))
        
        # Sort by score descending
        alternatives.sort(key=lambda x: -x[1])
        return alternatives[:3]  # Top 3 alternatives

    def _find_possible_meters(
        self,
        weights: List[SyllableType]
    ) -> List[Tuple[str, float]]:
        """Find possible meters when no confident match is found"""
        possibilities = []
        
        for meter in self.meter_registry.get_all_meters():
            result = self._match_against_meter(weights, meter)
            if result:
                possibilities.append((meter.transliteration, result.accuracy_score))
        
        possibilities.sort(key=lambda x: -x[1])
        return possibilities[:5]  # Top 5 possibilities

    def _weights_to_string(self, weights: List[SyllableType]) -> str:
        """Convert weights to symbolic string"""
        symbols = {
            SyllableType.LIGHT: "∪",
            SyllableType.HEAVY: "—",
            SyllableType.ANCEPS: "×",
            SyllableType.UNKNOWN: "?"
        }
        return " ".join(symbols.get(w, "?") for w in weights)

    def _create_empty_analysis(self, reason: str = "") -> AruzAnalysis:
        """Create an empty analysis for error cases"""
        return AruzAnalysis(
            identified_meter="unknown",
            meter_arabic="غير معروف",
            pattern_match="",
            full_pattern="",
            confidence=MeterConfidence.NONE,
            pattern_accuracy=0.0,
            variations_detected=[],
            line_scansion=[],
            caesura_positions=[],
            feet_breakdown=[],
            alternative_meters=[]
        )


# =============================================================================
# RHYME AND RADĪF ANALYZER
# =============================================================================

class RhymeRadifAnalyzer:
    """Analyzer for rhyme (Qāfiyeh) and refrain (Radīf)"""
    
    def __init__(self):
        self.phonetics = PersianTajikPhonetics()
        self.logger = logging.getLogger(f"{__name__}.RhymeRadifAnalyzer")

        # Common radīf patterns
        self.common_radif_patterns = [
            'است', 'شد', 'شود', 'کرد', 'کند', 'بود', 'نیست', 'هست',
            'می‌شود', 'می‌کند', 'خواهد شد', 'خواهد بود',
            # Tajik Cyrillic equivalents
            'аст', 'шуд', 'шавад', 'кард', 'кунад', 'буд', 'нест', 'ҳаст',
            'мешавад', 'мекунад', 'хоҳад шуд', 'хоҳад буд'
        ]
    
    def analyze_rhyme(self, line: str) -> RhymeAnalysis:
        """Analyze rhyme for a single line"""
        line = line.strip()
        
        if not line:
            return RhymeAnalysis(
                qafiya="",
                radif="",
                rawi="",
                phonetic_rhyme="",
                rhyme_type="none",
                rhyme_position=-1,
                confidence=0.0
            )
        
        # Extract words from line
        words = re.findall(r'[\wӣӯ]+', line)
        
        if not words:
            return RhymeAnalysis(
                qafiya="",
                radif="",
                rawi="",
                phonetic_rhyme="",
                rhyme_type="none",
                rhyme_position=-1,
                confidence=0.0
            )
        
        # Last word is typically the qafiya (rhyme word)
        last_word = words[-1]
        
        # Get phonetic representation
        try:
            phonetic_result = self.phonetics.to_phonetic(last_word)
            
            # Try different access patterns
            if hasattr(phonetic_result, 'phonetic_transcription'):
                phonetic = phonetic_result.phonetic_transcription
            elif hasattr(phonetic_result, 'transcription'):
                phonetic = phonetic_result.transcription
            elif isinstance(phonetic_result, str):
                phonetic = phonetic_result
            else:
                phonetic = str(phonetic_result)
                
        except Exception as e:
            self.logger.warning(f"Phonetic analysis failed for '{last_word}': {e}")
            phonetic = last_word.lower()
        
        # Extract rawi (the key consonant of the rhyme - usually last consonant before final vowel)
        vowels = set('аеёиоуыэюяaeiouāīūēōө')
        rawi = ""
        
        # Find rawi - last consonant
        for char in reversed(last_word.lower()):
            if char.isalpha() and char not in vowels:
                rawi = char
                break
        
        # Get phonetic ending (last 2-4 characters for rhyme matching)
        phonetic_rhyme = phonetic[-4:] if len(phonetic) >= 4 else phonetic
        
        # Determine rhyme type
        if len(last_word) <= 2:
            rhyme_type = "poor"
            confidence = 0.4
        elif rawi:
            rhyme_type = "standard"
            confidence = 0.7
        else:
            rhyme_type = "vowel"
            confidence = 0.5
        
        return RhymeAnalysis(
            qafiya=last_word,
            radif="",  # Radif is detected separately at poem level
            rawi=rawi,
            phonetic_rhyme=phonetic_rhyme,
            rhyme_type=rhyme_type,
            rhyme_position=len(words) - 1,
            confidence=confidence
        )
            

    def analyze_radif(self, lines: List[str]) -> RadifAnalysis:
        """
        Detect and analyze the radīf in a poem.
        """
        if not lines or len(lines) < 2:
            return RadifAnalysis(
                radif_present=False,
                radif_text="",
                radif_words=[],
                radif_frequency=0.0,
                qafiya_pattern="",
                lines_with_radif=[],
                lines_without_radif=list(range(len(lines)))
            )
        
        # Extract line endings for comparison
        line_endings = [self._get_line_ending(line) for line in lines]
        
        # Try to find common ending (radīf)
        radif_text, radif_words = self._find_common_radif(line_endings)
        
        if not radif_text:
            # No radīf found, analyze only qāfiya
            qafiya_pattern = self._analyze_qafiya_only(line_endings)
            return RadifAnalysis(
                radif_present=False,
                radif_text="",
                radif_words=[],
                radif_frequency=0.0,
                qafiya_pattern=qafiya_pattern,
                lines_with_radif=[],
                lines_without_radif=list(range(len(lines)))
            )
        
        # Count lines with this radīf
        lines_with_radif = []
        lines_without_radif = []
        
        for i, ending in enumerate(line_endings):
            if ending.endswith(radif_text) or radif_text in ending:
                lines_with_radif.append(i)
            else:
                lines_without_radif.append(i)
        
        radif_frequency = len(lines_with_radif) / len(lines) if lines else 0.0
        
        # Analyze qāfiya (what comes before radīf)
        qafiya_pattern = self._analyze_qafiya_with_radif(line_endings, radif_text)
        
        return RadifAnalysis(
            radif_present=True,
            radif_text=radif_text,
            radif_words=radif_words,
            radif_frequency=radif_frequency,
            qafiya_pattern=qafiya_pattern,
            lines_with_radif=lines_with_radif,
            lines_without_radif=lines_without_radif
        )

    def analyze_rhyme_scheme(self, lines: List[str]) -> Tuple[str, List[RhymeAnalysis]]:
        """
        Analyze the rhyme scheme of a poem.
        """
        if not lines:
            return ("", [])
        
        analyses = []
        rhyme_classes = {}  # Maps rhyme sound to letter
        current_letter_index = 0
        scheme_parts = []
        
        for i, line in enumerate(lines):
            ending = self._get_line_ending(line)
            rhyme_sound = self._extract_rhyme_sound(ending)
            
            # Find or assign rhyme class
            if rhyme_sound not in rhyme_classes:
                rhyme_classes[rhyme_sound] = chr(ord('A') + current_letter_index)
                current_letter_index = min(current_letter_index + 1, 25)  # Max 26 classes
            
            rhyme_letter = rhyme_classes[rhyme_sound]
            scheme_parts.append(rhyme_letter)
            
            # Create detailed analysis
            analysis = RhymeAnalysis(
                qafiya=self._extract_qafiya(ending),
                radif=self._extract_radif_from_ending(ending),
                rawi=self._extract_rawi(ending),
                phonetic_rhyme=self.phonetics._to_ipa(rhyme_sound, is_cyrillic=True),
                rhyme_type=self._classify_rhyme_type(rhyme_sound),
                rhyme_position="end",
                confidence=0.85
            )
            analyses.append(analysis)
        
        # Format scheme (group by pairs for couplets)
        formatted_scheme = self._format_rhyme_scheme(scheme_parts)
        
        return (formatted_scheme, analyses)

    def _get_line_ending(self, line: str, num_words: int = 3) -> str:
        """Get the last N words of a line"""
        words = line.strip().split()
        return ' '.join(words[-num_words:]) if len(words) >= num_words else line.strip()

    def _find_common_radif(self, endings: List[str]) -> Tuple[str, List[str]]:
        """Find the longest common suffix (radīf) among line endings"""
        if not endings or len(endings) < 2:
            return ("", [])
        
        # Try word-level matching first
        word_lists = [ending.split() for ending in endings]
        
        # Find common suffix words
        min_words = min(len(wl) for wl in word_lists)
        common_words = []
        
        for i in range(1, min_words + 1):
            # Check if last i words are the same across all endings
            reference = word_lists[0][-i:]
            all_match = all(wl[-i:] == reference for wl in word_lists)
            
            if all_match:
                common_words = reference
            else:
                break
        
        if common_words:
            return (' '.join(common_words), common_words)
        
        # Try character-level matching as fallback
        min_len = min(len(e) for e in endings)
        common_suffix = ""
        
        for i in range(1, min_len + 1):
            # Check if last i characters are the same
            reference = endings[0][-i:]
            all_match = all(e[-i:] == reference for e in endings)
            
            if all_match:
                common_suffix = reference
            else:
                break
        
        # Only count as radīf if it's at least 2 characters
        if len(common_suffix) >= 2:
            return (common_suffix.strip(), common_suffix.strip().split())
        
        return ("", [])

    def _analyze_qafiya_only(self, endings: List[str]) -> str:
        """Analyze qāfiya pattern when there's no radīf"""
        rhyme_sounds = [self._extract_rhyme_sound(e) for e in endings]
        
        if not rhyme_sounds:
            return ""
        
        # Find most common rhyme sound
        sound_counts = Counter(rhyme_sounds)
        most_common, count = sound_counts.most_common(1)[0]
        
        return f"-{most_common}" if most_common else ""

    def _analyze_qafiya_with_radif(self, endings: List[str], radif: str) -> str:
        """Analyze qāfiya pattern when radīf is present"""
        qafiya_parts = []
        
        for ending in endings:
            # Remove radīf to get qāfiya
            if ending.endswith(radif):
                before_radif = ending[:-len(radif)].strip()
                qafiya = self._extract_rhyme_sound(before_radif)
                qafiya_parts.append(qafiya)
            else:
                qafiya_parts.append(self._extract_rhyme_sound(ending))
        
        # Find most common pattern
        if qafiya_parts:
            sound_counts = Counter(qafiya_parts)
            most_common, _ = sound_counts.most_common(1)[0]
            return f"-{most_common}"
        
        return ""

    def _extract_rhyme_sound(self, text: str) -> str:
        """Extract the rhyming sound from text"""
        text = text.strip().lower()
        if not text:
            return ""
        
        # Get last word
        words = text.split()
        if not words:
            return ""
        
        last_word = words[-1]
        
        # Find the last vowel and everything after
        vowels = set('аеиоуэӣӯяюёaeiouāīū')
        
        last_vowel_pos = -1
        for i, char in enumerate(last_word):
            if char in vowels:
                last_vowel_pos = i
        
        if last_vowel_pos >= 0:
            return last_word[last_vowel_pos:]
        
        # Fallback: return last 2-3 characters
        return last_word[-3:] if len(last_word) >= 3 else last_word

    def _extract_qafiya(self, ending: str) -> str:
        """Extract the full qāfiya from a line ending"""
        return self._extract_rhyme_sound(ending)

    def _extract_radif_from_ending(self, ending: str) -> str:
        """Try to extract radīf from a single line ending"""
        words = ending.strip().split()
        
        # Check if last word is a common radīf
        if words:
            last_word = words[-1]
            if last_word in self.common_radif_patterns:
                return last_word
            
            # Check last two words
            if len(words) >= 2:
                last_two = ' '.join(words[-2:])
                if last_two in self.common_radif_patterns:
                    return last_two
        
        return ""

    def _extract_rawi(self, ending: str) -> str:
        """
        Extract the rawi (روی) - the main rhyme consonant.
        """
        text = ending.strip().lower()
        if not text:
            return ""
        
        words = text.split()
        if not words:
            return ""
        
        last_word = words[-1]
        vowels = set('аеиоуэӣӯяюёaeiouāīū')
        consonants = set('бвгғдежзйклмнпрстфхҳчҷшъьқbcdfghjklmnpqrstvwxyz')
        
        # Find the last consonant before final vowel(s)
        in_final_vowels = True
        for i in range(len(last_word) - 1, -1, -1):
            char = last_word[i]
            if char in vowels:
                continue
            elif char in consonants:
                return char
            in_final_vowels = False
        
        return ""

    def _classify_rhyme_type(self, rhyme_sound: str) -> str:
        """Classify the type of rhyme"""
        if not rhyme_sound:
            return "none"
        
        vowels = set('аеиоуэӣӯяюёaeiouāīū')
        consonants = set('бвгғдежзйклмнпрстфхҳчҷшъьқbcdfghjklmnpqrstvwxyz')
        
        vowel_count = sum(1 for c in rhyme_sound if c in vowels)
        consonant_count = sum(1 for c in rhyme_sound if c in consonants)
        
        if vowel_count >= 1 and consonant_count >= 1:
            return "perfect"  # قافیه تامه
        elif vowel_count >= 1:
            return "vowel"  # قافیه مقیده
        elif consonant_count >= 1:
            return "consonant"
        else:
            return "weak"

    def _format_rhyme_scheme(self, scheme_parts: List[str]) -> str:
        """Format rhyme scheme, grouping by couplets"""
        if not scheme_parts:
            return ""
        
        # Group by pairs (hemistichs in a couplet)
        formatted_parts = []
        for i in range(0, len(scheme_parts), 2):
            if i + 1 < len(scheme_parts):
                formatted_parts.append(f"{scheme_parts[i]}{scheme_parts[i+1]}")
            else:
                formatted_parts.append(scheme_parts[i])
        
        return ' '.join(formatted_parts)


# =============================================================================
# STANZA FORM DETECTOR
# =============================================================================

class StanzaFormDetector:
    """
    Detects the classical Persian/Tajik poetic form.
    """
    
    def __init__(self):
        self.logger = logging.getLogger(f"{__name__}.StanzaFormDetector")
        self.rhyme_analyzer = RhymeRadifAnalyzer()
    
    def detect_form(
        self,
        lines: List[str],
        rhyme_scheme: str,
        radif_analysis: RadifAnalysis,
        syllable_counts: List[int]
    ) -> Tuple[StanzaForm, float]:
        """
        Detect the poetic form based on structural analysis.
        """
        num_lines = len(lines)
        
        if num_lines == 0:
            return (StanzaForm.UNKNOWN, 0.0)
        
        # === Check for Rubāʿī (رباعی) - exactly 4 lines ===
        if num_lines == 4:
            if self._is_rubai(rhyme_scheme, syllable_counts):
                return (StanzaForm.RUBAI, 0.9)
        
        # === Check for Ghazal/Qaṣīda patterns ===
        if self._is_ghazal_qasida_pattern(rhyme_scheme):
            # Distinguish by length and radīf presence
            num_couplets = num_lines // 2
            
            if radif_analysis.radif_present:
                if num_couplets <= 15:
                    return (StanzaForm.GHAZAL_WITH_RADIF, 0.85)
                else:
                    return (StanzaForm.QASIDA_WITH_RADIF, 0.80)
            else:
                if num_couplets <= 15:
                    return (StanzaForm.GHAZAL, 0.80)
                else:
                    return (StanzaForm.QASIDA, 0.75)
        
        # === Check for Masnavi (مثنوی) - AA BB CC pattern ===
        if self._is_masnavi_pattern(rhyme_scheme):
            return (StanzaForm.MASNAVI, 0.85)
        
        # === Check for Qiṭʿa (قطعه) - xA xA pattern without maṭlaʿ ===
        if self._is_qita_pattern(rhyme_scheme):
            return (StanzaForm.QITA, 0.75)
        
        # === Check for Mukhammas (مخمس) - 5-line stanzas ===
        if num_lines % 5 == 0 and self._is_mukhammas_pattern(rhyme_scheme):
            return (StanzaForm.MUKHAMMAS, 0.80)
        
        # === Check for Musaddas (مسدس) - 6-line stanzas ===
        if num_lines % 6 == 0 and self._is_musaddas_pattern(rhyme_scheme):
            return (StanzaForm.MUSADDAS, 0.80)
        
        # === Check for free verse (شعر آزاد) ===
        if self._is_free_verse(rhyme_scheme, syllable_counts):
            return (StanzaForm.FREE_VERSE, 0.70)
        
        return (StanzaForm.UNKNOWN, 0.5)
    
    def _is_rubai(self, rhyme_scheme: str, syllable_counts: List[int]) -> bool:
        """Check if the poem is a Rubāʿī"""
        # Rubāʿī rhyme scheme: AABA or AAAA
        scheme_letters = rhyme_scheme.replace(' ', '')
        
        if len(scheme_letters) != 4:
            return False
        
        valid_patterns = ['AABA', 'AAAA', 'ABAB']  # Classic rubāʿī patterns
        
        if scheme_letters in valid_patterns:
            # Additional check: syllable count should be consistent (typically 13)
            if syllable_counts and len(syllable_counts) == 4:
                avg = sum(syllable_counts) / 4
                variance = sum((s - avg) ** 2 for s in syllable_counts) / 4
                if variance < 4:  # Low variance indicates consistent meter
                    return True
        
        return False
    
    def _is_ghazal_qasida_pattern(self, rhyme_scheme: str) -> bool:
        """Check for Ghazal/Qaṣīda pattern: AA BA CA DA ..."""
        parts = rhyme_scheme.split()
        
        if len(parts) < 2:
            return False
        
        # First couplet should be AA (maṭlaʿ)
        if len(parts[0]) >= 2 and parts[0][0] == parts[0][1]:
            # Subsequent couplets should end with same rhyme
            rhyme_letter = parts[0][0]
            for part in parts[1:]:
                if len(part) >= 2 and part[-1] != rhyme_letter:
                    return False
            return True
        
        return False
    
    def _is_masnavi_pattern(self, rhyme_scheme: str) -> bool:
        """Check for Masnavi pattern: AA BB CC DD ..."""
        parts = rhyme_scheme.split()
        
        if len(parts) < 2:
            return False
        
        for part in parts:
            if len(part) >= 2 and part[0] != part[1]:
                return False
        
        # Check that couplets have different rhymes
        unique_rhymes = set(part[0] for part in parts if part)
        return len(unique_rhymes) > 1
    
    def _is_qita_pattern(self, rhyme_scheme: str) -> bool:
        """Check for Qiṭʿa pattern: xA xA xA ... (no maṭlaʿ)"""
        parts = rhyme_scheme.split()
        
        if len(parts) < 2:
            return False
        
        # First couplet should NOT have matching rhymes (unlike Ghazal)
        if len(parts[0]) >= 2 and parts[0][0] == parts[0][1]:
            return False
        
        # Second hemistich of each couplet should rhyme
        rhyme_letter = parts[0][-1] if parts[0] else ''
        for part in parts:
            if len(part) >= 1 and part[-1] != rhyme_letter:
                return False
        
        return True
    
    def _is_mukhammas_pattern(self, rhyme_scheme: str) -> bool:
        """Check for Mukhammas pattern: AAAAB CCCDB EEEAB ..."""
        # Simplified check
        scheme_letters = rhyme_scheme.replace(' ', '')
        if len(scheme_letters) < 5:
            return False
        
        # Each 5-line stanza should have specific pattern
        for i in range(0, len(scheme_letters), 5):
            stanza = scheme_letters[i:i+5]
            if len(stanza) == 5:
                # First 4 lines should rhyme with each other
                if not (stanza[0] == stanza[1] == stanza[2] == stanza[3]):
                    return False
        
        return True
    
    def _is_musaddas_pattern(self, rhyme_scheme: str) -> bool:
        """Check for Musaddas pattern: 6-line stanzas"""
        scheme_letters = rhyme_scheme.replace(' ', '')
        if len(scheme_letters) < 6:
            return False
        
        # Each 6-line stanza check
        for i in range(0, len(scheme_letters), 6):
            stanza = scheme_letters[i:i+6]
            if len(stanza) == 6:
                # Typically: AAAAAB pattern
                if not (stanza[0] == stanza[1] == stanza[2] == stanza[3] == stanza[4]):
                    return False
        
        return True
    
    def _is_free_verse(self, rhyme_scheme: str, syllable_counts: List[int]) -> bool:
        """Check if the poem is free verse"""
        if not syllable_counts:
            return False
        
        # High variance in syllable counts suggests free verse
        if len(syllable_counts) >= 3:
            avg = sum(syllable_counts) / len(syllable_counts)
            variance = sum((s - avg) ** 2 for s in syllable_counts) / len(syllable_counts)
            std_dev = variance ** 0.5
            
            # High coefficient of variation suggests free verse
            cv = std_dev / avg if avg > 0 else 0
            if cv > 0.3:  # More than 30% variation
                return True
        
        # No consistent rhyme scheme
        scheme_letters = rhyme_scheme.replace(' ', '')
        unique_letters = set(scheme_letters)
        
        # If almost all lines have different rhymes, likely free verse
        if len(unique_letters) > len(scheme_letters) * 0.7:
            return True
        
        return False


# =============================================================================
# MODERN VERSE ANALYZER
# =============================================================================

class ModernVerseAnalyzer:
    """Analyzer for modern/free verse poetry"""

    def __init__(self):
        self.punctuation = set('.,!?;:—–-()[]{}"\'«»')
        self.sentence_enders = set('.!?;')
        self.prose_words = {'ва', 'ки', 'дар', 'бо', 'аз', 'то', 'барои', 'аммо', 'лекин'}

    def analyze(self, poem_content: str, syllable_counts: List[int] = None) -> ModernVerseMetrics:
        """Comprehensive analysis of free verse"""
        lines = [line.rstrip() for line in poem_content.split('\n') if line.strip()]

        if not lines:
            return ModernVerseMetrics()

        # Initialize metrics
        metrics = ModernVerseMetrics()

        # Enjambement analysis
        metrics.enjambement_count = self._count_enjambements(lines)
        metrics.enjambement_ratio = metrics.enjambement_count / max(len(lines) - 1, 1)

        # Semantic density (words per line)
        metrics.semantic_density = self._calculate_semantic_density(lines)

        # Line length variation
        if syllable_counts:
            metrics.line_length_variation = self._calculate_line_variation(syllable_counts)
        else:
            # Fallback: characters per line
            char_counts = [len(line) for line in lines]
            if char_counts:
                mean_val = statistics.mean(char_counts)
                if mean_val != 0:
                    stdev = statistics.stdev(char_counts) if len(char_counts) > 1 else 0
                    metrics.line_length_variation = stdev / mean_val

        # Prose-poetry score
        metrics.prose_poetry_score = self._calculate_prose_poetry_score(lines)

        # Visual structure
        metrics.visual_structure_score = self._analyze_visual_structure(poem_content)

        # Caesura distribution
        metrics.caesura_distribution = self._analyze_caesura_distribution(lines)

        # Syntactic parallelism
        metrics.syntactic_parallelism = self._calculate_syntactic_parallelism(lines)

        # Lexical repetition
        metrics.lexical_repetition_score = self._calculate_lexical_repetition(lines)

        # Breath group length
        metrics.breath_group_length = self._analyze_breath_groups(poem_content)

        # Pause frequency
        metrics.pause_frequency = self._calculate_pause_frequency(lines)

        return metrics

    def _count_enjambements(self, lines: List[str]) -> int:
        """Count enjambements (line breaks)"""
        count = 0
        for i in range(len(lines) - 1):
            current_line = lines[i].strip()
            next_line = lines[i + 1].strip()

            if not current_line or not next_line:
                continue

            # Enjambement if line doesn't end with sentence ending punctuation
            # AND next line doesn't start with uppercase (not a new sentence)
            if (current_line[-1] not in self.sentence_enders and
                    not next_line[0].isupper()):
                count += 1

        return count

    def _calculate_semantic_density(self, lines: List[str]) -> float:
        """Calculate semantic density (words per line)"""
        word_counts = [len(re.findall(r'[\wӣӯ]+', line)) for line in lines]
        return statistics.mean(word_counts) if word_counts else 0

    def _calculate_line_variation(self, syllable_counts: List[int]) -> float:
        """Calculate coefficient of variation of line lengths"""
        # Handle cases with insufficient data
        if len(syllable_counts) < 2:
            # If all lines have the same count, variation is 0. 
            # If only one line, we can't determine variation, so also return 0.
            return 0.0 

        mean_val = statistics.mean(syllable_counts)
        if mean_val == 0:
            return 0.0

        try:
            stdev = statistics.stdev(syllable_counts)
        except statistics.StatisticsError:
            # This should theoretically not happen due to the len check, 
            # but catch it just in case.
            stdev = 0.0

        return stdev / mean_val if mean_val != 0 else 0.0

    def _calculate_prose_poetry_score(self, lines: List[str]) -> float:
        """Calculate prose-poetry score (0 = purely poetic, 1 = prosaic)"""
        if not lines:
            return 0

        scores = []

        for line in lines:
            # Shorter lines are more poetic
            length_score = min(1.0, len(line) / 100)

            # Punctuation at end = more prosaic
            punctuation_score = 1.0 if line[-1] in self.sentence_enders else 0

            # Contains everyday language?
            words = set(re.findall(r'[\wӣӯ]+', line.lower()))
            prose_word_score = len(words & self.prose_words) / max(len(words), 1)

            line_score = (length_score * 0.3 +
                         punctuation_score * 0.4 +
                         prose_word_score * 0.3)
            scores.append(line_score)

        return statistics.mean(scores) if scores else 0

    def _analyze_visual_structure(self, poem_content: str) -> float:
        """Analyze visual structure (indentations, blank lines)"""
        lines = poem_content.split('\n')

        if len(lines) < 2:
            return 0

        # Count indentations
        indent_count = 0
        for line in lines:
            if line.startswith((' ', '\t')) and line.strip():
                indent_count += 1

        # Count blank line blocks
        empty_line_blocks = 0
        in_empty_block = False

        for line in lines:
            if not line.strip():
                if not in_empty_block:
                    empty_line_blocks += 1
                    in_empty_block = True
            else:
                in_empty_block = False

        visual_score = (indent_count / len(lines) * 0.5 +
                       empty_line_blocks / len(lines) * 0.5)

        return min(1.0, visual_score)

    def _analyze_caesura_distribution(self, lines: List[str]) -> List[int]:
        """Analyze caesura distribution"""
        caesura_positions = []

        for i, line in enumerate(lines):
            if ',' in line or '—' in line or '–' in line or ';' in line:
                caesura_positions.append(i)

        return caesura_positions

    def _calculate_syntactic_parallelism(self, lines: List[str]) -> float:
        """Calculate syntactic parallelism"""
        if len(lines) < 2:
            return 0

        parallel_pairs = 0
        total_pairs = 0

        for i in range(len(lines) - 1):
            line1 = lines[i].lower()
            line2 = lines[i + 1].lower()

            # Remove punctuation
            line1 = re.sub(r'[^\wӣӯ\s]', '', line1)
            line2 = re.sub(r'[^\wӣӯ\s]', '', line2)

            # Count words
            words1 = line1.split()
            words2 = line2.split()

            if len(words1) > 1 and len(words2) > 1:
                # Check for similar word order (beginnings)
                if words1[0] == words2[0]:
                    parallel_pairs += 1
                total_pairs += 1

        return parallel_pairs / total_pairs if total_pairs > 0 else 0

    def _calculate_lexical_repetition(self, lines: List[str]) -> float:
        """Calculate lexical repetition"""
        all_words = []
        for line in lines:
            words = re.findall(r'[\wӣӯ]+', line.lower())
            all_words.extend(words)

        if not all_words:
            return 0

        word_counts = {}
        for word in all_words:
            word_counts[word] = word_counts.get(word, 0) + 1

        # Percentage of words that appear more than once
        repeated_words = sum(1 for count in word_counts.values() if count > 1)
        total_unique_words = len(word_counts)

        return repeated_words / total_unique_words if total_unique_words > 0 else 0

    def _analyze_breath_groups(self, poem_content: str) -> float:
        """Analyze breath groups (sentence lengths)"""
        # Find sentence endings
        sentences = re.split(r'[.!?;]+', poem_content)
        sentences = [s.strip() for s in sentences if s.strip()]

        if not sentences:
            return 0

        # Words per sentence
        words_per_sentence = []
        for sentence in sentences:
            words = re.findall(r'[\wӣӯ]+', sentence)
            if words:
                words_per_sentence.append(len(words))

        return statistics.mean(words_per_sentence) if words_per_sentence else 0

    def _calculate_pause_frequency(self, lines: List[str]) -> float:
        """Calculate pause frequency (punctuation per line)"""
        if not lines:
            return 0

        punctuation_count = 0
        for line in lines:
            punctuation_count += sum(1 for char in line if char in self.punctuation)

        return punctuation_count / len(lines)


class HaikuDetector:
    """
    Detector for Haiku poetry (Japanese form adapted to Persian/Tajik)
    """
    
    # Persian haiku is more flexible - total syllables typically 10-20
    MIN_TOTAL_SYLLABLES = 8
    MAX_TOTAL_SYLLABLES = 25
    IDEAL_SYLLABLE_PATTERN = [5, 7, 5]
    MAX_SYLLABLE_DEVIATION = 2 # For near 5-7-5 patterns
    
    # Nature/seasonal keywords (kigo equivalents in Tajik Cyrillic)
    SEASONAL_KEYWORDS = {
        'baħor': ['баҳор', 'гул', 'навбаҳор', 'шукуфтан', 'сабза', 'булбул', 'лола', 'сабз'],
        'tobiston': ['офтоб', 'гармо', 'дарё', 'мева', 'ангур', 'тобистон'],
        'tirmoh': ['барг', 'хазон', 'зард', 'тирмоҳ'],
        'zimiston': ['барф', 'сармо', 'ях', 'замистон', 'шабнам']
    }
    
    # Nature elements (kigo-like)
    NATURE_KEYWORDS = [
        'моҳ', 'офтоб', 'осмон', 'ситора', 'абр', 'борон', 'шабнам',  # sky
        'дарё', 'об', 'чашма', 'дарьо', 'кӯл',  # water
        'кӯҳ', 'санг', 'замин', 'хок',  # earth
        'дарахт', 'гул', 'барг', 'мева', 'сабза',  # plants
        'парранда', 'булбул', 'кабутар', 'мурғ',  # birds/animals
        'бод', 'шаб', 'рӯз', 'субҳ', 'шом'  # time/elements
    ]
    
    # Mystical/Sufi keywords (Persian haiku links Buddhism & Sufism)
    MYSTICAL_KEYWORDS = [
        'дил', 'рӯҳ', 'ишқ', 'жон', 'маъно', 'ҳақиқат',  # soul/love/truth
        'сукут', 'хомӯшӣ', 'оромӣ',  # silence/peace
        'фано', 'бақо', 'ягонагӣ',  # Sufi concepts
        'лаҳза', 'дам', 'нафас', 'вақт',  # moment/breath/time
        'нур', 'равшанӣ', 'зулмат'  # light/darkness
    ]
    
    def __init__(self):
        self.phonetics = PersianTajikPhonetics()
        self.logger = logging.getLogger(f"{__name__}.HaikuDetector")
    
    def detect(self, poem_content: str, syllable_counts: List[int] = None) -> Dict[str, Any]:
        """
        Detect if a poem is a Persian/Tajik Haiku.
        """
        lines = [line.strip() for line in poem_content.split('\n') if line.strip()]
        
        # REQUIRED: exactly 3 lines
        if len(lines) != 3:
            return self._not_haiku()
        
        # Calculate syllables if not provided
        if syllable_counts is None or len(syllable_counts) != 3:
            syllable_counts = [self._count_syllables(line) for line in lines]
        
        total_syllables = sum(syllable_counts)
        
        # Check reasonable syllable range for Persian haiku
        if not (self.MIN_TOTAL_SYLLABLES <= total_syllables <= self.MAX_TOTAL_SYLLABLES):
            return self._not_haiku()
        
        # Check for nature OR mystical elements
        nature_elements = self._find_nature_elements(poem_content)
        mystical_elements = self._find_mystical_elements(poem_content)
        seasonal_elements = self._find_seasonal_elements(poem_content)
        
        all_elements = nature_elements + mystical_elements + seasonal_elements
        
        # Calculate confidence
        confidence = self._calculate_haiku_confidence(
            syllable_counts, all_elements, lines, nature_elements, mystical_elements
        )
        
        # Determine form variant
        form_variant = self._determine_variant(
            syllable_counts, nature_elements, mystical_elements
        )
        
        # Threshold: at least some haiku characteristics
        if confidence < 0.4:
            return self._not_haiku()
        
        self.logger.info(f"Haiku detected: {form_variant} with confidence {confidence:.0%}")
        
        return {
            'is_haiku': True,
            'confidence': confidence,
            'syllable_pattern': syllable_counts,
            'total_syllables': total_syllables,
            'is_575_pattern': syllable_counts == [5, 7, 5],
            'nature_elements': nature_elements,
            'mystical_elements': mystical_elements,
            'seasonal_elements': seasonal_elements,
            'form_variant': form_variant
        }
    
    def _count_syllables(self, line: str) -> int:
        """Count syllables in a line using phonetic analysis."""
        syllables = self.phonetics._syllabify(line)
        return len(syllables)
    
    def _find_seasonal_elements(self, text: str) -> List[str]:
        """Find seasonal keywords (kigo)."""
        text_lower = text.lower()
        found = []
        
        for season, keywords in self.SEASONAL_KEYWORDS.items():
            for kw in keywords:
                if kw in text_lower:
                    found.append(f"{kw} ({season})")
        
        return found
    
    def _find_nature_elements(self, text: str) -> List[str]:
        """Find nature keywords."""
        text_lower = text.lower()
        return [kw for kw in self.NATURE_KEYWORDS if kw in text_lower]
    
    def _find_mystical_elements(self, text: str) -> List[str]:
        """Find mystical/Sufi keywords (Persian haiku's Buddhist-Sufi nexus)."""
        text_lower = text.lower()
        return [kw for kw in self.MYSTICAL_KEYWORDS if kw in text_lower]
    
    def _calculate_haiku_confidence(self, syllables: List[int], 
                                    all_elements: List[str], 
                                    lines: List[str],
                                    nature_elements: List[str],
                                    mystical_elements: List[str]) -> float:
        """Calculate confidence score for Haiku classification."""
        confidence = 0.3  # Base confidence for 3-line structure
        
        # Traditional 5-7-5 pattern: +0.25 (but not required per Iranica)
        if syllables == [5, 7, 5]:
            confidence += 0.25
        # Close to 5-7-5: +0.1
        elif all(abs(syllables[i] - [5, 7, 5][i]) <= 2 for i in range(3)):
            confidence += 0.1
        
        # Nature elements: +0.15 per element, max +0.3
        confidence += min(0.3, len(nature_elements) * 0.15)
        
        # Mystical/Sufi elements: +0.1 per element, max +0.2
        # (Persian haiku often connects Buddhism & Sufism)
        confidence += min(0.2, len(mystical_elements) * 0.1)
        
        # Seasonal elements: +0.1
        if any('(' in e for e in all_elements):  # seasonal elements have season in parentheses
            confidence += 0.1
        
        # Short lines (typical for haiku): +0.1
        avg_words = sum(len(line.split()) for line in lines) / 3
        if avg_words <= 5:
            confidence += 0.1
        
        # Balanced line lengths: +0.05
        if max(syllables) - min(syllables) <= 4:
            confidence += 0.05
        
        return min(1.0, confidence)
    
    def _determine_variant(self, syllables: List[int], 
                          nature_elements: List[str],
                          mystical_elements: List[str]) -> str:
        """Determine the Haiku variant."""
        has_575 = syllables == [5, 7, 5]
        has_nature = len(nature_elements) > 0
        has_mystical = len(mystical_elements) > 0
        
        if has_575 and has_nature:
            return 'traditional_haiku'
        elif has_mystical and not has_nature:
            return 'sufi_haiku'  # Persian-specific: Buddhist-Sufi connection
        elif has_575:
            return 'senryū'  # Japanese-style without nature
        elif has_nature or has_mystical:
            return 'persian_haiku'  # Free-form Persian adaptation
        else:
            return 'free_haiku'
    
    def _not_haiku(self) -> Dict[str, Any]:
        """Return negative Haiku result."""
        return {
            'is_haiku': False,
            'confidence': 0.0,
            'syllable_pattern': [],
            'nature_elements': [],
            'mystical_elements': [],
            'form_variant': None
        }


class FreeVerseClassifier:
    """Classifier for free verse poetry"""

    @staticmethod
    def is_free_verse(structural_analysis, modern_metrics: ModernVerseMetrics) -> bool:
        """Determine if a poem is free verse"""
        criteria = {
            'meter_confidence_low': structural_analysis.meter_confidence.value in ['low', 'none'],
            'prosodic_inconsistency': structural_analysis.prosodic_consistency < 0.5,
            'enjambement_high': modern_metrics.enjambement_ratio > 0.3,
            'line_variation_high': modern_metrics.line_length_variation > 0.4,
            'prose_score_high': modern_metrics.prose_poetry_score > 0.6
        }

        weights = {
            'meter_confidence_low': 2.0,
            'prosodic_inconsistency': 1.5,
            'enjambement_high': 1.0,
            'line_variation_high': 1.0,
            'prose_score_high': 0.5
        }

        score = sum(weights[k] for k, v in criteria.items() if v)

        return score >= 2.5  # Threshold


# =============================================================================
# ENHANCED RADĪF DETECTOR
# =============================================================================

class EnhancedRadifDetector:
    """
    Enhanced Radīf (refrain) detector with metric correction.
    """
    
    def __init__(self):
        self.min_radif_frequency = 0.3
        self.phonetics = PersianTajikPhonetics()
        self.logger = logging.getLogger(f"{__name__}.EnhancedRadifDetector")
    
    def detect_radif_pattern(self, poem_content: str) -> RadifAnalysis:
        """Detect Radīf pattern and prepare corrected lines for meter analysis."""
        lines = [line.strip() for line in poem_content.split('\n') if line.strip()]
        
        if len(lines) < 2:
            return self._empty_radif_analysis(lines)
        
        # 1. Find common endings (Radīf candidates)
        radif_candidates = self._find_common_endings(lines)
        
        # 2. Select best Radīf
        best_radif = self._select_best_radif(radif_candidates, lines)

        # Safety net: discard "radifs" that occur on fewer than 2 lines.
        if best_radif:
            occurrence = sum(1 for line in lines if line.strip().endswith(best_radif))
            if occurrence < 2:
                best_radif = ""

        # 3. Remove Radīf and prepare cleaned lines for meter analysis
        if best_radif:
            cleaned_lines = self._remove_radif(lines, best_radif)
            radif_frequency = self._calculate_radif_frequency(lines, best_radif)
            
            # Find which lines have the radif
            lines_with_radif_indices = [
                i for i, line in enumerate(lines) 
                if line.strip().endswith(best_radif)
            ]
            
            lines_without_radif_indices = [
                i for i, line in enumerate(lines) 
                if not line.strip().endswith(best_radif)
            ]
            
            self.logger.info(f"Detected Radīf: '{best_radif}' in {radif_frequency:.0%} of lines")
        
            return RadifAnalysis(
                radif_present=True,
                radif_text=best_radif,
                radif_words=best_radif.split(),
                radif_frequency=radif_frequency,
                qafiya_pattern="",
                lines_with_radif=lines_with_radif_indices,
                lines_without_radif=lines_without_radif_indices,
                cleaned_lines=cleaned_lines
            )
        
        # No radif found - return empty radif analysis
        return self._empty_radif_analysis(lines)

    def _empty_radif_analysis(self, lines: List[str]) -> RadifAnalysis:
        """Return empty Radīf analysis with original lines as cleaned lines."""
        return RadifAnalysis(
            radif_present=False,
            radif_text="",
            radif_words=[],
            radif_frequency=0.0,
            qafiya_pattern="",
            lines_with_radif=[],
            lines_without_radif=list(range(len(lines))),
            cleaned_lines=lines  # Use original lines when no radif
        )
    
    def _find_common_endings(self, lines: List[str]) -> Dict[str, int]:
        """Find frequent endings in the last 1-3 words."""
        endings = Counter()
        
        for line in lines:
            words = re.findall(r'[\wӣӯ]+', line)
            if not words:
                continue
            
            for length in range(1, min(4, len(words) + 1)):
                ending = " ".join(words[-length:])
                endings[ending] += 1
        
        # A radif is by definition a repetition across verses: require at
        # least 2 occurrences regardless of poem length (fixes false
        # positives on free verse, where a single line ending was reported
        # as a radif).
        min_count = max(2, int(len(lines) * self.min_radif_frequency))
        significant = {e: c for e, c in endings.items() if c >= min_count}
        
        return significant
    
    def _select_best_radif(self, candidates: Dict[str, int], lines: List[str]) -> str:
        """Select the best Radīf from candidates."""
        if not candidates:
            return ""
        
        best_radif = ""
        best_score = 0
        
        for radif, count in candidates.items():
            word_count = len(radif.split())
            score = count / len(lines) * (1.5 / word_count)
            
            line_end_count = sum(1 for line in lines if line.strip().endswith(radif))
            line_end_bonus = line_end_count / len(lines)
            score += line_end_bonus
            
            if score > best_score:
                best_score = score
                best_radif = radif
        
        if best_score < 0.4:
            return ""
        
        return best_radif
    
    def _remove_radif(self, lines: List[str], radif: str) -> List[str]:
        """Remove Radīf from lines for correct meter analysis."""
        cleaned_lines = []
        
        for line in lines:
            line = line.strip()
            if line.endswith(radif):
                cleaned = line[:-len(radif)].rstrip(' ,;:!?.—–-')
                cleaned_lines.append(cleaned.strip())
            else:
                cleaned_lines.append(line)
        
        return cleaned_lines
    
    def _calculate_radif_frequency(self, lines: List[str], radif: str) -> float:
        """Calculate frequency of Radīf appearance."""
        if not lines or not radif:
            return 0.0
        count = sum(1 for line in lines if line.strip().endswith(radif))
        return count / len(lines)


# =============================================================================
# STRUCTURAL ANALYZER
# =============================================================================

class StructuralAnalyzer:
    """Enhanced structural analyzer with Radīf detection and Haiku support"""

    def __init__(self, config: Optional[AnalysisConfig] = None):
        self.config = config or AnalysisConfig()
        self.aruz_analyzer = AruzMeterAnalyzer(self.config)
        self.rhyme_analyzer = RhymeRadifAnalyzer()
        self.phonetics = PersianTajikPhonetics()
        self.radif_detector = EnhancedRadifDetector()  # Radīf detection
        self.haiku_detector = HaikuDetector()  # Haiku detection
        self.stanza_detector = StanzaFormDetector()

    def analyze(self, poem_content: str) -> StructuralAnalysis:
        """Comprehensive structural analysis with proper preprocessing"""
    
        # PREPROCESSING STEP - Remove titles/dedications
        preprocessor = PoemPreprocessor()
        processed_data = preprocessor.extract_poem_body(poem_content)
        
        # Use cleaned content for analysis (titles removed)
        clean_content = processed_data['body']
        clean_lines = processed_data['body_lines']
        
        if not clean_lines:
            raise ValueError("No valid lines found in poem after preprocessing")

        # Detect Radīf pattern first for accurate meter analysis
        radif_analysis = self.radif_detector.detect_radif_pattern(clean_content)
        has_radif = radif_analysis.radif_present
        radif_text = radif_analysis.radif_text

        # Use cleaned lines (Radif removed) for meter analysis to avoid interference
        analysis_lines = radif_analysis.cleaned_lines

        if has_radif:
            logger.info(f"Radīf detected: '{radif_text}' - using cleaned lines for meter analysis")

        line_analyses = []
        syllable_counts = []
        syllable_patterns = []
        rhyme_analyses = []

        # Analyze with original lines for rhyme, cleaned lines for meter
        for i, original_line in enumerate(clean_lines):
            # Use cleaned line for meter analysis
            meter_line = analysis_lines[i] if i < len(analysis_lines) else original_line
            
            # Analyze meter with cleaned line
            aruz_analysis_for_line = self.aruz_analyzer.analyze_meter(meter_line)
            
            # Extract syllables and count from the AruzAnalysis result
            syllables = aruz_analysis_for_line.line_scansion
            syllable_count_for_line = len(syllables)
            
            syllable_patterns.append(syllables)
            syllable_counts.append(syllable_count_for_line)

            # Use original line for rhyme analysis
            rhyme = self.rhyme_analyzer.analyze_rhyme(original_line)
            # Update radif field if global Radīf detected
            if has_radif and not rhyme.radif:
                rhyme = RhymeAnalysis(
                    qafiya=rhyme.qafiya,
                    radif=radif_text,
                    rawi=rhyme.rawi,
                    phonetic_rhyme=rhyme.phonetic_rhyme,
                    rhyme_type=rhyme.rhyme_type,
                    rhyme_position=rhyme.rhyme_position,
                    confidence=rhyme.confidence
                )
            rhyme_analyses.append(rhyme)

            line_analyses.append({
                'syllables': syllables,
                'rhyme': rhyme,
                'aruz': aruz_analysis_for_line,
                'has_radif': has_radif,
                'radif_text': radif_text if has_radif else ''
            })

        # Use correct line count from preprocessing
        actual_line_count = len(clean_lines)
        rhyme_pattern = self._generate_rhyme_pattern(rhyme_analyses, radif_text if has_radif else '')
        
        # Detect stanza structure
        stanza_structure, _ = self.stanza_detector.detect_form(
            clean_lines, rhyme_pattern, radif_analysis, syllable_counts
        )
        
        avg_syllables = sum(syllable_counts) / len(syllable_counts) if syllable_counts else 0
        syllable_std_dev = statistics.stdev(syllable_counts) if len(syllable_counts) > 1 else 0
        prosodic_consistency = self._calculate_prosodic_consistency(line_analyses)

        meters = [la['aruz'] for la in line_analyses]
        overall_aruz = self._determine_overall_meter(meters)

        return StructuralAnalysis(
            lines=actual_line_count,  # CORRECTED LINE COUNT
            syllables_per_line=syllable_counts,
            syllable_patterns=syllable_patterns,
            aruz_analysis=overall_aruz,
            rhyme_scheme=rhyme_analyses,
            rhyme_pattern=rhyme_pattern,
            radif_analysis=radif_analysis,
            stanza_structure=stanza_structure,
            avg_syllables=round(avg_syllables, 2),
            syllable_std_dev=syllable_std_dev,
            prosodic_consistency=prosodic_consistency,
            meter_confidence=overall_aruz.confidence
        )


    def _generate_rhyme_pattern(self, rhyme_analyses: List[RhymeAnalysis], global_radif: str = '') -> str:
        """Generate rhyme scheme pattern with Radīf awareness"""
        if not rhyme_analyses:
            return ""

        pattern = []
        rhyme_groups = {}
        next_label = 'A'

        for rhyme in rhyme_analyses:
            # For poems with global Radīf, focus on Qāfiyeh similarity
            if global_radif:
                rhyme_key = (rhyme.qafiya, global_radif, rhyme.phonetic_rhyme)
            else:
                rhyme_key = (rhyme.qafiya, rhyme.radif, rhyme.phonetic_rhyme)
            matched = False

            for prev_key, label in rhyme_groups.items():
                prev_rhyme = RhymeAnalysis(
                    qafiya=prev_key[0],
                    radif=prev_key[1],
                    rawi="", # Not needed for comparison
                    phonetic_rhyme=prev_key[2],
                    rhyme_type="",
                    rhyme_position="end",
                    confidence=0.0
                )
                similarity = self._calculate_rhyme_similarity(rhyme, prev_rhyme)
                # Higher threshold for Radīf poems (they often rhyme consistently)
                threshold = 0.6 if global_radif else 0.7
                if similarity > threshold:
                    pattern.append(label)
                    matched = True
                    break

            if not matched:
                pattern.append(next_label)
                rhyme_groups[rhyme_key] = next_label
                next_label = chr(ord(next_label) + 1)

        return ''.join(pattern)

    def _calculate_rhyme_similarity(self, rhyme1: RhymeAnalysis, rhyme2: RhymeAnalysis) -> float:
        """Calculate phonetic similarity between two rhymes"""
        if not rhyme1.phonetic_rhyme or not rhyme2.phonetic_rhyme:
            return 0.0

        phone1, phone2 = rhyme1.phonetic_rhyme, rhyme2.phonetic_rhyme
        matches = sum(1 for a, b in zip(phone1, phone2) if a == b)
        max_len = max(len(phone1), len(phone2))

        if max_len == 0:
            return 1.0

        radif_bonus = 0.2 if rhyme1.radif == rhyme2.radif and rhyme1.radif else 0.0
        return min(1.0, (matches / max_len) + radif_bonus)

    def _calculate_prosodic_consistency(self, line_analyses: List[Dict]) -> float:
        """Calculate prosodic consistency"""
        if not line_analyses:
            return 0.0

        meters = [la['aruz'].identified_meter for la in line_analyses]
        unique_meters = set(meters)
        meter_consistency = 1.0 / len(unique_meters) if unique_meters else 0.0

        syllable_counts = [len(la['syllables']) for la in line_analyses]
        if syllable_counts:
            avg = sum(syllable_counts) / len(syllable_counts)
            variance = sum((c - avg) ** 2 for c in syllable_counts) / len(syllable_counts)
            syllable_consistency = 1.0 / (1.0 + variance / max(avg, 1))
        else:
            syllable_consistency = 0.0

        return (meter_consistency + syllable_consistency) / 2

    def _determine_overall_meter(self, meters: List[AruzAnalysis]) -> AruzAnalysis:
        """Determine overall meter"""
        if not meters:
            return AruzAnalysis(
                identified_meter="unknown",
                meter_arabic="غير معروف",
                pattern_match="",
                full_pattern="",
                confidence=MeterConfidence.NONE,
                pattern_accuracy=0.0,
                variations_detected=[],
                line_scansion=[],
                caesura_positions=[],
                feet_breakdown=[],
                alternative_meters=[]
            )

        meter_counts = Counter(m.identified_meter for m in meters)
        most_common = meter_counts.most_common(1)[0][0]

        return max(
            (m for m in meters if m.identified_meter == most_common),
            key=lambda m: m.pattern_accuracy
        )


# =============================================================================
# CONTENT ANALYZER (Lexicon, Neologisms, Themes)
# =============================================================================

class ContentAnalyzer:
    """Enhanced content analyzer with lexicon support and neologism detection"""

    def __init__(self, config: Optional[AnalysisConfig] = None):
        self.config = config or AnalysisConfig()
        self.lexicon = self._load_lexicon()
        
        # Define archaic words (classical Persian/Tajik)
        self.archaisms = {
            'зи', 'ки', 'чу', 'зеро', 'балки', 'андар', 'бар', 'аз-ан-ки',
            'ҳамана', 'бадин', 'бад-он', 'з-он', 'к-он', 'чунон', 'чунин',
            'инак', 'онак', 'биҳишт', 'дӯзах', 'фалак', 'қазо', 'қадар'
        }
        
        # Modern/neologistic patterns
        self.modern_patterns = [
            r'\w+изм$',  # -ism words
            r'\w+ция$',  # -tion words (Russian influence)
            r'\w+логӣ$',  # -logy words
        ]
        
        # Persian/Arabic loanword patterns
        self.arabic_patterns = [
            r'^ал-',  # Arabic definite article
            r'ият$',  # -iyyat suffix
            r'ӣ$',   # Persian/Arabic -ī suffix
        ]
        
        logger.info(f"ContentAnalyzer initialized with {len(self.lexicon)} lexicon entries")

    def _load_lexicon(self) -> Set[str]:
        """Load lexicon from configured file path"""
        try:
            lexicon_path = Path(self.config.lexicon_path)
            if lexicon_path.exists():
                with open(lexicon_path, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                    # Handle both list and dict formats
                    if isinstance(data, list):
                        return set(word.lower() for word in data)
                    elif isinstance(data, dict):
                        return set(word.lower() for word in data.keys())
                    else:
                        logger.warning(f"Unexpected lexicon format in {lexicon_path}. Expected list or dict.")
                        return set()
            else:
                logger.warning(f"Lexicon file not found at: {lexicon_path}")
        except json.JSONDecodeError as e:
            logger.error(f"Error decoding JSON from lexicon file {self.config.lexicon_path}: {e}")
        except Exception as e:
            logger.error(f"Error loading lexicon from {self.config.lexicon_path}: {e}")
        return set() # Return empty set on any failure

    def _calculate_mtld(self, words: List[str], threshold: float = 0.72) -> float:
        """Calculate MTLD (Measure of Textual Lexical Diversity) - McCarthy & Jarvis 2010"""
        if len(words) < 10:
            return 0.0
        
        def mtld_forward(word_list):
            factors = 0.0
            start = 0
            for i in range(1, len(word_list) + 1):
                segment = word_list[start:i]
                ttr = len(set(segment)) / len(segment)
                if ttr <= threshold:
                    factors += 1
                    start = i
            if start < len(word_list):
                remaining = word_list[start:]
                ttr = len(set(remaining)) / len(remaining)
                factors += (1 - ttr) / (1 - threshold) if threshold < 1 else 0
            return factors
        
        forward = mtld_forward(words)
        backward = mtld_forward(words[::-1])
        avg_factors = (forward + backward) / 2
        
        return len(words) / avg_factors if avg_factors > 0 else float(len(words))

    def analyze(self, poem_content: str) -> ContentAnalysis:
        """Comprehensive content analysis"""
        # Extract words
        words = re.findall(r'[\wӣӯ]+', poem_content.lower())
        word_freq = Counter(words)
        
        # Find neologisms and archaisms
        neologisms = self._find_neologisms(words)
        archaisms = self._find_archaisms(words)
        
        # Analyze themes
        theme_distribution = self._analyze_themes(words)
        
        # Calculate lexical diversity using MTLD (McCarthy & Jarvis 2010)
        total_words = len(words)
        unique_words = len(set(words))
        lexical_diversity = self._calculate_mtld(words)
        
        # Determine stylistic register
        stylistic_register = self._determine_register(words, archaisms, neologisms)
        
        # Persian/Arabic ratio
        persian_arabic_ratio = self._calculate_persian_arabic_ratio(words)
        
        # Primary theme
        # Only report a primary theme when at least one theme keyword was
        # actually found; otherwise "Love" (the first dict key) was returned
        # for every poem with zero theme hits.
        if theme_distribution and max(theme_distribution.values()) > 0:
            primary_theme = max(theme_distribution.items(), key=lambda x: x[1])[0]
        else:
            primary_theme = "Undetermined"
        
        return ContentAnalysis(
            word_frequencies=word_freq.most_common(20),
            neologisms=neologisms[:self.config.max_neologisms],
            archaisms=list(archaisms),
            theme_distribution=theme_distribution,
            primary_theme=primary_theme,
            lexical_diversity=round(lexical_diversity, 3),
            stylistic_register=stylistic_register,
            total_words=total_words,
            unique_words=unique_words,
            persian_arabic_ratio=persian_arabic_ratio
        )

    # Tajik nominal suffixes, longest first: plural+izofat combinations,
    # plural, possessive enclitics, object marker, izofat, indefinite -e.
    _STRIP_SUFFIXES = (
        'ҳоямон', 'ҳоятон', 'ҳояшон',
        'ҳоям', 'ҳоят', 'ҳояш', 'ҳои', 'ҳоро', 'ҳое', 'ҳо',
        'амон', 'атон', 'ашон',
        'ям', 'ят', 'яш', 'ам', 'ат', 'аш',
        'еро', 'ро', 'ии', 'и', 'е',
    )

    def _lexicon_forms(self, word: str) -> Set[str]:
        """Generate candidate base forms by iteratively stripping nominal
        suffixes (izofat, plural, possessives, object marker). Surface forms
        like 'ғамгини' (ғамгин + izofat) or 'гумкардаҳои' (гумкарда + plural
        + izofat) previously produced false-positive neologisms because only
        the inflected form was checked against the lexicon."""
        forms = {word}
        frontier = {word}
        for _ in range(3):  # at most 3 stacked suffixes
            next_frontier = set()
            for w in frontier:
                for suf in self._STRIP_SUFFIXES:
                    if w.endswith(suf) and len(w) - len(suf) >= 3:
                        stem = w[: -len(suf)]
                        if stem not in forms:
                            forms.add(stem)
                            next_frontier.add(stem)
            if not next_frontier:
                break
            frontier = next_frontier
        return forms

    def _find_neologisms(self, words: List[str]) -> List[str]:
        """Find neologisms (words whose base form is not in the lexicon)"""
        if not self.lexicon:
            logger.warning("No lexicon loaded - neologism detection disabled")
            return []
        
        neologisms = []
        for word in set(words):
            if word.isdigit() or len(word) <= 2:
                continue
            # Known if any candidate base form is in lexicon or archaisms
            forms = self._lexicon_forms(word)
            if any(f in self.lexicon or f in self.archaisms for f in forms):
                continue
            neologisms.append(word)
        
        return sorted(neologisms)

    def _find_archaisms(self, words: List[str]) -> Set[str]:
        """Find archaic words"""
        return set(word for word in words if word in self.archaisms)

    def _analyze_themes(self, words: List[str]) -> Dict[str, int]:
        """Analyze thematic distribution"""
        theme_counts = {}
        
        for theme, keywords in self.config.themes.items():
            count = sum(1 for word in words if word in keywords)
            theme_counts[theme] = count
        
        return theme_counts

    def _determine_register(self, words: List[str], archaisms: Set[str],
                           neologisms: List[str]) -> str:
        """Determine stylistic register"""
        total_words = len(words)
        
        if not total_words:
            return "unknown"
        
        archaic_ratio = len(archaisms) / total_words
        neologism_ratio = len(neologisms) / total_words
        
        if archaic_ratio > 0.05:
            return "classical"
        elif neologism_ratio > 0.05:
            return "modern"
        elif archaic_ratio > 0.02 and neologism_ratio < 0.02:
            return "neo-classical"
        else:
            return "contemporary"

    def _calculate_persian_arabic_ratio(self, words: List[str]) -> float:
        """Calculate ratio of Persian/Arabic loanwords"""
        arabic_count = 0
        
        for word in words:
            for pattern in self.arabic_patterns:
                if re.search(pattern, word):
                    arabic_count += 1
                    break
        
        return arabic_count / len(words) if words else 0.0

    def build_vocabulary_from_corpus(self, corpus_path: str) -> Dict[str, int]:
        """Build vocabulary dictionary from corpus file"""
        vocabulary = Counter()
        
        try:
            corpus_file = Path(corpus_path)
            if not corpus_file.exists():
                logger.error(f"Corpus file not found: {corpus_path}")
                return {}
            
            logger.info(f"Building vocabulary from {corpus_path}...")
            
            with open(corpus_file, 'r', encoding='utf-8') as f:
                for line in f:
                    words = re.findall(r'[\wӣӯ]+', line.lower())
                    vocabulary.update(words)
            
            logger.info(f"Built vocabulary with {len(vocabulary)} unique words")
            return dict(vocabulary)
            
        except Exception as e:
            logger.error(f"Error building vocabulary: {e}")
            return {}

    def save_vocabulary(self, vocabulary: Dict[str, int], output_path: str):
        """Save vocabulary to JSON file"""
        try:
            with open(output_path, 'w', encoding='utf-8') as f:
                json.dump(vocabulary, f, ensure_ascii=False, indent=2)
            logger.info(f"Vocabulary saved to {output_path}")
        except Exception as e:
            logger.error(f"Error saving vocabulary: {e}")


# =============================================================================
# LITERARY ASSESSOR
# =============================================================================

class LiteraryAssessor:
    """
    EXPERIMENTAL HEURISTIC — not a validated instrument.

    The scores produced here (german_perspective, persian_tradition,
    tajik_elements, modernist_features, overall_quality) are rule-of-thumb
    composites without any empirical grounding or inter-rater validation.
    They MUST NOT be cited as measurements in academic work. Kept only for
    exploratory sorting in the UI; treat as ordinal hints at best.
    """
    
    @staticmethod
    def assess(
        structural: StructuralAnalysis,
        content: ContentAnalysis
    ) -> LiteraryAssessment:
        """
        Provide multi-perspective literary assessment.
        """
        # Classical conformity based on meter accuracy and form
        classical_conformity = LiteraryAssessor._assess_classical_conformity(structural)
        
        # German literary perspective (focus on form, structure)
        german_perspective = LiteraryAssessor._assess_german_perspective(structural)
        
        # Persian tradition perspective (focus on classical elements)
        persian_tradition = LiteraryAssessor._assess_persian_tradition(structural, content)
        
        # Tajik elements (regional/linguistic features)
        tajik_elements = LiteraryAssessor._assess_tajik_elements(content)
        
        # Modernist features
        modernist_features = LiteraryAssessor._assess_modernist_features(structural, content)
        
        # Overall quality (weighted average)
        overall_quality = (
            classical_conformity * 0.25 +
            (german_perspective / 10) * 0.15 +
            (persian_tradition / 10) * 0.25 +
            (tajik_elements / 10) * 0.20 +
            (modernist_features / 10) * 0.15
        )
        
        return LiteraryAssessment(
            classical_conformity=classical_conformity,
            german_perspective=german_perspective,
            persian_tradition=persian_tradition,
            tajik_elements=tajik_elements,
            modernist_features=modernist_features,
            overall_quality=overall_quality
        )
    
    @staticmethod
    def _assess_classical_conformity(structural: StructuralAnalysis) -> float:
        """Assess conformity to classical prosodic rules"""
        score = 0.0
        
        # Meter identification confidence
        if structural.aruz_analysis.confidence == MeterConfidence.HIGH:
            score += 0.4
        elif structural.aruz_analysis.confidence == MeterConfidence.MEDIUM:
            score += 0.25
        elif structural.aruz_analysis.confidence == MeterConfidence.LOW:
            score += 0.1
        
        # Pattern accuracy
        score += structural.aruz_analysis.pattern_accuracy * 0.3
        
        # Prosodic consistency
        score += structural.prosodic_consistency * 0.2
        
        # Stanza form (classical forms get higher score)
        classical_forms = {
            StanzaForm.GHAZAL, StanzaForm.QASIDA, StanzaForm.RUBAI,
            StanzaForm.GHAZAL_WITH_RADIF, StanzaForm.QASIDA_WITH_RADIF,
            StanzaForm.MASNAVI
        }
        if structural.stanza_structure in classical_forms:
            score += 0.1
        
        return min(score, 1.0)
    
    @staticmethod
    def _assess_german_perspective(structural: StructuralAnalysis) -> int:
        """
        German literary perspective (1-10).
        Focuses on formal rigor, structural consistency.
        """
        score = 5  # Base score
        
        # Meter identification
        if structural.aruz_analysis.identified_meter != "unknown":
            score += 2
        
        # Prosodic consistency
        if structural.prosodic_consistency > 0.8:
            score += 1
        elif structural.prosodic_consistency < 0.5:
            score -= 1
        
        # Syllable regularity
        if structural.syllable_std_dev < 2:
            score += 1
        elif structural.syllable_std_dev > 4:
            score -= 1
        
        # Clear structure
        if structural.stanza_structure != StanzaForm.UNKNOWN:
            score += 1
        
        return max(1, min(10, score))
    
    @staticmethod
    def _assess_persian_tradition(
        structural: StructuralAnalysis,
        content: ContentAnalysis
    ) -> int:
        """
        Persian tradition perspective (1-10).
        Focuses on classical elements, literary conventions.
        """
        score = 5  # Base score
        
        # Classical meter
        if structural.aruz_analysis.identified_meter != "unknown":
            score += 2
        
        # Radīf presence (valued in Persian poetry)
        if structural.radif_analysis.radif_present:
            score += 1
        
        # Classical themes
        classical_themes = {"Love", "Mysticism", "Philosophy", "Nature"}
        if content.primary_theme in classical_themes:
            score += 1
        
        # Archaic vocabulary
        if content.archaisms:
            score += 1
        
        # Persian/Arabic literary vocabulary
        if content.persian_arabic_ratio > 0.1:
            score += 1
        
        # Classical form
        classical_forms = {StanzaForm.GHAZAL, StanzaForm.QASIDA, StanzaForm.RUBAI}
        if structural.stanza_structure in classical_forms:
            score += 1
        
        return max(1, min(10, score))
    
    @staticmethod
    def _assess_tajik_elements(content: ContentAnalysis) -> int:
        """
        Tajik elements perspective (1-10).
        Focuses on regional and linguistic features.
        """
        score = 5  # Base score
        
        # Homeland theme
        if content.theme_distribution.get("Homeland", 0) > 0:
            score += 2
        
        # Unity theme (important in Tajik poetry)
        if content.theme_distribution.get("Unity", 0) > 0:
            score += 1
        
        # Contemporary register suggests modern Tajik style
        if content.stylistic_register == "contemporary":
            score += 1
        elif content.stylistic_register == "traditional":
            score += 1
        
        # Lexical diversity
        if content.lexical_diversity > 0.5:
            score += 1
        
        return max(1, min(10, score))
    
    @staticmethod
    def _assess_modernist_features(
        structural: StructuralAnalysis,
        content: ContentAnalysis
    ) -> int:
        """
        Modernist features assessment (1-10).
        """
        score = 5  # Base score
        
        # Free verse
        if structural.stanza_structure == StanzaForm.FREE_VERSE:
            score += 3
        
        # Neologisms
        if content.neologisms:
            score += 1
        
        # Modern register
        if content.stylistic_register == "modern":
            score += 2
        
        # High syllable variation (enjambement, etc.)
        if structural.syllable_std_dev > 3:
            score += 1
        
        # Absence of traditional forms
        traditional_forms = {
            StanzaForm.GHAZAL, StanzaForm.QASIDA, StanzaForm.RUBAI,
            StanzaForm.MASNAVI
        }
        if structural.stanza_structure not in traditional_forms:
            score += 1
        
        return max(1, min(10, score))


# =============================================================================
# ENHANCED TAJIK POEM ANALYZER
# =============================================================================

class EnhancedTajikPoemAnalyzer:
    """
    Enhanced analyzer with free verse detection and modern metrics
    Works in parallel with classical TajikPoemAnalyzer
    """

    def __init__(self, config: Optional[AnalysisConfig] = None, enable_corpus: bool = True):
        self.config = config or AnalysisConfig()
        self.structural_analyzer = StructuralAnalyzer(self.config)
        self.content_analyzer = ContentAnalyzer(self.config)
        self.modern_analyzer = ModernVerseAnalyzer()
        self.free_verse_classifier = FreeVerseClassifier()
        self.excel_reporter = ExcelReporter()
        self.enable_corpus = enable_corpus

        logger.info("EnhancedTajikPoemAnalyzer initialized with free verse detection")

    def analyze_poem(self, poem_content: str) -> ComprehensiveAnalysis:
        """Perform enhanced analysis of a single poem with proper preprocessing"""
        
        # Input validation
        if not isinstance(poem_content, str):
            raise TypeError("poem_content must be a string")
        
        if not poem_content or len(poem_content.strip()) < self.config.min_poem_length:
            raise ValueError(f"Poem content is too short. Minimum length is {self.config.min_poem_length} characters.")

        # PREPROCESSING STEP - Remove titles/dedications BEFORE analysis
        preprocessor = PoemPreprocessor()
        processed_data = preprocessor.extract_poem_body(poem_content)
        clean_content = processed_data['body']  # This excludes titles/dedications
        
        # Perform basic analysis on cleaned content
        structural = self.structural_analyzer.analyze(clean_content)
        content = self.content_analyzer.analyze(clean_content)
        literary = LiteraryAssessor.assess(structural, content)

        # Analyze modern metrics
        modern_metrics = self.modern_analyzer.analyze(
            clean_content,  # Use cleaned content
            structural.syllables_per_line
        )

        # Classify free verse
        is_free_verse = self.free_verse_classifier.is_free_verse(
            structural,
            modern_metrics
        )

        # Enhance structural analysis
        enhanced_structural = self._enhance_structural_analysis(
            structural,
            modern_metrics,
            is_free_verse
        )

        # Create enhanced analysis
        analysis = ComprehensiveAnalysis(
            structural=enhanced_structural,
            content=content,
            literary=literary,
            quality_metrics={},
            corpus_ready=self.enable_corpus,
            contribution_id=None
        )

        # Enhanced quality validation
        validation = self._enhance_quality_metrics(
            analysis,
            modern_metrics
        )
        analysis.quality_metrics = validation

        return analysis

    def _enhance_structural_analysis(self, structural: StructuralAnalysis,
                                   modern_metrics: ModernVerseMetrics,
                                   is_free_verse: bool) -> EnhancedStructuralAnalysis:
        """Enhance structural analysis for free verse"""
        
        # Simplify rhyme pattern for free verse
        rhyme_pattern = structural.rhyme_pattern
        if is_free_verse and len(rhyme_pattern) > 20:
            unique_rhymes = len(set(rhyme_pattern))
            rhyme_pattern = f"free_rhyme_{unique_rhymes}unique"
        
        # Adjust meter for free verse and special forms
        identified_meter = structural.aruz_analysis.identified_meter
        meter_confidence = structural.meter_confidence
        
        # Check if it's a Haiku
        is_haiku = structural.stanza_structure == StanzaForm.HAIKU
        
        if is_haiku:
            # Haiku is a special non-ʿArūḍ form
            identified_meter = "haiku"
            meter_confidence = MeterConfidence.HIGH  # Haiku detection is reliable
            is_free_verse = False  # Haiku is not free verse, it has a fixed form
        elif is_free_verse and identified_meter == "ṭawīl":
            # ṭawīl is often false-positive for free verse
            identified_meter = "free_verse"
            meter_confidence = MeterConfidence.LOW
        elif is_free_verse:
            identified_meter = "free_verse"
        
        # Extract modern features
        modern_features = {
            "enjambement_density": modern_metrics.enjambement_ratio,
            "line_variation": modern_metrics.line_length_variation,
            "prose_tendency": modern_metrics.prose_poetry_score,
            "visual_complexity": modern_metrics.visual_structure_score,
            "syntactic_parallelism": modern_metrics.syntactic_parallelism,
            "lexical_repetition": modern_metrics.lexical_repetition_score
        }
        
        return EnhancedStructuralAnalysis(
            lines=structural.lines,
            syllables_per_line=structural.syllables_per_line,
            syllable_patterns=structural.syllable_patterns,
            aruz_analysis=structural.aruz_analysis,
            rhyme_scheme=structural.rhyme_scheme,
            rhyme_pattern=rhyme_pattern,
            radif_analysis=structural.radif_analysis,
            stanza_structure=structural.stanza_structure,
            avg_syllables=structural.avg_syllables,
            syllable_std_dev=structural.syllable_std_dev,
            prosodic_consistency=structural.prosodic_consistency,
            meter_confidence=meter_confidence,
            modern_metrics=modern_metrics,
            is_free_verse=is_free_verse,
            free_verse_confidence=self._calculate_free_verse_confidence(
                structural, modern_metrics),
            modern_features=modern_features
        )

    def _calculate_free_verse_confidence(self, structural: StructuralAnalysis,
                                       modern_metrics: ModernVerseMetrics) -> float:
        """Calculate confidence for free verse classification"""
        indicators = [
            (structural.prosodic_consistency < 0.4, 0.8),
            (modern_metrics.enjambement_ratio > 0.3, 0.6),
            (modern_metrics.line_length_variation > 0.5, 0.7),
            (modern_metrics.prose_poetry_score > 0.6, 0.5),
            (structural.meter_confidence.value in ['low', 'none'], 0.9)
        ]
        
        confidence = sum(weight for condition, weight in indicators if condition)
        return min(1.0, confidence / 2.5)  # Normalize

    def _enhance_quality_metrics(self, analysis: ComprehensiveAnalysis,
                               modern_metrics: ModernVerseMetrics) -> Dict[str, Any]:
        """Enhanced quality validation"""
        warnings = []
        recommendations = []
        quality_score = 1.0

        if analysis.structural.meter_confidence == MeterConfidence.NONE:
            warnings.append("No reliable meter detected")
            recommendations.append("Manual prosodic verification recommended")
            quality_score *= 0.7

        if analysis.structural.prosodic_consistency < 0.5:
            warnings.append("Low prosodic consistency")
            recommendations.append("Check for textual corruption or free verse intention")
            quality_score *= 0.8

        if analysis.structural.lines < 2:
            warnings.append("Very short poem")
            recommendations.append("Statistical analysis not reliable for single lines")
            quality_score *= 0.5

        # Free verse specific assessments
        free_verse_analysis = {}
        if analysis.structural.is_free_verse:
            free_verse_assessment = self._assess_free_verse_quality(
                analysis.structural, modern_metrics)
            
            # Remove conflicting warnings
            if "Low prosodic consistency" in warnings:
                warnings.remove("Low prosodic consistency")
                warnings.append("Free verse detected - prosodic analysis limited")
            
            free_verse_analysis = {
                "confidence": analysis.structural.free_verse_confidence,
                "enjambement_score": modern_metrics.enjambement_ratio,
                "prose_poetry_score": modern_metrics.prose_poetry_score,
                "line_variation_score": modern_metrics.line_length_variation,
                "assessment": free_verse_assessment
            }

        reliability = "high" if quality_score > 0.8 else "medium" if quality_score > 0.6 else "low"

        result = {
            'quality_score': round(quality_score, 2),
            'reliability': reliability,
            'warnings': warnings,
            'recommendations': recommendations,
            'timestamp': datetime.now().isoformat()
        }
        
        if analysis.structural.is_free_verse:
            result['free_verse_analysis'] = free_verse_analysis
            
        return result

    def _assess_free_verse_quality(self, structural: EnhancedStructuralAnalysis,
                                 modern_metrics: ModernVerseMetrics) -> str:
        """Assess quality of free verse"""
        scores = []
        
        # Enjambement assessment
        if 0.2 <= modern_metrics.enjambement_ratio <= 0.6:
            scores.append(1.0)
        else:
            scores.append(0.5)
            
        # Line variation assessment
        if 0.3 <= modern_metrics.line_length_variation <= 0.8:
            scores.append(1.0)
        else:
            scores.append(0.5)
            
        # Visual structure
        if modern_metrics.visual_structure_score > 0.2:
            scores.append(0.8)
            
        avg_score = sum(scores) / len(scores) if scores else 0
        
        if avg_score > 0.8:
            return "excellent_free_verse"
        elif avg_score > 0.6:
            return "good_free_verse"
        elif avg_score > 0.4:
            return "experimental_free_verse"
        else:
            return "irregular_free_verse"

    def analyze_text(self, text: str) -> List[Dict[str, Any]]:
        """Analyze text containing multiple poems"""
        poems = self._split_poems(text)
        logger.info(f"Found {len(poems)} poems to analyze")

        results = []
        for poem in poems:
            try:
                analysis = self.analyze_poem(poem.content)

                results.append({
                    "poem_id": poem.poem_id,
                    "title": poem.title,
                    "content": poem.content,
                    "analysis": analysis,
                    "validation": analysis.quality_metrics
                })
            except Exception as e:
                logger.error(f"Error analyzing poem {poem.poem_id}: {e}")
                continue

        return results

    def analyze_file(self, filepath: str, output_file: Optional[str] = None) -> List[Dict[str, Any]]:
        """Analyze poems from a file"""
        try:
            file_path = Path(filepath)
            if not file_path.exists():
                raise FileNotFoundError(f"File not found: {filepath}")

            with open(file_path, 'r', encoding='utf-8') as f:
                text = f.read()

            results = self.analyze_text(text)

            if output_file is None:
                output_file = f"{file_path.stem}_enhanced_analysis.xlsx"

            self.excel_reporter.create_report(results, output_file)

            logger.info(f"Enhanced analysis complete. Results saved to {output_file}")
            return results

        except Exception as e:
            logger.error(f"Error analyzing file {filepath}: {e}")
            raise

    def _split_poems(self, text: str) -> List[PoemData]:
        """Split text into individual poems"""
        text = unicodedata.normalize('NFC', text)

        separators = [
            r'\*{5,}', r'-{5,}', r'={5,}', r'_{5,}',
            r'#+\s*\d+\s*#+', r'\n\s*\n\s*\n+'
        ]

        pattern = '|'.join(separators)
        blocks = re.split(pattern, text)

        poems = []
        for i, block in enumerate(blocks, 1):
            block = block.strip()
            if len(block) < self.config.min_poem_length:
                continue

            lines = block.split('\n')
            if lines:
                first_line = lines[0].strip()
                if (self.config.min_title_length <= len(first_line) <= self.config.max_title_length
                        and not first_line.endswith(('.', '!', '?'))):
                    title = first_line
                    content = '\n'.join(lines[1:]).strip()
                else:
                    title = f"Poem {i}"
                    content = block

                poems.append(PoemData(
                    title=title,
                    content=content,
                    poem_id=f"poem_{i:03d}"
                ))

        return poems


# =============================================================================
# EXCEL REPORTER
# =============================================================================

class ExcelReporter:
    """Excel report generation"""

    def __init__(self):
        self.header_font = Font(bold=True, color="FFFFFF")
        self.header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        self.border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        self.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    def create_report(self, results: List[Dict[str, Any]], filename: str = "tajik_poetry_analysis.xlsx"):
        """Create Excel report"""
        try:
            wb = openpyxl.Workbook()
            self._create_overview_sheet(wb, results)
            self._create_structural_sheet(wb, results)
            self._create_quality_sheet(wb, results)

            wb.save(filename)
            logger.info(f"Report saved as: {filename}")

        except Exception as e:
            logger.error(f"Error creating report: {e}")
            raise

    def _create_overview_sheet(self, wb: openpyxl.Workbook, results: List[Dict[str, Any]]):
        """Create overview sheet with correct line counts"""
        ws = wb.active
        ws.title = "Overview"

        headers = [
            "ID", "Title", "Lines", "Meter", "Confidence", "Rhyme Pattern",
            "Stanza Form", "Radīf", "Avg Syllables", "Quality Score"
        ]

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.border = self.border

        for row_num, result in enumerate(results, 2):
            analysis = result["analysis"]
            validation = result.get("validation", {})
            
            # Detect global Radīf
            structural = analysis.structural
            radif_values = [r.radif for r in structural.rhyme_scheme if r.radif]
            global_radif = radif_values[0] if radif_values and len(set(radif_values)) == 1 else "—"

            values = [
                result["poem_id"],
                result["title"],
                analysis.structural.lines,  # CORRECT LINE COUNT FROM PREPROCESSING
                analysis.structural.aruz_analysis.identified_meter,
                analysis.structural.meter_confidence.value,
                analysis.structural.rhyme_pattern,
                analysis.structural.stanza_structure.value,
                global_radif,
                analysis.structural.avg_syllables,
                validation.get("quality_score", "N/A")
            ]

            for col_num, value in enumerate(values, 1):
                cell = ws.cell(row=row_num, column=col_num, value=value)
                cell.border = self.border

    def _create_structural_sheet(self, wb: openpyxl.Workbook, results: List[Dict[str, Any]]):
        """Create structural analysis sheet with Radīf detection"""
        ws = wb.create_sheet(title="Structural Analysis")

        # NEW: Added "Global Radīf" and "Radīf Frequency" columns
        headers = [
            "Poem ID", "Line #", "Line Text", "Syllables", "Meter Pattern",
            "Qāfiyeh", "Radīf", "Rhyme Type", "Global Radīf", "Stanza Form"
        ]

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill

        row_num = 2
        for result in results:
            poem_id = result["poem_id"]
            content = result["content"]
            structural = result["analysis"].structural
            
            # Detect if poem has global Radīf
            # Check if all rhyme_scheme entries have same radif
            radif_values = [r.radif for r in structural.rhyme_scheme if r.radif]
            global_radif = radif_values[0] if radif_values and len(set(radif_values)) == 1 else ""
            stanza_form = structural.stanza_structure.value

            lines = [line.strip() for line in content.split('\n') if line.strip()]

            for line_idx, line in enumerate(lines):
                syllable_count = structural.syllables_per_line[line_idx] if line_idx < len(
                    structural.syllables_per_line) else 0

                if line_idx < len(structural.syllable_patterns):
                    pattern = ''.join([s.weight.value for s in structural.syllable_patterns[line_idx]])
                else:
                    pattern = ""

                if line_idx < len(structural.rhyme_scheme):
                    rhyme = structural.rhyme_scheme[line_idx]
                    qafiya = rhyme.qafiya
                    radif = rhyme.radif
                    rhyme_type = rhyme.rhyme_type
                else:
                    qafiya = radif = rhyme_type = ""

                values = [
                    poem_id, line_idx + 1, line, syllable_count, pattern,
                    qafiya, radif, rhyme_type, global_radif, stanza_form
                ]

                for col_num, value in enumerate(values, 1):
                    ws.cell(row=row_num, column=col_num, value=value)

                row_num += 1

    def _create_quality_sheet(self, wb: openpyxl.Workbook, results: List[Dict[str, Any]]):
        """Create quality metrics sheet"""
        ws = wb.create_sheet(title="Quality Metrics")

        headers = ["Poem ID", "Quality Score", "Reliability", "Warnings", "Recommendations"]

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill

        for row_num, result in enumerate(results, 2):
            validation = result.get("validation", {})

            values = [
                result["poem_id"],
                validation.get("quality_score", "N/A"),
                validation.get("reliability", "N/A"),
                "; ".join(validation.get("warnings", [])),
                "; ".join(validation.get("recommendations", []))
            ]

            for col_num, value in enumerate(values, 1):
                ws.cell(row=row_num, column=col_num, value=value)


# =============================================================================
# MAIN FUNCTION
# =============================================================================

def main():
    """Main function demonstrating the enhanced analyzer"""
    sample_text = """
Дар кӯҳсори ватан гулҳо мешукуфанд,
Дили ошиқ аз муҳаббат меларзад.
Баҳори нав ба замин таҷдид меорад,
Навиди хушҳолии мардум мерасад.

*****

Эй ватан, эй модари меҳрубон,
Дар оғӯши ту ёфтам ҷон.
Кӯҳҳои ту сари фалак расида,
Дарёҳои ту ҷовидон.
"""

    try:
        print("=== TAJIK POETRY ANALYZER DEMONSTRATION ===\n")
        
        # Test Enhanced Analyzer
        print("ENHANCED ANALYSIS MODE (with free verse detection and improved ʿArūḍ):")
        enhanced_analyzer = EnhancedTajikPoemAnalyzer()
        enhanced_results = enhanced_analyzer.analyze_text(sample_text)
        
        for result in enhanced_results:
            analysis = result["analysis"]
            print(f"--- {result['title']} (Enhanced) ---")
            print(f"  Lines: {analysis.structural.lines}")
            print(f"  Meter: {analysis.structural.aruz_analysis.identified_meter}")
            print(f"  Meter Arabic Name: {analysis.structural.aruz_analysis.meter_arabic}")
            print(f"  Confidence: {analysis.structural.meter_confidence.value}")
            print(f"  Pattern Accuracy: {analysis.structural.aruz_analysis.pattern_accuracy:.2f}")
            print(f"  Variations Detected: {', '.join(analysis.structural.aruz_analysis.variations_detected) or 'None'}")
            print(f"  Rhyme Pattern: {analysis.structural.rhyme_pattern}")
            print(f"  Stanza Form: {analysis.structural.stanza_structure.value}")
            print(f"  Radif Present: {analysis.structural.radif_analysis.radif_present}")
            if analysis.structural.radif_analysis.radif_present:
                print(f"  Radif Text: {analysis.structural.radif_analysis.radif_text}")
            print(f"  Avg Syllables: {analysis.structural.avg_syllables}")
            print(f"  Prosodic Consistency: {analysis.structural.prosodic_consistency:.2f}")
            print(f"  Free Verse: {analysis.structural.is_free_verse}")
            if analysis.structural.is_free_verse:
                print(f"  Free Verse Confidence: {analysis.structural.free_verse_confidence:.0%}")
            print(f"  Quality Score: {result['validation']['quality_score']}")
            print()

    except Exception as e:
        logger.error(f"Error in main: {e}")
        import traceback
        traceback.print_exc()


if __name__ == "__main__":
    main()